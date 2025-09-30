"""Excel to ODCS parser - Convert Excel files back to ODCS JSON/YAML format."""

from pathlib import Path
from typing import Any, Dict, Union, Optional
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from .models import ODCSDataContract
from .logging_config import get_logger
from .logging_utils import PerformanceTracker

logger = get_logger(__name__)
performance_tracker = PerformanceTracker()


class ExcelToODCSParser:
    """Parse Excel files and convert them back to ODCS format."""

    def __init__(self):
        """Initialize the Excel parser."""
        self.workbook = None
        self.worksheets = {}

    def parse_from_file(self, excel_path: Union[str, Path]) -> Dict[str, Any]:
        """Parse Excel file and return ODCS dictionary.

        Args:
            excel_path: Path to Excel file

        Returns:
            Dictionary containing ODCS data

        Raises:
            FileNotFoundError: If Excel file doesn't exist
            ValueError: If Excel file format is invalid
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        try:
            self.workbook = load_workbook(excel_path, read_only=True)
            self._load_all_worksheets()

            odcs_data = self._parse_odcs_data()

            logger.info(f"Successfully parsed Excel file: {excel_path}")
            return odcs_data

        except Exception as e:
            logger.error(f"Failed to parse Excel file {excel_path}: {e}")
            raise ValueError(f"Invalid Excel file format: {e}")
        finally:
            if self.workbook:
                self.workbook.close()

    def _load_all_worksheets(self) -> None:
        """Load all worksheets into pandas DataFrames."""
        self.worksheets = {}
        for sheet_name in self.workbook.sheetnames:
            try:
                # Read worksheet into DataFrame, treating first row as headers
                # Use openpyxl engine and get data directly from workbook
                sheet = self.workbook[sheet_name]
                data = []
                headers = []

                # Get headers from first row
                first_row = next(
                    sheet.iter_rows(min_row=1, max_row=1, values_only=True), None
                )
                if first_row:
                    headers = [cell for cell in first_row if cell is not None]

                # Get data from subsequent rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        row_data = list(row[: len(headers)] if headers else row)
                        data.append(row_data)

                # Create DataFrame
                if headers and data:
                    df = pd.DataFrame(data, columns=headers)
                elif headers:
                    df = pd.DataFrame(columns=headers)
                else:
                    df = pd.DataFrame()
                # Convert empty strings to None for better data handling
                df = df.where(df != "", None)
                self.worksheets[sheet_name] = df
                logger.debug(f"Loaded worksheet '{sheet_name}' with {len(df)} rows")
            except Exception as e:
                logger.warning(f"Could not load worksheet '{sheet_name}': {e}")

    def _parse_odcs_data(self) -> Dict[str, Any]:
        """Parse all worksheets and construct ODCS data structure."""
        odcs_data = {}

        # Parse each worksheet type
        # Parse all sections
        odcs_data.update(self._parse_basic_information())
        odcs_data.update(self._parse_tags())
        odcs_data.update(self._parse_description())
        odcs_data.update(self._parse_servers())
        odcs_data.update(self._parse_schema())

        # Parse enhanced schema data
        self._enhance_schema_with_properties(odcs_data)
        self._enhance_schema_with_logical_type_options(odcs_data)
        self._enhance_schema_with_quality_rules(odcs_data)

        odcs_data.update(self._parse_support())
        odcs_data.update(self._parse_pricing())
        odcs_data.update(self._parse_team())
        odcs_data.update(self._parse_roles())
        odcs_data.update(self._parse_sla_properties())
        odcs_data.update(self._parse_authoritative_definitions())
        odcs_data.update(self._parse_custom_properties())

        # Remove None values and empty structures
        odcs_data = self._clean_data(odcs_data)

        return odcs_data

    def _parse_basic_information(self) -> Dict[str, Any]:
        """Parse Basic Information worksheet."""
        sheet_name = "Basic Information"
        if sheet_name not in self.worksheets:
            logger.warning(f"Worksheet '{sheet_name}' not found")
            return {}

        df = self.worksheets[sheet_name]
        if df.empty or "Field" not in df.columns or "Value" not in df.columns:
            return {}

        data = {}
        for _, row in df.iterrows():
            field = row.get("Field")
            value = row.get("Value")

            if field and value is not None:
                # Handle special field mappings
                if field == "contractCreatedTs" and value:
                    try:
                        # Try to parse datetime if it's a string
                        if isinstance(value, str):
                            data[field] = datetime.fromisoformat(
                                value.replace("Z", "+00:00")
                            ).isoformat()
                        else:
                            data[field] = value
                    except (ValueError, AttributeError):
                        data[field] = str(value)
                else:
                    data[field] = self._convert_value(value)

        return data

    def _parse_tags(self) -> Dict[str, Any]:
        """Parse Tags worksheet."""
        sheet_name = "Tags"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty or "Tag" not in df.columns:
            return {}

        tags = []
        for _, row in df.iterrows():
            tag = row.get("Tag")
            if tag:
                tags.append(str(tag))

        return {"tags": tags} if tags else {}

    def _parse_description(self) -> Dict[str, Any]:
        """Parse Description worksheet."""
        sheet_name = "Description"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty or "Field" not in df.columns or "Value" not in df.columns:
            return {}

        description = {}
        for _, row in df.iterrows():
            field = row.get("Field")
            value = row.get("Value")

            if field and value is not None:
                description[field] = str(value)

        return {"description": description} if description else {}

    def _parse_servers(self) -> Dict[str, Any]:
        """Parse Servers worksheet."""
        sheet_name = "Servers"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        servers = []
        for _, row in df.iterrows():
            server = {}

            # Map common server fields
            field_mappings = {
                "Server": "server",
                "Type": "type",
                "Description": "description",
                "Environment": "environment",
                "Location": "location",
                "Host": "host",
                "Port": "port",
                "Database": "database",
                "Schema": "schema",
                "Project": "project",
                "Catalog": "catalog",
                "Format": "format",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    if odcs_field == "port":
                        try:
                            server[odcs_field] = int(value)
                        except (ValueError, TypeError):
                            server[odcs_field] = value
                    else:
                        server[odcs_field] = str(value)

            if server:  # Only add if we have some data
                servers.append(server)

        return {"servers": servers} if servers else {}

    def _parse_schema(self) -> Dict[str, Any]:
        """Parse Schema worksheet."""
        sheet_name = "Schema"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        schema_objects = []
        for _, row in df.iterrows():
            schema_obj = {}

            # Map schema object fields
            field_mappings = {
                "Object Name": "name",
                "Physical Name": "physicalName",
                "Logical Type": "logicalType",
                "Physical Type": "physicalType",
                "Description": "description",
                "Business Name": "businessName",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    schema_obj[odcs_field] = str(value)

            if schema_obj:  # Only add if we have some data
                schema_objects.append(schema_obj)

        return {"schema": schema_objects} if schema_objects else {}

    def _parse_support(self) -> Dict[str, Any]:
        """Parse Support worksheet."""
        sheet_name = "Support"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        support_items = []
        for _, row in df.iterrows():
            support_item = {}

            field_mappings = {
                "Channel": "channel",
                "URL": "url",
                "Description": "description",
                "Tool": "tool",
                "Scope": "scope",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    support_item[odcs_field] = str(value)

            if support_item:
                support_items.append(support_item)

        return {"support": support_items} if support_items else {}

    def _parse_pricing(self) -> Dict[str, Any]:
        """Parse Pricing worksheet."""
        sheet_name = "Pricing"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty or "Field" not in df.columns or "Value" not in df.columns:
            return {}

        pricing = {}
        for _, row in df.iterrows():
            field = row.get("Field")
            value = row.get("Value")

            if field and value is not None:
                if field == "priceAmount":
                    try:
                        pricing[field] = float(value)
                    except (ValueError, TypeError):
                        pricing[field] = value
                else:
                    pricing[field] = str(value)

        return {"price": pricing} if pricing else {}

    def _parse_team(self) -> Dict[str, Any]:
        """Parse Team worksheet."""
        sheet_name = "Team"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        team_members = []
        for _, row in df.iterrows():
            member = {}

            field_mappings = {
                "Username": "username",
                "Name": "name",
                "Role": "role",
                "Description": "description",
                "Date In": "dateIn",
                "Date Out": "dateOut",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    member[odcs_field] = str(value)

            if member:
                team_members.append(member)

        return {"team": team_members} if team_members else {}

    def _parse_roles(self) -> Dict[str, Any]:
        """Parse Roles worksheet."""
        sheet_name = "Roles"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        roles = []
        for _, row in df.iterrows():
            role = {}

            field_mappings = {
                "Role": "role",
                "Description": "description",
                "Access": "access",
                "First Level Approvers": "firstLevelApprovers",
                "Second Level Approvers": "secondLevelApprovers",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    role[odcs_field] = str(value)

            if role:
                roles.append(role)

        return {"roles": roles} if roles else {}

    def _parse_sla_properties(self) -> Dict[str, Any]:
        """Parse SLA Properties worksheet."""
        sheet_name = "SLA Properties"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        sla_properties = []
        for _, row in df.iterrows():
            sla_prop = {}

            field_mappings = {
                "Property": "property",
                "Value": "value",
                "Value Ext": "valueExt",
                "Unit": "unit",
                "Element": "element",
                "Driver": "driver",
            }

            for excel_field, odcs_field in field_mappings.items():
                value = row.get(excel_field)
                if value is not None and value != "":
                    if odcs_field in ["value", "valueExt"]:
                        # Try to convert to appropriate type
                        sla_prop[odcs_field] = self._convert_value(value)
                    else:
                        sla_prop[odcs_field] = str(value)

            if sla_prop:
                sla_properties.append(sla_prop)

        return {"slaProperties": sla_properties} if sla_properties else {}

    def _parse_authoritative_definitions(self) -> Dict[str, Any]:
        """Parse Authoritative Definitions worksheet."""
        sheet_name = "Authoritative Definitions"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty:
            return {}

        definitions = []
        for _, row in df.iterrows():
            definition = {}

            url = row.get("URL")
            type_val = row.get("Type")

            if url and url != "":
                definition["url"] = str(url)
            if type_val and type_val != "":
                definition["type"] = str(type_val)

            if definition:
                definitions.append(definition)

        return {"authoritativeDefinitions": definitions} if definitions else {}

    def _parse_custom_properties(self) -> Dict[str, Any]:
        """Parse Custom Properties worksheet."""
        sheet_name = "Custom Properties"
        if sheet_name not in self.worksheets:
            return {}

        df = self.worksheets[sheet_name]
        if df.empty or "Property" not in df.columns or "Value" not in df.columns:
            return {}

        custom_properties = []
        for _, row in df.iterrows():
            prop = row.get("Property")
            value = row.get("Value")

            if prop and value is not None:
                custom_properties.append(
                    {"property": str(prop), "value": self._convert_value(value)}
                )

        return {"customProperties": custom_properties} if custom_properties else {}

    def _enhance_schema_with_properties(self, odcs_data: Dict[str, Any]) -> None:
        """Enhance schema objects with detailed property information."""
        sheet_name = "Schema Properties"
        if sheet_name not in self.worksheets or "schema" not in odcs_data:
            return

        df = self.worksheets[sheet_name]
        if df.empty:
            return

        # Create a map of schema objects by name for easy lookup
        schema_objects = {obj["name"]: obj for obj in odcs_data["schema"]}

        # Initialize properties array for each schema object if not exists
        for obj in odcs_data["schema"]:
            if "properties" not in obj:
                obj["properties"] = []

        for _, row in df.iterrows():
            obj_name = str(row.get("Object Name", ""))
            if obj_name in schema_objects and pd.notna(row.get("Property Name")):
                prop = {
                    "name": str(row.get("Property Name", "")),
                    "logicalType": self._parse_string(row.get("Logical Type")),
                    "physicalType": self._parse_string(row.get("Physical Type")),
                    "physicalName": self._parse_string(row.get("Physical Name")),
                    "description": self._parse_string(row.get("Description")),
                    "businessName": self._parse_string(row.get("Business Name")),
                    "required": self._parse_boolean(row.get("Required", False)),
                    "unique": self._parse_boolean(row.get("Unique", False)),
                    "primaryKey": self._parse_boolean(row.get("Primary Key", False)),
                    "primaryKeyPosition": self._parse_int(row.get("PK Position", -1)),
                    "partitioned": self._parse_boolean(row.get("Partitioned", False)),
                    "partitionKeyPosition": self._parse_int(
                        row.get("Partition Position", -1)
                    ),
                    "classification": self._parse_string(row.get("Classification")),
                    "encryptedName": self._parse_string(row.get("Encrypted Name")),
                    "criticalDataElement": self._parse_boolean(
                        row.get("Critical Data Element", False)
                    ),
                    "transformLogic": self._parse_string(row.get("Transform Logic")),
                    "transformDescription": self._parse_string(
                        row.get("Transform Description")
                    ),
                }

                # Parse transform sources
                transform_sources_str = self._parse_string(row.get("Transform Sources"))
                if transform_sources_str:
                    prop["transformSourceObjects"] = [
                        s.strip() for s in transform_sources_str.split(",") if s.strip()
                    ]

                # Parse examples
                examples_str = self._parse_string(row.get("Examples"))
                if examples_str:
                    prop["examples"] = [
                        ex.strip() for ex in examples_str.split(",") if ex.strip()
                    ]

                # Parse tags
                tags_str = self._parse_string(row.get("Tags"))
                if tags_str:
                    prop["tags"] = [
                        tag.strip() for tag in tags_str.split(",") if tag.strip()
                    ]

                # Initialize empty collections
                prop["quality"] = []
                prop["authoritativeDefinitions"] = []

                # Remove None values and empty strings
                prop = {k: v for k, v in prop.items() if v is not None and v != ""}
                schema_objects[obj_name]["properties"].append(prop)

    def _enhance_schema_with_logical_type_options(
        self, odcs_data: Dict[str, Any]
    ) -> None:
        """Enhance schema properties with logical type options."""
        sheet_name = "Logical Type Options"
        if sheet_name not in self.worksheets or "schema" not in odcs_data:
            return

        df = self.worksheets[sheet_name]
        if df.empty:
            return

        # Create a map for easy property lookup
        for _, row in df.iterrows():
            obj_name = str(row.get("Object Name", ""))
            prop_name = str(row.get("Property Name", ""))

            # Find the property in the schema
            for obj in odcs_data["schema"]:
                if obj["name"] == obj_name:
                    for prop in obj.get("properties", []):
                        if prop["name"] == prop_name:
                            options = {}

                            # String options
                            if pd.notna(row.get("Format")):
                                options["format"] = self._parse_string(
                                    row.get("Format")
                                )
                            if pd.notna(row.get("Min Length")):
                                options["minLength"] = self._parse_int(
                                    row.get("Min Length")
                                )
                            if pd.notna(row.get("Max Length")):
                                options["maxLength"] = self._parse_int(
                                    row.get("Max Length")
                                )
                            if pd.notna(row.get("Pattern")):
                                options["pattern"] = self._parse_string(
                                    row.get("Pattern")
                                )

                            # Number options
                            if pd.notna(row.get("Minimum")):
                                options["minimum"] = self._parse_number(
                                    row.get("Minimum")
                                )
                            if pd.notna(row.get("Maximum")):
                                options["maximum"] = self._parse_number(
                                    row.get("Maximum")
                                )
                            if pd.notna(row.get("Exclusive Minimum")):
                                options["exclusiveMinimum"] = self._parse_boolean(
                                    row.get("Exclusive Minimum")
                                )
                            if pd.notna(row.get("Exclusive Maximum")):
                                options["exclusiveMaximum"] = self._parse_boolean(
                                    row.get("Exclusive Maximum")
                                )
                            if pd.notna(row.get("Multiple Of")):
                                options["multipleOf"] = self._parse_number(
                                    row.get("Multiple Of")
                                )

                            # Array options
                            if pd.notna(row.get("Min Items")):
                                options["minItems"] = self._parse_int(
                                    row.get("Min Items")
                                )
                            if pd.notna(row.get("Max Items")):
                                options["maxItems"] = self._parse_int(
                                    row.get("Max Items")
                                )
                            if pd.notna(row.get("Unique Items")):
                                options["uniqueItems"] = self._parse_boolean(
                                    row.get("Unique Items")
                                )

                            # Object options
                            if pd.notna(row.get("Min Properties")):
                                options["minProperties"] = self._parse_int(
                                    row.get("Min Properties")
                                )
                            if pd.notna(row.get("Max Properties")):
                                options["maxProperties"] = self._parse_int(
                                    row.get("Max Properties")
                                )

                            required_props_str = self._parse_string(
                                row.get("Required Properties")
                            )
                            if required_props_str:
                                options["required"] = [
                                    p.strip()
                                    for p in required_props_str.split(",")
                                    if p.strip()
                                ]

                            if options:
                                prop["logicalTypeOptions"] = options

    def _enhance_schema_with_quality_rules(self, odcs_data: Dict[str, Any]) -> None:
        """Enhance schema objects and properties with quality rules."""
        sheet_name = "Quality Rules"
        if sheet_name not in self.worksheets or "schema" not in odcs_data:
            return

        df = self.worksheets[sheet_name]
        if df.empty:
            return

        # Initialize quality arrays for each schema object and property if not exists
        for obj in odcs_data["schema"]:
            if "quality" not in obj:
                obj["quality"] = []
            for prop in obj.get("properties", []):
                if "quality" not in prop:
                    prop["quality"] = []

        for _, row in df.iterrows():
            obj_name = str(row.get("Object Name", ""))
            prop_name = str(row.get("Property Name", ""))
            level = str(row.get("Level", ""))

            if not pd.notna(row.get("Name")) and not pd.notna(row.get("Description")):
                continue

            rule = {
                "name": self._parse_string(row.get("Name")),
                "description": self._parse_string(row.get("Description")),
                "type": self._parse_string(row.get("Type")) or "library",
                "rule": self._parse_string(row.get("Rule")),
                "dimension": self._parse_string(row.get("Dimension")),
                "severity": self._parse_string(row.get("Severity")),
                "businessImpact": self._parse_string(row.get("Business Impact")),
                "unit": self._parse_string(row.get("Unit")),
                "query": self._parse_string(row.get("Query")),
                "engine": self._parse_string(row.get("Engine")),
                "implementation": self._parse_string(row.get("Implementation")),
                "method": self._parse_string(row.get("Method")),
                "scheduler": self._parse_string(row.get("Scheduler")),
                "schedule": self._parse_string(row.get("Schedule")),
            }

            # Parse comparison operators
            if pd.notna(row.get("Must Be")):
                rule["mustBe"] = self._parse_number(row.get("Must Be"))
            if pd.notna(row.get("Must Not Be")):
                rule["mustNotBe"] = self._parse_number(row.get("Must Not Be"))
            if pd.notna(row.get("Must Be Greater Than")):
                rule["mustBeGreaterThan"] = self._parse_number(
                    row.get("Must Be Greater Than")
                )
            if pd.notna(row.get("Must Be Greater Or Equal")):
                rule["mustBeGreaterOrEqualTo"] = self._parse_number(
                    row.get("Must Be Greater Or Equal")
                )
            if pd.notna(row.get("Must Be Less Than")):
                rule["mustBeLessThan"] = self._parse_number(
                    row.get("Must Be Less Than")
                )
            if pd.notna(row.get("Must Be Less Or Equal")):
                rule["mustBeLessOrEqualTo"] = self._parse_number(
                    row.get("Must Be Less Or Equal")
                )

            # Parse between operators
            between_str = self._parse_string(row.get("Must Be Between"))
            if between_str:
                try:
                    rule["mustBeBetween"] = [
                        self._parse_number(v.strip())
                        for v in between_str.split(",")
                        if v.strip()
                    ]
                except (ValueError, TypeError):
                    pass

            not_between_str = self._parse_string(row.get("Must Not Be Between"))
            if not_between_str:
                try:
                    rule["mustNotBeBetween"] = [
                        self._parse_number(v.strip())
                        for v in not_between_str.split(",")
                        if v.strip()
                    ]
                except (ValueError, TypeError):
                    pass

            # Parse valid values
            valid_values_str = self._parse_string(row.get("Valid Values"))
            if valid_values_str:
                rule["validValues"] = [
                    v.strip() for v in valid_values_str.split(",") if v.strip()
                ]

            # Parse tags
            tags_str = self._parse_string(row.get("Tags"))
            if tags_str:
                rule["tags"] = [
                    tag.strip() for tag in tags_str.split(",") if tag.strip()
                ]

            # Remove None values
            rule = {k: v for k, v in rule.items() if v is not None and v != ""}

            # Add rule to appropriate object or property
            for obj in odcs_data["schema"]:
                if obj["name"] == obj_name:
                    if level == "object" or not prop_name:
                        if "quality" not in obj:
                            obj["quality"] = []
                        obj["quality"].append(rule)
                    elif level == "property" and prop_name:
                        for prop in obj.get("properties", []):
                            if prop["name"] == prop_name:
                                if "quality" not in prop:
                                    prop["quality"] = []
                                prop["quality"].append(rule)

    def _parse_boolean(self, value) -> bool:
        """Parse boolean value from Excel."""
        if pd.isna(value):
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("true", "yes", "1", "on")
        return bool(value)

    def _parse_int(self, value) -> int:
        """Parse integer value from Excel."""
        if pd.isna(value):
            return -1
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return -1

    def _parse_number(self, value) -> Union[int, float]:
        """Parse numeric value from Excel."""
        if pd.isna(value):
            return None
        try:
            num = float(value)
            return int(num) if num.is_integer() else num
        except (ValueError, TypeError):
            return None

    def _parse_string(self, value) -> Optional[str]:
        """Parse string value from Excel, handling None/NaN properly."""
        if pd.isna(value) or value is None:
            return None
        str_val = str(value).strip()
        if str_val == "" or str_val.lower() == "nan":
            return None
        return str_val

    def _convert_value(self, value: Any) -> Any:
        """Convert Excel value to appropriate Python type."""
        if value is None or value == "":
            return None

        # If it's already a basic type, return as-is
        if isinstance(value, (bool, int, float)):
            return value

        value_str = str(value).strip()

        # Try boolean
        if value_str.lower() in ("true", "false"):
            return value_str.lower() == "true"

        # Try integer
        try:
            if "." not in value_str and value_str.isdigit():
                return int(value_str)
        except (ValueError, AttributeError):
            pass

        # Try float
        try:
            if "." in value_str:
                return float(value_str)
        except (ValueError, AttributeError):
            pass

        # Return as string
        return value_str

    def _clean_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Remove None values and empty structures from data."""
        cleaned = {}

        for key, value in data.items():
            if value is None:
                continue
            elif isinstance(value, dict):
                cleaned_dict = self._clean_data(value)
                if cleaned_dict:  # Only add non-empty dicts
                    cleaned[key] = cleaned_dict
            elif isinstance(value, list):
                cleaned_list = []
                for item in value:
                    if isinstance(item, dict):
                        cleaned_item = self._clean_data(item)
                        if cleaned_item:
                            cleaned_list.append(cleaned_item)
                    elif item is not None and item != "":
                        cleaned_list.append(item)
                if cleaned_list:  # Only add non-empty lists
                    cleaned[key] = cleaned_list
            elif value != "" and value is not None:
                cleaned[key] = value

        return cleaned

    def validate_odcs_data(self, data: Dict[str, Any]) -> bool:
        """Validate parsed data against ODCS schema.

        Args:
            data: Parsed ODCS data dictionary

        Returns:
            True if valid, False otherwise
        """
        try:
            ODCSDataContract(**data)
            logger.info("ODCS data validation successful")
            return True
        except Exception as e:
            logger.warning(f"ODCS data validation failed: {e}")
            return False
