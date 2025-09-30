"""Main ODCS Converter implementation."""

import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from .models import ODCSDataContract


logger = logging.getLogger(__name__)


class ODCSToExcelConverter:
    """Convert ODCS JSON/YAML data to Excel format with separate worksheets."""

    def __init__(self, style_config: Optional[Dict[str, Any]] = None):
        """Initialize the generator with optional styling configuration.

        Args:
            style_config: Optional styling configuration for Excel output
        """
        self.style_config = style_config or self._default_style_config()

    def _default_style_config(self) -> Dict[str, Any]:
        """Get default Excel styling configuration."""
        return {
            "header_font": Font(bold=True, color="FFFFFF"),
            "header_fill": PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            ),
            "cell_font": Font(name="Calibri", size=11),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        }

    def generate_from_file(
        self, input_path: Union[str, Path], output_path: Union[str, Path]
    ) -> None:
        """Generate Excel file from ODCS JSON file.

        Args:
            input_path: Path to input ODCS JSON file
            output_path: Path to output Excel file
        """
        input_path = Path(input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found: {input_path}")

        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        self.generate_from_dict(data, output_path)

    def generate_from_url(self, url: str, output_path: Union[str, Path]) -> None:
        """Generate Excel file from ODCS JSON URL.

        Args:
            url: URL to ODCS JSON data
            output_path: Path to output Excel file
        """
        try:
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as e:
            raise ValueError(f"Failed to fetch data from URL {url}: {e}")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON data from URL {url}: {e}")

        self.generate_from_dict(data, output_path)

    def generate_from_dict(
        self, data: Dict[str, Any], output_path: Union[str, Path]
    ) -> None:
        """Generate Excel file from ODCS data dictionary.

        Args:
            data: ODCS data as dictionary
            output_path: Path to output Excel file
        """
        try:
            # Validate data against ODCS schema
            contract = ODCSDataContract(**data)
        except ValidationError as e:
            logger.warning(f"Data validation failed: {e}")
            # Continue with raw data if validation fails
            contract = None

        # Generate Excel workbook
        workbook = self._create_workbook(data, contract)

        # Save to file
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        logger.info(f"Excel file generated successfully: {output_path}")

    def _create_workbook(
        self, data: Dict[str, Any], contract: Optional[ODCSDataContract]
    ) -> Workbook:
        """Create Excel workbook with ODCS data.

        Args:
            data: Raw ODCS data
            contract: Validated ODCS contract (if validation succeeded)

        Returns:
            Excel workbook
        """
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        # Create worksheets for each top-level field
        self._create_basic_info_sheet(workbook, data)
        self._create_tags_sheet(workbook, data.get("tags", []))
        self._create_description_sheet(workbook, data.get("description", {}))
        self._create_servers_sheet(workbook, data.get("servers", []))
        self._create_schema_sheet(workbook, data.get("schema", []))
        self._create_schema_properties_sheet(workbook, data.get("schema", []))
        self._create_logical_type_options_sheet(workbook, data.get("schema", []))
        self._create_quality_rules_sheet(workbook, data.get("schema", []))
        self._create_support_sheet(workbook, data.get("support", []))
        self._create_pricing_sheet(workbook, data.get("price", {}))
        self._create_team_sheet(workbook, data.get("team", []))
        self._create_roles_sheet(workbook, data.get("roles", []))
        self._create_sla_properties_sheet(workbook, data.get("slaProperties", []))
        self._create_authoritative_definitions_sheet(
            workbook, data.get("authoritativeDefinitions", [])
        )
        self._create_custom_properties_sheet(workbook, data.get("customProperties", []))

        return workbook

    def _create_basic_info_sheet(
        self, workbook: Workbook, data: Dict[str, Any]
    ) -> None:
        """Create Basic Information worksheet."""
        sheet = workbook.create_sheet("Basic Information")

        # Define basic info fields
        basic_fields = [
            ("version", "Contract Version"),
            ("kind", "Contract Kind"),
            ("apiVersion", "API Version"),
            ("id", "Unique Identifier"),
            ("name", "Contract Name"),
            ("tenant", "Tenant"),
            ("status", "Status"),
            ("dataProduct", "Data Product"),
            ("domain", "Domain"),
            ("slaDefaultElement", "SLA Default Element"),
            ("contractCreatedTs", "Created Timestamp"),
        ]

        # Add headers
        headers = ["Field", "Value", "Description"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add data rows
        for row, (field_key, field_desc) in enumerate(basic_fields, 2):
            sheet.cell(row=row, column=1, value=field_key)
            sheet.cell(row=row, column=2, value=str(data.get(field_key, "")))
            sheet.cell(row=row, column=3, value=field_desc)

        self._auto_adjust_columns(sheet)

    def _create_tags_sheet(self, workbook: Workbook, tags: List[str]) -> None:
        """Create Tags worksheet."""
        sheet = workbook.create_sheet("Tags")

        # Add header
        cell = sheet.cell(row=1, column=1, value="Tag")
        self._apply_header_style(cell)

        # Add tags
        for row, tag in enumerate(tags, 2):
            sheet.cell(row=row, column=1, value=tag)

        self._auto_adjust_columns(sheet)

    def _create_description_sheet(
        self, workbook: Workbook, description: Dict[str, Any]
    ) -> None:
        """Create Description worksheet."""
        sheet = workbook.create_sheet("Description")

        # Add headers
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add description fields
        desc_fields = [
            ("usage", "Usage"),
            ("purpose", "Purpose"),
            ("limitations", "Limitations"),
        ]

        for row, (field_key, _) in enumerate(desc_fields, 2):
            sheet.cell(row=row, column=1, value=field_key)
            sheet.cell(row=row, column=2, value=str(description.get(field_key, "")))

        self._auto_adjust_columns(sheet)

    def _create_servers_sheet(
        self, workbook: Workbook, servers: List[Dict[str, Any]]
    ) -> None:
        """Create Servers worksheet."""
        sheet = workbook.create_sheet("Servers")

        if not servers:
            sheet.cell(row=1, column=1, value="No servers defined")
            return

        # Add headers
        headers = [
            "Server",
            "Type",
            "Description",
            "Environment",
            "Location",
            "Host",
            "Port",
            "Database",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add server data
        for row, server in enumerate(servers, 2):
            sheet.cell(row=row, column=1, value=server.get("server", ""))
            sheet.cell(row=row, column=2, value=server.get("type", ""))
            sheet.cell(row=row, column=3, value=server.get("description", ""))
            sheet.cell(row=row, column=4, value=server.get("environment", ""))
            sheet.cell(row=row, column=5, value=server.get("location", ""))
            sheet.cell(row=row, column=6, value=server.get("host", ""))
            sheet.cell(row=row, column=7, value=server.get("port", ""))
            sheet.cell(row=row, column=8, value=server.get("database", ""))

        self._auto_adjust_columns(sheet)

    def _create_schema_sheet(
        self, workbook: Workbook, schema: List[Dict[str, Any]]
    ) -> None:
        """Create Schema worksheet."""
        sheet = workbook.create_sheet("Schema")

        if not schema:
            sheet.cell(row=1, column=1, value="No schema objects defined")
            return

        # Add headers
        headers = [
            "Object Name",
            "Physical Name",
            "Logical Type",
            "Physical Type",
            "Description",
            "Business Name",
            "Data Granularity",
            "Tags",
            "Quality Rules Count",
            "Properties Count",
            "Auth Definitions Count",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add schema objects
        for row, obj in enumerate(schema, 2):
            sheet.cell(row=row, column=1, value=obj.get("name", ""))
            sheet.cell(row=row, column=2, value=obj.get("physicalName", ""))
            sheet.cell(row=row, column=3, value=obj.get("logicalType", ""))
            sheet.cell(row=row, column=4, value=obj.get("physicalType", ""))
            sheet.cell(row=row, column=5, value=obj.get("description", ""))
            sheet.cell(row=row, column=6, value=obj.get("businessName", ""))
            sheet.cell(
                row=row, column=7, value=obj.get("dataGranularityDescription", "")
            )
            sheet.cell(row=row, column=8, value=", ".join(obj.get("tags", [])))
            sheet.cell(row=row, column=9, value=len(obj.get("quality", [])))
            sheet.cell(row=row, column=10, value=len(obj.get("properties", [])))
            sheet.cell(
                row=row, column=11, value=len(obj.get("authoritativeDefinitions", []))
            )

        self._auto_adjust_columns(sheet)

    def _create_support_sheet(
        self, workbook: Workbook, support: List[Dict[str, Any]]
    ) -> None:
        """Create Support worksheet."""
        sheet = workbook.create_sheet("Support")

        if not support:
            sheet.cell(row=1, column=1, value="No support channels defined")
            return

        # Add headers
        headers = ["Channel", "URL", "Description", "Tool", "Scope"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add support data
        for row, item in enumerate(support, 2):
            sheet.cell(row=row, column=1, value=item.get("channel", ""))
            sheet.cell(row=row, column=2, value=item.get("url", ""))
            sheet.cell(row=row, column=3, value=item.get("description", ""))
            sheet.cell(row=row, column=4, value=item.get("tool", ""))
            sheet.cell(row=row, column=5, value=item.get("scope", ""))

        self._auto_adjust_columns(sheet)

    def _create_pricing_sheet(
        self, workbook: Workbook, pricing: Dict[str, Any]
    ) -> None:
        """Create Pricing worksheet."""
        sheet = workbook.create_sheet("Pricing")

        # Add headers
        headers = ["Field", "Value"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add pricing fields
        pricing_fields = [
            ("priceAmount", "Price Amount"),
            ("priceCurrency", "Price Currency"),
            ("priceUnit", "Price Unit"),
        ]

        for row, (field_key, _) in enumerate(pricing_fields, 2):
            sheet.cell(row=row, column=1, value=field_key)
            sheet.cell(row=row, column=2, value=str(pricing.get(field_key, "")))

        self._auto_adjust_columns(sheet)

    def _create_team_sheet(
        self, workbook: Workbook, team: List[Dict[str, Any]]
    ) -> None:
        """Create Team worksheet."""
        sheet = workbook.create_sheet("Team")

        if not team:
            sheet.cell(row=1, column=1, value="No team members defined")
            return

        # Add headers
        headers = ["Username", "Name", "Role", "Description", "Date In", "Date Out"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add team data
        for row, member in enumerate(team, 2):
            sheet.cell(row=row, column=1, value=member.get("username", ""))
            sheet.cell(row=row, column=2, value=member.get("name", ""))
            sheet.cell(row=row, column=3, value=member.get("role", ""))
            sheet.cell(row=row, column=4, value=member.get("description", ""))
            sheet.cell(row=row, column=5, value=member.get("dateIn", ""))
            sheet.cell(row=row, column=6, value=member.get("dateOut", ""))

        self._auto_adjust_columns(sheet)

    def _create_roles_sheet(
        self, workbook: Workbook, roles: List[Dict[str, Any]]
    ) -> None:
        """Create Roles worksheet."""
        sheet = workbook.create_sheet("Roles")

        if not roles:
            sheet.cell(row=1, column=1, value="No roles defined")
            return

        # Add headers
        headers = [
            "Role",
            "Description",
            "Access",
            "First Level Approvers",
            "Second Level Approvers",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add roles data
        for row, role in enumerate(roles, 2):
            sheet.cell(row=row, column=1, value=role.get("role", ""))
            sheet.cell(row=row, column=2, value=role.get("description", ""))
            sheet.cell(row=row, column=3, value=role.get("access", ""))
            sheet.cell(row=row, column=4, value=role.get("firstLevelApprovers", ""))
            sheet.cell(row=row, column=5, value=role.get("secondLevelApprovers", ""))

        self._auto_adjust_columns(sheet)

    def _create_schema_properties_sheet(
        self, workbook: Workbook, schema: List[Dict[str, Any]]
    ) -> None:
        """Create Schema Properties worksheet with detailed property information."""
        sheet = workbook.create_sheet("Schema Properties")

        # Collect all properties from all schema objects
        all_properties = []
        for obj in schema:
            obj_name = obj.get("name", "")
            for prop in obj.get("properties", []):
                prop_with_object = {"object_name": obj_name, **prop}
                all_properties.append(prop_with_object)

        if not all_properties:
            sheet.cell(row=1, column=1, value="No schema properties defined")
            return

        # Add headers
        headers = [
            "Object Name",
            "Property Name",
            "Logical Type",
            "Physical Type",
            "Physical Name",
            "Description",
            "Business Name",
            "Required",
            "Unique",
            "Primary Key",
            "PK Position",
            "Partitioned",
            "Partition Position",
            "Classification",
            "Encrypted Name",
            "Critical Data Element",
            "Transform Sources",
            "Transform Logic",
            "Transform Description",
            "Examples",
            "Tags",
            "Quality Rules Count",
            "Auth Definitions Count",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add properties
        for row, prop in enumerate(all_properties, 2):
            sheet.cell(row=row, column=1, value=prop.get("object_name", ""))
            sheet.cell(row=row, column=2, value=prop.get("name", ""))
            sheet.cell(row=row, column=3, value=prop.get("logicalType", ""))
            sheet.cell(row=row, column=4, value=prop.get("physicalType", ""))
            sheet.cell(row=row, column=5, value=prop.get("physicalName", ""))
            sheet.cell(row=row, column=6, value=prop.get("description", ""))
            sheet.cell(row=row, column=7, value=prop.get("businessName", ""))
            sheet.cell(row=row, column=8, value=prop.get("required", False))
            sheet.cell(row=row, column=9, value=prop.get("unique", False))
            sheet.cell(row=row, column=10, value=prop.get("primaryKey", False))
            sheet.cell(row=row, column=11, value=prop.get("primaryKeyPosition", -1))
            sheet.cell(row=row, column=12, value=prop.get("partitioned", False))
            sheet.cell(row=row, column=13, value=prop.get("partitionKeyPosition", -1))
            sheet.cell(row=row, column=14, value=prop.get("classification", ""))
            sheet.cell(row=row, column=15, value=prop.get("encryptedName", ""))
            sheet.cell(row=row, column=16, value=prop.get("criticalDataElement", False))
            sheet.cell(
                row=row,
                column=17,
                value=", ".join(prop.get("transformSourceObjects", [])),
            )
            sheet.cell(row=row, column=18, value=prop.get("transformLogic", ""))
            sheet.cell(row=row, column=19, value=prop.get("transformDescription", ""))
            sheet.cell(
                row=row, column=20, value=", ".join(map(str, prop.get("examples", [])))
            )
            sheet.cell(row=row, column=21, value=", ".join(prop.get("tags", [])))
            sheet.cell(row=row, column=22, value=len(prop.get("quality", [])))
            sheet.cell(
                row=row, column=23, value=len(prop.get("authoritativeDefinitions", []))
            )

        self._auto_adjust_columns(sheet)

    def _create_logical_type_options_sheet(
        self, workbook: Workbook, schema: List[Dict[str, Any]]
    ) -> None:
        """Create Logical Type Options worksheet."""
        sheet = workbook.create_sheet("Logical Type Options")

        # Collect all properties with logical type options
        options_data = []
        for obj in schema:
            obj_name = obj.get("name", "")
            for prop in obj.get("properties", []):
                if prop.get("logicalTypeOptions"):
                    options = prop["logicalTypeOptions"]
                    options_data.append(
                        {
                            "object_name": obj_name,
                            "property_name": prop.get("name", ""),
                            "logical_type": prop.get("logicalType", ""),
                            **options,
                        }
                    )

        if not options_data:
            sheet.cell(row=1, column=1, value="No logical type options defined")
            return

        # Add headers
        headers = [
            "Object Name",
            "Property Name",
            "Logical Type",
            "Format",
            "Min Length",
            "Max Length",
            "Pattern",
            "Minimum",
            "Maximum",
            "Exclusive Minimum",
            "Exclusive Maximum",
            "Multiple Of",
            "Min Items",
            "Max Items",
            "Unique Items",
            "Min Properties",
            "Max Properties",
            "Required Properties",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add options data
        for row, options in enumerate(options_data, 2):
            sheet.cell(row=row, column=1, value=options.get("object_name", ""))
            sheet.cell(row=row, column=2, value=options.get("property_name", ""))
            sheet.cell(row=row, column=3, value=options.get("logical_type", ""))
            sheet.cell(row=row, column=4, value=options.get("format", ""))
            sheet.cell(row=row, column=5, value=options.get("minLength", ""))
            sheet.cell(row=row, column=6, value=options.get("maxLength", ""))
            sheet.cell(row=row, column=7, value=options.get("pattern", ""))
            sheet.cell(row=row, column=8, value=options.get("minimum", ""))
            sheet.cell(row=row, column=9, value=options.get("maximum", ""))
            sheet.cell(row=row, column=10, value=options.get("exclusiveMinimum", ""))
            sheet.cell(row=row, column=11, value=options.get("exclusiveMaximum", ""))
            sheet.cell(row=row, column=12, value=options.get("multipleOf", ""))
            sheet.cell(row=row, column=13, value=options.get("minItems", ""))
            sheet.cell(row=row, column=14, value=options.get("maxItems", ""))
            sheet.cell(row=row, column=15, value=options.get("uniqueItems", ""))
            sheet.cell(row=row, column=16, value=options.get("minProperties", ""))
            sheet.cell(row=row, column=17, value=options.get("maxProperties", ""))
            sheet.cell(row=row, column=18, value=", ".join(options.get("required", [])))

        self._auto_adjust_columns(sheet)

    def _create_quality_rules_sheet(
        self, workbook: Workbook, schema: List[Dict[str, Any]]
    ) -> None:
        """Create Quality Rules worksheet."""
        sheet = workbook.create_sheet("Quality Rules")

        # Collect all quality rules from objects and properties
        all_rules = []
        for obj in schema:
            obj_name = obj.get("name", "")

            # Object-level quality rules
            for rule in obj.get("quality", []):
                rule_with_context = {
                    "object_name": obj_name,
                    "property_name": "",
                    "level": "object",
                    **rule,
                }
                all_rules.append(rule_with_context)

            # Property-level quality rules
            for prop in obj.get("properties", []):
                prop_name = prop.get("name", "")
                for rule in prop.get("quality", []):
                    rule_with_context = {
                        "object_name": obj_name,
                        "property_name": prop_name,
                        "level": "property",
                        **rule,
                    }
                    all_rules.append(rule_with_context)

        if not all_rules:
            sheet.cell(row=1, column=1, value="No quality rules defined")
            return

        # Add headers
        headers = [
            "Object Name",
            "Property Name",
            "Level",
            "Name",
            "Description",
            "Type",
            "Rule",
            "Dimension",
            "Severity",
            "Business Impact",
            "Unit",
            "Valid Values",
            "Query",
            "Engine",
            "Implementation",
            "Must Be",
            "Must Not Be",
            "Must Be Greater Than",
            "Must Be Greater Or Equal",
            "Must Be Less Than",
            "Must Be Less Or Equal",
            "Must Be Between",
            "Must Not Be Between",
            "Method",
            "Scheduler",
            "Schedule",
            "Tags",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add quality rules
        for row, rule in enumerate(all_rules, 2):
            sheet.cell(row=row, column=1, value=rule.get("object_name", ""))
            sheet.cell(row=row, column=2, value=rule.get("property_name", ""))
            sheet.cell(row=row, column=3, value=rule.get("level", ""))
            sheet.cell(row=row, column=4, value=rule.get("name", ""))
            sheet.cell(row=row, column=5, value=rule.get("description", ""))
            sheet.cell(row=row, column=6, value=rule.get("type", ""))
            sheet.cell(row=row, column=7, value=rule.get("rule", ""))
            sheet.cell(row=row, column=8, value=rule.get("dimension", ""))
            sheet.cell(row=row, column=9, value=rule.get("severity", ""))
            sheet.cell(row=row, column=10, value=rule.get("businessImpact", ""))
            sheet.cell(row=row, column=11, value=rule.get("unit", ""))
            sheet.cell(
                row=row,
                column=12,
                value=", ".join(map(str, rule.get("validValues", []))),
            )
            sheet.cell(row=row, column=13, value=rule.get("query", ""))
            sheet.cell(row=row, column=14, value=rule.get("engine", ""))
            sheet.cell(row=row, column=15, value=rule.get("implementation", ""))
            sheet.cell(row=row, column=16, value=rule.get("mustBe", ""))
            sheet.cell(row=row, column=17, value=rule.get("mustNotBe", ""))
            sheet.cell(row=row, column=18, value=rule.get("mustBeGreaterThan", ""))
            sheet.cell(row=row, column=19, value=rule.get("mustBeGreaterOrEqualTo", ""))
            sheet.cell(row=row, column=20, value=rule.get("mustBeLessThan", ""))
            sheet.cell(row=row, column=21, value=rule.get("mustBeLessOrEqualTo", ""))
            sheet.cell(
                row=row,
                column=22,
                value=", ".join(map(str, rule.get("mustBeBetween", []))),
            )
            sheet.cell(
                row=row,
                column=23,
                value=", ".join(map(str, rule.get("mustNotBeBetween", []))),
            )
            sheet.cell(row=row, column=24, value=rule.get("method", ""))
            sheet.cell(row=row, column=25, value=rule.get("scheduler", ""))
            sheet.cell(row=row, column=26, value=rule.get("schedule", ""))
            sheet.cell(row=row, column=27, value=", ".join(rule.get("tags", [])))

        self._auto_adjust_columns(sheet)

    def _create_sla_properties_sheet(
        self, workbook: Workbook, sla_properties: List[Dict[str, Any]]
    ) -> None:
        """Create SLA Properties worksheet."""
        sheet = workbook.create_sheet("SLA Properties")

        if not sla_properties:
            sheet.cell(row=1, column=1, value="No SLA properties defined")
            return

        # Add headers
        headers = ["Property", "Value", "Value Ext", "Unit", "Element", "Driver"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add SLA data
        for row, prop in enumerate(sla_properties, 2):
            sheet.cell(row=row, column=1, value=prop.get("property", ""))
            sheet.cell(row=row, column=2, value=str(prop.get("value", "")))
            sheet.cell(row=row, column=3, value=str(prop.get("valueExt", "")))
            sheet.cell(row=row, column=4, value=prop.get("unit", ""))
            sheet.cell(row=row, column=5, value=prop.get("element", ""))
            sheet.cell(row=row, column=6, value=prop.get("driver", ""))

        self._auto_adjust_columns(sheet)

    def _create_authoritative_definitions_sheet(
        self, workbook: Workbook, definitions: List[Dict[str, Any]]
    ) -> None:
        """Create Authoritative Definitions worksheet."""
        sheet = workbook.create_sheet("Authoritative Definitions")

        if not definitions:
            sheet.cell(row=1, column=1, value="No authoritative definitions defined")
            return

        # Add headers
        headers = ["URL", "Type"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add definitions data
        for row, definition in enumerate(definitions, 2):
            sheet.cell(row=row, column=1, value=definition.get("url", ""))
            sheet.cell(row=row, column=2, value=definition.get("type", ""))

        self._auto_adjust_columns(sheet)

    def _create_custom_properties_sheet(
        self, workbook: Workbook, custom_properties: List[Dict[str, Any]]
    ) -> None:
        """Create Custom Properties worksheet."""
        sheet = workbook.create_sheet("Custom Properties")

        if not custom_properties:
            sheet.cell(row=1, column=1, value="No custom properties defined")
            return

        # Add headers
        headers = ["Property", "Value"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add custom properties data
        for row, prop in enumerate(custom_properties, 2):
            sheet.cell(row=row, column=1, value=prop.get("property", ""))
            sheet.cell(row=row, column=2, value=str(prop.get("value", "")))

        self._auto_adjust_columns(sheet)

    def _apply_header_style(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.font = self.style_config["header_font"]
        cell.fill = self.style_config["header_fill"]
        cell.alignment = self.style_config["alignment"]

    def _auto_adjust_columns(self, sheet) -> None:
        """Auto-adjust column widths."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
