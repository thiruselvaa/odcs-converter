"""Template generator for creating sample Excel templates with field indicators."""

from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from .logging_config import get_logger

logger = get_logger(__name__)


class TemplateType(str, Enum):
    """Available template types."""

    MINIMAL = "minimal"
    REQUIRED = "required"
    FULL = "full"
    CUSTOM = "custom"


class TemplateGenerator:
    """Generate sample Excel templates with field indicators and examples."""

    def __init__(self):
        """Initialize the template generator."""
        self.style_config = self._default_style_config()

    def _default_style_config(self) -> Dict[str, Any]:
        """Get default Excel styling configuration."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        return {
            "header_font": Font(bold=True, color="FFFFFF", size=11),
            "required_header_fill": PatternFill(
                start_color="C00000", end_color="C00000", fill_type="solid"
            ),  # Dark Red for required
            "optional_header_fill": PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            ),  # Blue for optional
            "example_font": Font(name="Calibri", size=10, italic=True, color="808080"),
            "normal_font": Font(name="Calibri", size=10),
            "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "border": thin_border,
        }

    def generate_template(
        self,
        output_path: Union[str, Path],
        template_type: TemplateType = TemplateType.FULL,
        include_examples: bool = True,
    ) -> None:
        """Generate Excel template based on specified type.

        Args:
            output_path: Path to save the template
            template_type: Type of template to generate
            include_examples: Whether to include example values
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        # Create instruction sheet first
        self._create_instructions_sheet(workbook, template_type)

        # Create data sheets based on template type
        if template_type == TemplateType.MINIMAL:
            self._create_minimal_template(workbook, include_examples)
        elif template_type == TemplateType.REQUIRED:
            self._create_required_template(workbook, include_examples)
        elif template_type == TemplateType.FULL:
            self._create_full_template(workbook, include_examples)

        workbook.save(output_path)
        logger.info(f"Template generated successfully: {output_path}")

    def _create_instructions_sheet(
        self, workbook: Workbook, template_type: TemplateType
    ) -> None:
        """Create instructions worksheet."""
        sheet = workbook.create_sheet("📖 Instructions", 0)

        instructions = [
            ["ODCS Data Contract Excel Template", ""],
            ["", ""],
            ["Template Type:", template_type.value.upper()],
            ["", ""],
            ["Color Coding:", ""],
            ["🔴 Red Headers", "= REQUIRED fields (must be filled)"],
            ["🔵 Blue Headers", "= OPTIONAL fields (can be left empty)"],
            ["", ""],
            ["Instructions:", ""],
            ["1. Fill in the required fields (red headers)", ""],
            ["2. Optionally fill in blue header fields for additional details", ""],
            ["3. Delete example rows and replace with your actual data", ""],
            ["4. Do NOT modify the header row colors or names", ""],
            ["5. Save and convert using: odcs-converter excel-to-odcs <filename>", ""],
            ["", ""],
            ["Field Requirements by Sheet:", ""],
            ["", ""],
            ["Basic Information:", "Required: version, kind, apiVersion, id, name"],
            ["", "Optional: tenant, status, dataProduct, domain, etc."],
            ["", ""],
            ["Servers:", "Required: server, type"],
            ["", "Optional: description, environment, and type-specific fields"],
            ["", ""],
            ["Schema:", "Required: name, logicalType"],
            ["", "Optional: physicalName, description, businessName, tags, etc."],
            ["", ""],
            ["Schema Properties:", "Required: name, logicalType (for each property)"],
            ["", "Optional: physicalType, description, required, primaryKey, etc."],
            ["", ""],
            ["Tips:", ""],
            ["• Use the example values as a guide for format and content", ""],
            ["• Hover over cells with comments (red corner) for more help", ""],
            ["• Required fields must have a value to pass validation", ""],
            ["• Boolean fields accept: true, false, TRUE, FALSE, 1, 0", ""],
            ["• Date fields should be in ISO 8601 format: YYYY-MM-DDTHH:MM:SSZ", ""],
            ["", ""],
            ["For more information:", ""],
            ["Documentation: https://github.com/yourorg/odcs-converter", ""],
            ["ODCS Spec: https://open-data-contract-standard.github.io/", ""],
        ]

        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # Style the title
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16, color="1F4E78")
                # Style section headers
                elif value in [
                    "Color Coding:",
                    "Instructions:",
                    "Field Requirements by Sheet:",
                    "Tips:",
                    "For more information:",
                ]:
                    cell.font = Font(bold=True, size=12, color="1F4E78")
                # Style color coding examples
                elif "Red Headers" in value:
                    cell.fill = PatternFill(
                        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                    )
                elif "Blue Headers" in value:
                    cell.fill = PatternFill(
                        start_color="E6F0FF", end_color="E6F0FF", fill_type="solid"
                    )

        # Adjust column widths
        sheet.column_dimensions["A"].width = 50
        sheet.column_dimensions["B"].width = 60

    def _create_minimal_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create minimal template with only essential required fields."""
        self._create_basic_info_minimal(workbook, include_examples)
        self._create_schema_minimal(workbook, include_examples)
        self._create_schema_properties_minimal(workbook, include_examples)

    def _create_required_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create template with all required fields."""
        self._create_basic_info_required(workbook, include_examples)
        self._create_servers_template(workbook, include_examples, required_only=True)
        self._create_schema_required(workbook, include_examples)
        self._create_schema_properties_required(workbook, include_examples)

    def _create_full_template(self, workbook: Workbook, include_examples: bool) -> None:
        """Create full template with all required and optional fields."""
        self._create_basic_info_full(workbook, include_examples)
        self._create_tags_template(workbook, include_examples)
        self._create_description_template(workbook, include_examples)
        self._create_servers_template(workbook, include_examples, required_only=False)
        self._create_schema_full(workbook, include_examples)
        self._create_schema_properties_full(workbook, include_examples)
        self._create_logical_type_options_template(workbook, include_examples)
        self._create_quality_rules_template(workbook, include_examples)
        self._create_support_template(workbook, include_examples)
        self._create_pricing_template(workbook, include_examples)
        self._create_team_template(workbook, include_examples)
        self._create_roles_template(workbook, include_examples)
        self._create_sla_properties_template(workbook, include_examples)
        self._create_authoritative_definitions_template(workbook, include_examples)
        self._create_custom_properties_template(workbook, include_examples)

    def _create_basic_info_minimal(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create minimal basic information sheet."""
        sheet = workbook.create_sheet("Basic Information")

        headers = [
            ("version", True, "Contract version (e.g., 1.0.0)"),
            ("kind", True, "Must be 'DataContract'"),
            ("apiVersion", True, "ODCS API version (e.g., v3.0.2)"),
            ("id", True, "Unique identifier for this contract"),
            ("name", True, "Human-readable name for this contract"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example_row = [
                "1.0.0",
                "DataContract",
                "v3.0.2",
                "my-data-contract-001",
                "My Data Contract",
            ]
            self._add_example_row(sheet, example_row, 2)

        self._auto_adjust_columns(sheet)

    def _create_basic_info_required(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create basic information sheet with all required fields."""
        sheet = workbook.create_sheet("Basic Information")

        headers = [
            ("version", True, "Contract version (e.g., 1.0.0)"),
            ("kind", True, "Must be 'DataContract'"),
            ("apiVersion", True, "ODCS API version (e.g., v3.0.2)"),
            ("id", True, "Unique identifier for this contract"),
            ("name", True, "Human-readable name for this contract"),
            ("tenant", False, "Tenant or organization name"),
            ("status", False, "Contract status (e.g., active, draft)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example_row = [
                "1.0.0",
                "DataContract",
                "v3.0.2",
                "my-data-contract-001",
                "My Data Contract",
                "my-team",
                "active",
            ]
            self._add_example_row(sheet, example_row, 2)

        self._auto_adjust_columns(sheet)

    def _create_basic_info_full(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create full basic information sheet."""
        sheet = workbook.create_sheet("Basic Information")

        headers = [
            ("version", True, "Contract version (e.g., 1.0.0)"),
            ("kind", True, "Must be 'DataContract'"),
            ("apiVersion", True, "ODCS API version (e.g., v3.0.2)"),
            ("id", True, "Unique identifier for this contract"),
            ("name", True, "Human-readable name for this contract"),
            ("tenant", False, "Tenant or organization name"),
            ("status", False, "Contract status (e.g., active, draft, deprecated)"),
            ("dataProduct", False, "Name of the data product"),
            ("domain", False, "Business domain (e.g., finance, marketing)"),
            ("slaDefaultElement", False, "Default element for SLA properties"),
            (
                "contractCreatedTs",
                False,
                "ISO 8601 timestamp when contract was created",
            ),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example_row = [
                "1.0.0",
                "DataContract",
                "v3.0.2",
                "user-analytics-001",
                "User Analytics Data Contract",
                "analytics-team",
                "active",
                "User Behavior Analytics",
                "user_analytics",
                "user_sessions",
                "2024-01-15T09:00:00Z",
            ]
            self._add_example_row(sheet, example_row, 2)

        self._auto_adjust_columns(sheet)

    def _create_tags_template(self, workbook: Workbook, include_examples: bool) -> None:
        """Create tags template sheet."""
        sheet = workbook.create_sheet("Tags")

        headers = [("tag", False, "Tag or label for categorization")]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = ["analytics", "user-data", "behavioral", "production"]
            for idx, tag in enumerate(examples, 2):
                self._add_example_row(sheet, [tag], idx)

        self._auto_adjust_columns(sheet)

    def _create_description_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create description template sheet."""
        sheet = workbook.create_sheet("Description")

        headers = [
            ("usage", False, "How this data should be used"),
            ("purpose", False, "Why this data exists"),
            ("limitations", False, "Known limitations or constraints"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example = [
                "This dataset provides user behavioral analytics for product optimization",
                "Enable data-driven decision making for product development",
                "Data is aggregated daily with up to 24-hour latency",
            ]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_servers_template(
        self, workbook: Workbook, include_examples: bool, required_only: bool = False
    ) -> None:
        """Create servers template sheet."""
        sheet = workbook.create_sheet("Servers")

        if required_only:
            headers = [
                ("server", True, "Server identifier/name"),
                ("type", True, "Server type (e.g., snowflake, bigquery, postgresql)"),
                ("description", False, "Description of this server"),
            ]
        else:
            headers = [
                ("server", True, "Server identifier/name"),
                ("type", True, "Server type (e.g., snowflake, bigquery, postgresql)"),
                ("description", False, "Description of this server"),
                ("environment", False, "Environment (e.g., production, staging, dev)"),
                ("account", False, "Account identifier (Snowflake/cloud specific)"),
                ("database", False, "Database name"),
                ("schema", False, "Schema name"),
                ("warehouse", False, "Warehouse name (Snowflake specific)"),
                ("project", False, "Project ID (BigQuery/GCP specific)"),
                ("dataset", False, "Dataset name (BigQuery specific)"),
                ("host", False, "Host address"),
                ("port", False, "Port number"),
            ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            if required_only:
                example = [
                    "analytics-warehouse",
                    "snowflake",
                    "Primary analytics data warehouse",
                ]
            else:
                example = [
                    "analytics-warehouse",
                    "snowflake",
                    "Primary analytics data warehouse",
                    "production",
                    "company-analytics",
                    "ANALYTICS_DB",
                    "USER_BEHAVIOR",
                    "ANALYTICS_WH",
                    "",
                    "",
                    "",
                    "",
                ]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_schema_minimal(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create minimal schema template sheet."""
        sheet = workbook.create_sheet("Schema")

        headers = [
            ("name", True, "Schema object name (table/dataset name)"),
            ("logicalType", True, "Logical type (usually 'object' for tables)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example = ["user_sessions", "object"]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_schema_required(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create schema template with required fields."""
        sheet = workbook.create_sheet("Schema")

        headers = [
            ("name", True, "Schema object name (table/dataset name)"),
            ("logicalType", True, "Logical type (usually 'object' for tables)"),
            ("description", False, "Description of this schema object"),
            ("physicalName", False, "Physical name in the database"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example = [
                "user_sessions",
                "object",
                "User session data with behavioral metrics",
                "user_sessions_fact",
            ]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_schema_full(self, workbook: Workbook, include_examples: bool) -> None:
        """Create full schema template sheet."""
        sheet = workbook.create_sheet("Schema")

        headers = [
            ("name", True, "Schema object name (table/dataset name)"),
            ("logicalType", True, "Logical type (usually 'object' for tables)"),
            ("physicalName", False, "Physical name in the database"),
            ("description", False, "Description of this schema object"),
            ("businessName", False, "Business-friendly name"),
            ("dataGranularityDescription", False, "What each record represents"),
            ("tags", False, "Comma-separated tags for this object"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example = [
                "user_sessions",
                "object",
                "user_sessions_fact",
                "User session data with behavioral metrics",
                "User Sessions",
                "One record per user session",
                "fact-table,behavioral",
            ]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_schema_properties_minimal(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create minimal schema properties template."""
        sheet = workbook.create_sheet("Schema Properties")

        headers = [
            ("schemaName", True, "Name of the schema object this property belongs to"),
            ("name", True, "Property/column name"),
            (
                "logicalType",
                True,
                "Logical type (string, number, integer, date, boolean)",
            ),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                ["user_sessions", "session_id", "string"],
                ["user_sessions", "user_id", "string"],
                ["user_sessions", "session_start_ts", "date"],
                ["user_sessions", "page_views", "integer"],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_schema_properties_required(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create schema properties template with required fields."""
        sheet = workbook.create_sheet("Schema Properties")

        headers = [
            ("schemaName", True, "Name of the schema object this property belongs to"),
            ("name", True, "Property/column name"),
            (
                "logicalType",
                True,
                "Logical type (string, number, integer, date, boolean, object, array)",
            ),
            ("description", False, "Description of this property"),
            ("required", False, "Is this property required? (true/false)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "user_sessions",
                    "session_id",
                    "string",
                    "Unique session identifier",
                    "true",
                ],
                [
                    "user_sessions",
                    "user_id",
                    "string",
                    "Anonymized user identifier",
                    "true",
                ],
                [
                    "user_sessions",
                    "session_start_ts",
                    "date",
                    "Session start timestamp in UTC",
                    "true",
                ],
                [
                    "user_sessions",
                    "page_views",
                    "integer",
                    "Number of pages viewed",
                    "true",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_schema_properties_full(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create full schema properties template."""
        sheet = workbook.create_sheet("Schema Properties")

        headers = [
            ("schemaName", True, "Schema object this property belongs to"),
            ("name", True, "Property/column name"),
            (
                "logicalType",
                True,
                "Logical type (string, number, integer, date, boolean, object, array)",
            ),
            (
                "physicalType",
                False,
                "Physical database type (e.g., VARCHAR(100), INTEGER)",
            ),
            ("description", False, "Description of this property"),
            ("required", False, "Is this required? (true/false)"),
            ("primaryKey", False, "Is this a primary key? (true/false)"),
            (
                "primaryKeyPosition",
                False,
                "Position in composite primary key (1, 2, 3...)",
            ),
            (
                "classification",
                False,
                "Data classification (e.g., public, internal, confidential, restricted)",
            ),
            ("pii", False, "Contains PII? (true/false)"),
            ("examples", False, "Comma-separated example values"),
            ("pattern", False, "Regex pattern for validation"),
            ("minLength", False, "Minimum length (for strings)"),
            ("maxLength", False, "Maximum length (for strings)"),
            ("minimum", False, "Minimum value (for numbers)"),
            ("maximum", False, "Maximum value (for numbers)"),
            ("tags", False, "Comma-separated tags"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "user_sessions",
                    "session_id",
                    "string",
                    "VARCHAR(36)",
                    "Unique session identifier",
                    "true",
                    "true",
                    "1",
                    "internal",
                    "false",
                    "550e8400-e29b-41d4-a716-446655440000",
                    "",
                    "",
                    "36",
                    "",
                    "",
                    "",
                ],
                [
                    "user_sessions",
                    "user_id",
                    "string",
                    "VARCHAR(36)",
                    "Anonymized user identifier",
                    "true",
                    "false",
                    "",
                    "restricted",
                    "true",
                    "user_abc123xyz",
                    "",
                    "",
                    "36",
                    "",
                    "",
                    "",
                ],
                [
                    "user_sessions",
                    "session_start_ts",
                    "date",
                    "TIMESTAMP_NTZ",
                    "Session start timestamp in UTC",
                    "true",
                    "false",
                    "",
                    "internal",
                    "false",
                    "2024-01-15T10:30:00Z",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ],
                [
                    "user_sessions",
                    "page_views",
                    "integer",
                    "INTEGER",
                    "Number of pages viewed in session",
                    "true",
                    "false",
                    "",
                    "internal",
                    "false",
                    "5,12,3",
                    "",
                    "",
                    "",
                    "0",
                    "1000",
                    "",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_logical_type_options_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create logical type options template."""
        sheet = workbook.create_sheet("Logical Type Options")

        headers = [
            ("schemaName", True, "Schema object name"),
            ("propertyName", True, "Property name"),
            ("format", False, "Format specification"),
            ("enum", False, "Comma-separated allowed values"),
            ("precision", False, "Numeric precision"),
            ("scale", False, "Numeric scale"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                ["user_sessions", "session_start_ts", "date-time", "", "", ""],
                ["user_sessions", "status", "", "active,inactive,pending", "", ""],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_quality_rules_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create quality rules template."""
        sheet = workbook.create_sheet("Quality Rules")

        headers = [
            ("schemaName", True, "Schema object name"),
            ("name", True, "Quality rule name"),
            ("description", False, "Description of this quality rule"),
            (
                "dimension",
                False,
                "Quality dimension (uniqueness, completeness, accuracy, etc.)",
            ),
            ("type", False, "Rule type (library, custom, etc.)"),
            ("rule", False, "Rule specification"),
            ("mustBe", False, "Expected value"),
            ("mustNotBe", False, "Disallowed value"),
            ("severity", False, "Severity level (error, warning, info)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "user_sessions",
                    "session_id_uniqueness",
                    "Ensure session IDs are unique",
                    "uniqueness",
                    "library",
                    "duplicateCount",
                    "0",
                    "",
                    "error",
                ],
                [
                    "user_sessions",
                    "session_start_completeness",
                    "Session start timestamp must be present",
                    "completeness",
                    "library",
                    "nullCount",
                    "0",
                    "",
                    "error",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_support_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create support template."""
        sheet = workbook.create_sheet("Support")

        headers = [
            ("channel", False, "Support channel name"),
            ("url", False, "URL or contact information"),
            ("description", False, "Description of this support channel"),
            ("tool", False, "Tool used (email, slack, jira, etc.)"),
            ("scope", False, "Scope (issues, interactive, etc.)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "analytics-support",
                    "mailto:analytics-team@company.com",
                    "Primary support for analytics questions",
                    "email",
                    "issues",
                ],
                [
                    "analytics-slack",
                    "https://company.slack.com/channels/analytics",
                    "Real-time support and announcements",
                    "slack",
                    "interactive",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_pricing_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create pricing template."""
        sheet = workbook.create_sheet("Pricing")

        headers = [
            ("priceAmount", False, "Price amount"),
            ("priceCurrency", False, "Currency code (USD, EUR, etc.)"),
            (
                "priceUnit",
                False,
                "Unit of pricing (per query, per GB, per month, etc.)",
            ),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            example = ["0.01", "USD", "per query"]
            self._add_example_row(sheet, example, 2)

        self._auto_adjust_columns(sheet)

    def _create_team_template(self, workbook: Workbook, include_examples: bool) -> None:
        """Create team template."""
        sheet = workbook.create_sheet("Team")

        headers = [
            ("username", False, "Username or email"),
            ("name", False, "Full name"),
            ("role", False, "Role in the team"),
            ("description", False, "Description of responsibilities"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "jane.smith@company.com",
                    "Jane Smith",
                    "Data Product Owner",
                    "Responsible for data product strategy",
                ],
                [
                    "john.doe@company.com",
                    "John Doe",
                    "Senior Data Engineer",
                    "Technical lead for data pipeline development",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_roles_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create roles template."""
        sheet = workbook.create_sheet("Roles")

        headers = [
            ("role", False, "Role name"),
            ("description", False, "Description of this role"),
            ("access", False, "Access permissions (SELECT, INSERT, UPDATE, DELETE)"),
            (
                "firstLevelApprovers",
                False,
                "Comma-separated list of first-level approvers",
            ),
            (
                "secondLevelApprovers",
                False,
                "Comma-separated list of second-level approvers",
            ),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                [
                    "analytics_reader",
                    "Read-only access to analytics data",
                    "SELECT",
                    "jane.smith@company.com",
                    "",
                ],
                [
                    "analytics_writer",
                    "Read and write access to analytics data",
                    "SELECT, INSERT, UPDATE",
                    "jane.smith@company.com",
                    "data-governance@company.com",
                ],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_sla_properties_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create SLA properties template."""
        sheet = workbook.create_sheet("SLA Properties")

        headers = [
            (
                "property",
                False,
                "SLA property name (availability, freshness, latency, etc.)",
            ),
            ("value", False, "SLA value"),
            ("unit", False, "Unit (percent, hours, minutes, seconds, etc.)"),
            ("element", False, "Element this SLA applies to"),
            ("driver", False, "Driver (operational, regulatory, etc.)"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                ["availability", "99.9", "percent", "user_sessions", "operational"],
                ["freshness", "24", "hours", "user_sessions", "operational"],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_authoritative_definitions_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create authoritative definitions template."""
        sheet = workbook.create_sheet("Authoritative Definitions")

        headers = [
            ("url", False, "URL to authoritative definition"),
            (
                "type",
                False,
                "Type of definition (businessDefinition, implementation, etc.)",
            ),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                ["https://wiki.company.com/data/user-analytics", "businessDefinition"],
                ["https://github.com/company/analytics-pipelines", "implementation"],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _create_custom_properties_template(
        self, workbook: Workbook, include_examples: bool
    ) -> None:
        """Create custom properties template."""
        sheet = workbook.create_sheet("Custom Properties")

        headers = [
            ("property", False, "Custom property name"),
            ("value", False, "Property value"),
        ]

        self._add_headers_with_style(sheet, headers)

        if include_examples:
            examples = [
                ["dataRetentionDays", "2555"],
                ["complianceClassification", "internal"],
            ]
            for idx, example in enumerate(examples, 2):
                self._add_example_row(sheet, example, idx)

        self._auto_adjust_columns(sheet)

    def _add_headers_with_style(self, sheet, headers: List[tuple]) -> None:
        """Add headers with appropriate styling based on required/optional status.

        Args:
            sheet: Worksheet to add headers to
            headers: List of tuples (header_name, is_required, help_text)
        """
        for col_idx, (header_name, is_required, help_text) in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header_name)

            # Apply font
            cell.font = self.style_config["header_font"]

            # Apply fill based on required status
            if is_required:
                cell.fill = self.style_config["required_header_fill"]
            else:
                cell.fill = self.style_config["optional_header_fill"]

            # Apply alignment and border
            cell.alignment = self.style_config["alignment"]
            cell.border = self.style_config["border"]

            # Add comment with help text
            if help_text:
                comment = Comment(help_text, "ODCS Converter")
                cell.comment = comment

    def _add_example_row(self, sheet, values: List[Any], row_idx: int) -> None:
        """Add example row with italic gray text."""
        for col_idx, value in enumerate(values, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = self.style_config["example_font"]
            cell.alignment = self.style_config["alignment"]
            cell.border = self.style_config["border"]

    def _auto_adjust_columns(self, sheet) -> None:
        """Auto-adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
