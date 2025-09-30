"""Tests for ODCS Excel Generator."""

import json
import tempfile
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import load_workbook

from odcs_excel_generator.generator import ODCSExcelGenerator
from odcs_excel_generator.models import ODCSDataContract


class TestODCSExcelGenerator:
    """Test cases for ODCSExcelGenerator."""
    
    @pytest.fixture
    def sample_odcs_data(self):
        """Sample ODCS data for testing."""
        return {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "test-contract-123",
            "status": "active",
            "name": "Test Data Contract",
            "tenant": "test-tenant",
            "tags": ["test", "example"],
            "description": {
                "usage": "Test usage",
                "purpose": "Test purpose",
                "limitations": "Test limitations"
            },
            "servers": [
                {
                    "server": "test-server",
                    "type": "postgresql",
                    "description": "Test database server",
                    "environment": "test",
                    "host": "localhost",
                    "port": 5432,
                    "database": "testdb"
                }
            ],
            "schema": [
                {
                    "name": "users",
                    "logicalType": "object",
                    "physicalName": "users_table",
                    "description": "User data table",
                    "properties": [
                        {
                            "name": "id",
                            "logicalType": "integer",
                            "physicalType": "BIGINT",
                            "description": "User ID",
                            "primaryKey": True
                        },
                        {
                            "name": "email",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(255)",
                            "description": "User email address",
                            "required": True
                        }
                    ]
                }
            ]
        }
    
    @pytest.fixture
    def generator(self):
        """Generator instance for testing."""
        return ODCSExcelGenerator()
    
    def test_generator_initialization(self, generator):
        """Test generator initialization."""
        assert generator is not None
        assert generator.style_config is not None
        assert "header_font" in generator.style_config
        assert "header_fill" in generator.style_config
    
    def test_generate_from_dict(self, generator, sample_odcs_data):
        """Test generating Excel from dictionary."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            output_path = tmp_file.name
        
        try:
            generator.generate_from_dict(sample_odcs_data, output_path)
            
            # Verify file was created
            assert Path(output_path).exists()
            
            # Load and verify Excel content
            workbook = load_workbook(output_path)
            
            # Check that expected worksheets exist
            expected_sheets = [
                "Basic Information",
                "Tags", 
                "Description",
                "Servers",
                "Schema"
            ]
            
            for sheet_name in expected_sheets:
                assert sheet_name in workbook.sheetnames
            
            # Verify basic info sheet content
            basic_info_sheet = workbook["Basic Information"]
            assert basic_info_sheet["A1"].value == "Field"
            assert basic_info_sheet["B1"].value == "Value"
            assert basic_info_sheet["C1"].value == "Description"
            
            # Verify some data was written
            assert basic_info_sheet["A2"].value == "version"
            assert basic_info_sheet["B2"].value == "1.0.0"
            
        finally:
            # Clean up
            Path(output_path).unlink(missing_ok=True)
    
    def test_generate_from_file(self, generator, sample_odcs_data):
        """Test generating Excel from JSON file."""
        with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as input_file:
            json.dump(sample_odcs_data, input_file)
            input_path = input_file.name
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = output_file.name
        
        try:
            generator.generate_from_file(input_path, output_path)
            
            # Verify file was created
            assert Path(output_path).exists()
            
            # Load and verify content exists
            workbook = load_workbook(output_path)
            assert len(workbook.sheetnames) > 0
            
        finally:
            # Clean up
            Path(input_path).unlink(missing_ok=True)
            Path(output_path).unlink(missing_ok=True)
    
    @patch("requests.get")
    def test_generate_from_url(self, mock_get, generator, sample_odcs_data):
        """Test generating Excel from URL."""
        # Mock HTTP response
        mock_response = Mock()
        mock_response.json.return_value = sample_odcs_data
        mock_response.raise_for_status.return_value = None
        mock_get.return_value = mock_response
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = output_file.name
        
        try:
            generator.generate_from_url("https://example.com/contract.json", output_path)
            
            # Verify request was made
            mock_get.assert_called_once_with("https://example.com/contract.json", timeout=30)
            
            # Verify file was created
            assert Path(output_path).exists()
            
        finally:
            # Clean up
            Path(output_path).unlink(missing_ok=True)
    
    def test_file_not_found_error(self, generator):
        """Test handling of missing input file."""
        with pytest.raises(FileNotFoundError):
            generator.generate_from_file("nonexistent.json", "output.xlsx")
    
    def test_invalid_data_handling(self, generator):
        """Test handling of invalid ODCS data."""
        invalid_data = {"invalid": "data"}
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = output_file.name
        
        try:
            # Should not raise exception, but handle gracefully
            generator.generate_from_dict(invalid_data, output_path)
            
            # File should still be created
            assert Path(output_path).exists()
            
        finally:
            Path(output_path).unlink(missing_ok=True)
    
    def test_custom_style_config(self):
        """Test generator with custom style configuration."""
        custom_config = {
            "header_font": "Arial",
            "header_fill": "#FF0000"
        }
        
        generator = ODCSExcelGenerator(style_config=custom_config)
        assert generator.style_config["header_font"] == "Arial"
        assert generator.style_config["header_fill"] == "#FF0000"


class TestODCSDataContract:
    """Test cases for ODCSDataContract model."""
    
    def test_valid_contract_creation(self):
        """Test creating valid ODCS contract."""
        contract_data = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "test-123",
            "status": "active"
        }
        
        contract = ODCSDataContract(**contract_data)
        assert contract.version == "1.0.0"
        assert contract.kind == "DataContract"
        assert contract.apiVersion == "v3.0.2"
        assert contract.id == "test-123"
        assert contract.status == "active"
    
    def test_invalid_contract_creation(self):
        """Test validation of invalid contract data."""
        # Missing required fields
        with pytest.raises(Exception):  # ValidationError from pydantic
            ODCSDataContract(version="1.0.0")
        
        # Invalid enum value
        with pytest.raises(Exception):
            ODCSDataContract(
                version="1.0.0",
                kind="InvalidKind",
                apiVersion="v3.0.2",
                id="test-123", 
                status="active"
            )
