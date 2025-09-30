"""Pytest configuration and shared fixtures for ODCS Converter tests."""

import json
import tempfile
from pathlib import Path
from typing import Dict, Any, Generator
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.yaml_converter import YAMLConverter

# Import fixtures from utility modules to make them globally available
# These are re-exported to be available as pytest fixtures across all test modules
__all__ = [
    "unit_test_helper",
    "mock_factory",
    "validation_helper",
    "file_helper",
    "parameterized_test_data",
    "integration_test_helper",
    "excel_test_helper",
    "conversion_test_helper",
    "component_test_helper",
    "workflow_test_helper",
    "complete_odcs_data",
    "multi_table_odcs_data",
    "e2e_test_helper",
    "cli_test_helper",
    "performance_test_helper",
    "scenario_test_helper",
    "error_scenario_test_helper",
    "production_like_odcs",
    "complex_multi_domain_odcs",
]

try:
    from tests.unit.utils import (  # noqa: F401
        unit_test_helper,
        mock_factory,
        validation_helper,
        file_helper,
        parameterized_test_data,
    )
except ImportError:
    pass

try:
    from tests.integration.utils import (  # noqa: F401
        integration_test_helper,
        excel_test_helper,
        conversion_test_helper,
        component_test_helper,
        workflow_test_helper,
        complete_odcs_data,
        multi_table_odcs_data,
    )
except ImportError:
    pass

try:
    from tests.end_to_end.utils import (  # noqa: F401
        e2e_test_helper,
        cli_test_helper,
        performance_test_helper,
        scenario_test_helper,
        error_scenario_test_helper,
        production_like_odcs,
        complex_multi_domain_odcs,
    )
except ImportError:
    pass


# Test directory paths
TESTS_DIR = Path(__file__).parent
TEST_DATA_DIR = TESTS_DIR / "test_data"
FIXTURES_DIR = TESTS_DIR / "fixtures"
OUTPUTS_DIR = TESTS_DIR / "outputs"

# Ensure output directories exist
for output_dir in [
    OUTPUTS_DIR / "unit",
    OUTPUTS_DIR / "integration",
    OUTPUTS_DIR / "end_to_end",
]:
    output_dir.mkdir(parents=True, exist_ok=True)


@pytest.fixture(scope="session")
def test_data_dir() -> Path:
    """Path to test data directory."""
    return TEST_DATA_DIR


@pytest.fixture(scope="session")
def fixtures_dir() -> Path:
    """Path to fixtures directory."""
    return FIXTURES_DIR


@pytest.fixture(scope="session")
def outputs_dir() -> Path:
    """Path to outputs directory."""
    return OUTPUTS_DIR


@pytest.fixture
def temp_dir() -> Generator[Path, None, None]:
    """Create and cleanup temporary directory for tests."""
    with tempfile.TemporaryDirectory() as tmp_dir:
        yield Path(tmp_dir)


@pytest.fixture
def sample_odcs_minimal() -> Dict[str, Any]:
    """Minimal valid ODCS contract for testing."""
    return {
        "version": "1.0.0",
        "kind": "DataContract",
        "apiVersion": "v3.0.2",
        "id": "test-contract-minimal",
        "status": "active",
    }


@pytest.fixture
def sample_odcs_complete() -> Dict[str, Any]:
    """Complete ODCS contract with all fields for testing."""
    return {
        "version": "2.0.0",
        "kind": "DataContract",
        "apiVersion": "v3.0.2",
        "id": "test-contract-complete",
        "name": "Complete Test Contract",
        "tenant": "test-tenant",
        "status": "active",
        "dataProduct": "Test Data Product",
        "domain": "test_domain",
        "contractCreatedTs": "2024-01-15T09:00:00Z",
        "tags": ["test", "complete", "sample"],
        "description": {
            "usage": "Complete test contract for comprehensive testing",
            "purpose": "Testing all ODCS fields and conversion scenarios",
            "limitations": "Test environment only - not for production use",
        },
        "servers": [
            {
                "server": "test-db-primary",
                "type": "postgresql",
                "description": "Primary test database server",
                "environment": "test",
                "host": "test-db.example.com",
                "port": 5432,
                "database": "test_db",
                "schema": "public",
            },
            {
                "server": "test-warehouse",
                "type": "snowflake",
                "description": "Test data warehouse",
                "environment": "test",
                "account": "test-account",
                "database": "TEST_DW",
                "schema": "ANALYTICS",
                "warehouse": "TEST_WH",
            },
        ],
        "schema": [
            {
                "name": "test_table",
                "logicalType": "object",
                "physicalName": "test_table_v1",
                "description": "Test table for validation",
                "businessName": "Test Data Table",
                "dataGranularityDescription": "One record per test entity",
                "properties": [
                    {
                        "name": "id",
                        "logicalType": "integer",
                        "physicalType": "BIGINT",
                        "description": "Unique identifier",
                        "required": True,
                        "primaryKey": True,
                        "primaryKeyPosition": 1,
                    },
                    {
                        "name": "name",
                        "logicalType": "string",
                        "physicalType": "VARCHAR(255)",
                        "description": "Entity name",
                        "required": True,
                    },
                    {
                        "name": "created_at",
                        "logicalType": "date",
                        "physicalType": "TIMESTAMP",
                        "description": "Creation timestamp",
                        "required": True,
                    },
                ],
            }
        ],
        "support": [
            {
                "channel": "test-support",
                "url": "https://support.example.com/test",
                "description": "Test support channel",
                "tool": "web",
                "scope": "issues",
            }
        ],
        "team": [
            {
                "username": "test.user@example.com",
                "name": "Test User",
                "role": "owner",
                "description": "Test contract owner",
            }
        ],
        "roles": [
            {
                "role": "test_reader",
                "description": "Read-only access to test data",
                "access": "SELECT",
            }
        ],
        "slaProperties": [
            {
                "property": "availability",
                "value": 99.5,
                "unit": "percent",
                "driver": "operational",
            }
        ],
        "authoritativeDefinitions": [
            {
                "url": "https://docs.example.com/test-contract",
                "type": "businessDefinition",
            }
        ],
        "customProperties": [
            {"property": "testEnvironment", "value": "integration"},
            {"property": "autoCleanup", "value": True},
        ],
    }


@pytest.fixture
def odcs_to_excel_converter() -> ODCSToExcelConverter:
    """ODCSToExcelConverter instance for testing."""
    return ODCSToExcelConverter()


@pytest.fixture
def excel_to_odcs_parser() -> ExcelToODCSParser:
    """ExcelToODCSParser instance for testing."""
    return ExcelToODCSParser()


@pytest.fixture
def sample_excel_workbook() -> Generator[str, None, None]:
    """Create a sample Excel workbook for testing."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Basic Information sheet
    basic_sheet = wb.create_sheet("Basic Information")
    basic_data = [
        ["Field", "Value", "Description"],
        ["version", "1.0.0", "Contract Version"],
        ["kind", "DataContract", "Contract Kind"],
        ["apiVersion", "v3.0.2", "API Version"],
        ["id", "test-workbook-001", "Unique ID"],
        ["status", "active", "Status"],
    ]
    for row in basic_data:
        basic_sheet.append(row)

    # Tags sheet
    tags_sheet = wb.create_sheet("Tags")
    tags_sheet.append(["Tag"])
    for tag in ["test", "sample", "workbook"]:
        tags_sheet.append([tag])

    # Description sheet
    desc_sheet = wb.create_sheet("Description")
    desc_data = [
        ["Field", "Value"],
        ["usage", "Sample workbook for testing"],
        ["purpose", "Automated test validation"],
    ]
    for row in desc_data:
        desc_sheet.append(row)

    # Save to temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        wb.save(tmp_file.name)
        tmp_path = tmp_file.name

    yield tmp_path

    # Cleanup
    Path(tmp_path).unlink(missing_ok=True)


@pytest.fixture
def mock_requests_get():
    """Mock requests.get for testing URL fetching."""

    def _mock_get(url: str, response_data: Dict[str, Any], status_code: int = 200):
        mock = MagicMock()
        mock.json.return_value = response_data
        mock.raise_for_status.return_value = None
        mock.status_code = status_code
        return mock

    return _mock_get


@pytest.fixture(autouse=True)
def cleanup_test_outputs():
    """Automatically cleanup test outputs after each test."""
    yield  # Run the test

    # Cleanup outputs after test
    for output_dir in [
        OUTPUTS_DIR / "unit",
        OUTPUTS_DIR / "integration",
        OUTPUTS_DIR / "end_to_end",
    ]:
        if output_dir.exists():
            for file in output_dir.glob("test_*"):
                try:
                    file.unlink()
                except (OSError, PermissionError):
                    pass  # Ignore cleanup errors


@pytest.fixture
def json_file_factory(temp_dir):
    """Factory to create temporary JSON files for testing."""

    def _create_json_file(data: Dict[str, Any], filename: str = "test.json") -> Path:
        file_path = temp_dir / filename
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return file_path

    return _create_json_file


@pytest.fixture
def yaml_file_factory(temp_dir):
    """Factory to create temporary YAML files for testing."""

    def _create_yaml_file(data: Dict[str, Any], filename: str = "test.yaml") -> Path:
        file_path = temp_dir / filename
        YAMLConverter.dict_to_yaml(data, file_path)
        return file_path

    return _create_yaml_file


@pytest.fixture
def excel_file_factory(temp_dir):
    """Factory to create temporary Excel files for testing."""

    def _create_excel_file(data: Dict[str, Any], filename: str = "test.xlsx") -> Path:
        file_path = temp_dir / filename
        converter = ODCSToExcelConverter()
        converter.generate_from_dict(data, file_path)
        return file_path

    return _create_excel_file


# Test markers for different test categories
def pytest_configure(config):
    """Register custom pytest markers."""
    config.addinivalue_line(
        "markers", "unit: marks tests as unit tests (fast, isolated)"
    )
    config.addinivalue_line(
        "markers",
        "integration: marks tests as integration tests (medium speed, component interaction)",
    )
    config.addinivalue_line(
        "markers", "e2e: marks tests as end-to-end tests (slower, full workflow)"
    )
    config.addinivalue_line("markers", "slow: marks tests as slow running")


# Custom test result handling
@pytest.fixture(scope="session", autouse=True)
def test_session_setup():
    """Setup and teardown for test session."""
    # Setup: Ensure test directories exist
    for dir_path in [TEST_DATA_DIR, FIXTURES_DIR, OUTPUTS_DIR]:
        dir_path.mkdir(parents=True, exist_ok=True)

    yield

    # Teardown: Optional cleanup of session-level resources
    # (Individual test cleanup is handled by cleanup_test_outputs fixture)


class TestFileManager:
    """Utility class for managing test files and directories."""

    @staticmethod
    def create_test_file(content: str, filename: str, directory: Path) -> Path:
        """Create a test file with given content."""
        directory.mkdir(parents=True, exist_ok=True)
        file_path = directory / filename
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        return file_path

    @staticmethod
    def get_test_output_path(test_name: str, test_type: str, filename: str) -> Path:
        """Get standardized path for test outputs."""
        output_dir = OUTPUTS_DIR / test_type / test_name
        output_dir.mkdir(parents=True, exist_ok=True)
        return output_dir / filename

    @staticmethod
    def cleanup_test_directory(directory: Path) -> None:
        """Clean up all files in a test directory."""
        if directory.exists():
            for file in directory.iterdir():
                if file.is_file():
                    try:
                        file.unlink()
                    except (OSError, PermissionError):
                        pass


@pytest.fixture
def test_file_manager() -> TestFileManager:
    """Test file manager utility."""
    return TestFileManager()
