"""Utilities for end-to-end tests."""

import json
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional, Tuple
from unittest.mock import MagicMock, patch
import pytest
import shlex
import os

from odcs_converter.cli import main as cli_main
from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.yaml_converter import YAMLConverter


class EndToEndTestHelper:
    """Helper class for end-to-end test operations."""

    @staticmethod
    def create_production_like_odcs() -> Dict[str, Any]:
        """Create production-like ODCS data for E2E testing."""
        return {
            "version": "3.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "e2e-production-contract",
            "name": "E2E Production-like Contract",
            "tenant": "production-tenant",
            "status": "active",
            "dataProduct": "Customer Analytics Platform",
            "domain": "analytics",
            "contractCreatedTs": "2024-01-15T08:00:00Z",
            "tags": ["production", "analytics", "customer-data", "gdpr-compliant"],
            "description": {
                "usage": "Customer analytics data for business intelligence and reporting",
                "purpose": "Enable data-driven decision making for customer experience optimization",
                "limitations": "PII data requires special handling. Data retention limited to 7 years."
            },
            "servers": [
                {
                    "server": "prod-analytics-primary",
                    "type": "snowflake",
                    "description": "Primary production analytics warehouse",
                    "environment": "production",
                    "account": "prod-account.snowflakecomputing.com",
                    "database": "ANALYTICS_PROD",
                    "schema": "CUSTOMER_DATA",
                    "warehouse": "ANALYTICS_WH"
                },
                {
                    "server": "prod-replica-read",
                    "type": "postgresql",
                    "description": "Read replica for real-time queries",
                    "environment": "production",
                    "host": "prod-replica.company.com",
                    "port": 5432,
                    "database": "analytics_prod",
                    "schema": "public"
                }
            ],
            "schema": [
                {
                    "name": "customer_profiles",
                    "logicalType": "object",
                    "physicalName": "customer_profiles_v3",
                    "description": "Comprehensive customer profile information",
                    "businessName": "Customer Master Data",
                    "dataGranularityDescription": "One record per unique customer",
                    "properties": [
                        {
                            "name": "customer_id",
                            "logicalType": "integer",
                            "physicalType": "BIGINT",
                            "description": "Unique customer identifier",
                            "required": True,
                            "primaryKey": True,
                            "primaryKeyPosition": 1
                        },
                        {
                            "name": "customer_uuid",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(36)",
                            "description": "Customer UUID for external systems",
                            "required": True
                        },
                        {
                            "name": "email_hash",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(64)",
                            "description": "SHA-256 hash of customer email for privacy",
                            "required": True
                        },
                        {
                            "name": "registration_date",
                            "logicalType": "date",
                            "physicalType": "DATE",
                            "description": "Customer registration date",
                            "required": True
                        },
                        {
                            "name": "last_activity_ts",
                            "logicalType": "timestamp",
                            "physicalType": "TIMESTAMP_NTZ",
                            "description": "Last customer activity timestamp",
                            "required": False
                        },
                        {
                            "name": "is_active",
                            "logicalType": "boolean",
                            "physicalType": "BOOLEAN",
                            "description": "Customer active status",
                            "required": True
                        },
                        {
                            "name": "lifetime_value",
                            "logicalType": "number",
                            "physicalType": "DECIMAL(15,2)",
                            "description": "Customer lifetime value in USD",
                            "required": False
                        },
                        {
                            "name": "segment_data",
                            "logicalType": "object",
                            "physicalType": "VARIANT",
                            "description": "Customer segmentation data (JSON)",
                            "required": False
                        }
                    ]
                },
                {
                    "name": "transaction_events",
                    "logicalType": "object",
                    "physicalName": "transaction_events_v2",
                    "description": "Customer transaction events",
                    "businessName": "Transaction Activity Log",
                    "dataGranularityDescription": "One record per transaction event",
                    "properties": [
                        {
                            "name": "event_id",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(36)",
                            "description": "Unique event identifier",
                            "required": True,
                            "primaryKey": True,
                            "primaryKeyPosition": 1
                        },
                        {
                            "name": "customer_id",
                            "logicalType": "integer",
                            "physicalType": "BIGINT",
                            "description": "Reference to customer",
                            "required": True
                        },
                        {
                            "name": "event_timestamp",
                            "logicalType": "timestamp",
                            "physicalType": "TIMESTAMP_NTZ",
                            "description": "When the event occurred",
                            "required": True
                        },
                        {
                            "name": "event_type",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(50)",
                            "description": "Type of transaction event",
                            "required": True
                        },
                        {
                            "name": "amount",
                            "logicalType": "number",
                            "physicalType": "DECIMAL(12,2)",
                            "description": "Transaction amount",
                            "required": False
                        },
                        {
                            "name": "currency_code",
                            "logicalType": "string",
                            "physicalType": "CHAR(3)",
                            "description": "ISO currency code",
                            "required": False
                        }
                    ]
                }
            ],
            "support": [
                {
                    "channel": "data-platform-support",
                    "url": "https://company.slack.com/channels/data-platform",
                    "description": "Primary support channel for data platform issues",
                    "tool": "slack",
                    "scope": "issues"
                },
                {
                    "channel": "analytics-docs",
                    "url": "https://docs.company.com/analytics/customer-data",
                    "description": "Documentation for customer analytics data",
                    "tool": "confluence",
                    "scope": "documentation"
                }
            ],
            "team": [
                {
                    "username": "data.team@company.com",
                    "name": "Data Platform Team",
                    "role": "owner",
                    "description": "Responsible for data platform and contract maintenance"
                },
                {
                    "username": "analytics.team@company.com",
                    "name": "Analytics Team",
                    "role": "consumer",
                    "description": "Primary consumer of customer analytics data"
                }
            ],
            "roles": [
                {
                    "role": "analytics_reader",
                    "description": "Read access to analytics data",
                    "access": "SELECT"
                },
                {
                    "role": "analytics_writer",
                    "description": "Write access for ETL processes",
                    "access": "INSERT, UPDATE, DELETE"
                }
            ],
            "slaProperties": [
                {
                    "property": "availability",
                    "value": 99.9,
                    "unit": "percent",
                    "driver": "operational"
                },
                {
                    "property": "freshness",
                    "value": 15,
                    "unit": "minutes",
                    "driver": "operational"
                },
                {
                    "property": "retention",
                    "value": 7,
                    "unit": "years",
                    "driver": "compliance"
                }
            ],
            "authoritativeDefinitions": [
                {
                    "url": "https://docs.company.com/data-dictionary/customer-profiles",
                    "type": "businessDefinition"
                },
                {
                    "url": "https://github.com/company/data-contracts/customer-analytics",
                    "type": "technicalDefinition"
                }
            ],
            "customProperties": [
                {
                    "property": "gdprCompliant",
                    "value": True
                },
                {
                    "property": "dataClassification",
                    "value": "confidential"
                },
                {
                    "property": "backupFrequency",
                    "value": "daily"
                },
                {
                    "property": "monitoringEnabled",
                    "value": True
                }
            ]
        }

    @staticmethod
    def create_complex_multi_domain_odcs() -> Dict[str, Any]:
        """Create complex multi-domain ODCS for comprehensive E2E testing."""
        base = EndToEndTestHelper.create_production_like_odcs()
        base.update({
            "id": "e2e-multi-domain-contract",
            "name": "Multi-Domain Analytics Contract",
            "domain": "cross_functional",
            "servers": [
                {
                    "server": "finance-warehouse",
                    "type": "bigquery",
                    "description": "Finance data warehouse",
                    "environment": "production",
                    "project": "company-finance-prod",
                    "dataset": "financial_data"
                },
                {
                    "server": "marketing-lake",
                    "type": "databricks",
                    "description": "Marketing data lake",
                    "environment": "production",
                    "workspace": "marketing-prod.databricks.com",
                    "catalog": "marketing",
                    "schema": "campaigns"
                },
                {
                    "server": "operations-db",
                    "type": "mongodb",
                    "description": "Operations document database",
                    "environment": "production",
                    "host": "ops-cluster.company.com",
                    "database": "operations",
                    "collection": "events"
                }
            ]
        })
        return base


class CLITestHelper:
    """Helper for testing CLI functionality in E2E scenarios."""

    @staticmethod
    def run_cli_command(
        command: List[str],
        input_data: Optional[str] = None,
        timeout: int = 30
    ) -> Tuple[bool, str, str]:
        """Run CLI command and return success status, stdout, stderr."""
        try:
            # Prepare the command with proper Python path
            full_command = ["python", "-m", "odcs_converter.cli"] + command[1:]  # Skip the first 'odcs-converter'

            process = subprocess.Popen(
                full_command,
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=timeout
            )

            stdout, stderr = process.communicate(input=input_data)
            success = process.returncode == 0

            return success, stdout, stderr

        except subprocess.TimeoutExpired:
            return False, "", "Command timed out"
        except Exception as e:
            return False, "", str(e)

    @staticmethod
    def test_cli_help() -> Tuple[bool, str]:
        """Test CLI help functionality."""
        success, stdout, stderr = CLITestHelper.run_cli_command(["odcs-converter", "--help"])
        return success, stdout if success else stderr

    @staticmethod
    def test_cli_version() -> Tuple[bool, str]:
        """Test CLI version functionality."""
        success, stdout, stderr = CLITestHelper.run_cli_command(["odcs-converter", "--version"])
        return success, stdout if success else stderr

    @staticmethod
    def test_odcs_to_excel_cli(
        input_file: Path,
        output_file: Path
    ) -> Tuple[bool, str, str]:
        """Test ODCS to Excel conversion via CLI."""
        command = [
            "odcs-converter",
            "to-excel",
            str(input_file),
            str(output_file)
        ]
        return CLITestHelper.run_cli_command(command)

    @staticmethod
    def test_excel_to_odcs_cli(
        input_file: Path,
        output_file: Path,
        output_format: str = "json"
    ) -> Tuple[bool, str, str]:
        """Test Excel to ODCS conversion via CLI."""
        command = [
            "odcs-converter",
            "to-odcs",
            str(input_file),
            str(output_file),
            "--format",
            output_format
        ]
        return CLITestHelper.run_cli_command(command)


class PerformanceTestHelper:
    """Helper for performance testing in E2E scenarios."""

    @staticmethod
    def measure_conversion_time(
        conversion_func,
        *args,
        **kwargs
    ) -> Tuple[float, Any]:
        """Measure time taken for a conversion operation."""
        start_time = time.time()
        result = conversion_func(*args, **kwargs)
        end_time = time.time()
        return end_time - start_time, result

    @staticmethod
    def test_large_dataset_performance(
        base_odcs: Dict[str, Any],
        table_count: int = 10,
        properties_per_table: int = 50
    ) -> Dict[str, float]:
        """Test performance with large datasets."""
        # Create large ODCS data
        large_odcs = base_odcs.copy()
        large_odcs["schema"] = []

        for table_idx in range(table_count):
            table = {
                "name": f"large_table_{table_idx}",
                "logicalType": "object",
                "physicalName": f"large_table_{table_idx}_v1",
                "description": f"Large test table {table_idx}",
                "properties": []
            }

            for prop_idx in range(properties_per_table):
                prop = {
                    "name": f"column_{prop_idx}",
                    "logicalType": "string",
                    "physicalType": "VARCHAR(255)",
                    "description": f"Column {prop_idx} description",
                    "required": prop_idx < 5  # First 5 are required
                }
                table["properties"].append(prop)

            large_odcs["schema"].append(table)

        results = {}

        # Test ODCS to Excel conversion
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_excel:
            excel_path = Path(tmp_excel.name)

        try:
            converter = ODCSToExcelConverter()
            excel_time, _ = PerformanceTestHelper.measure_conversion_time(
                converter.generate_from_dict,
                large_odcs,
                excel_path
            )
            results["odcs_to_excel"] = excel_time

            # Test Excel to ODCS conversion
            parser = ExcelToODCSParser()
            odcs_time, _ = PerformanceTestHelper.measure_conversion_time(
                parser.parse_from_file,
                excel_path
            )
            results["excel_to_odcs"] = odcs_time

        finally:
            excel_path.unlink(missing_ok=True)

        return results

    @staticmethod
    def memory_usage_test(conversion_func, *args, **kwargs) -> Tuple[Any, Dict[str, float]]:
        """Test memory usage during conversion (requires psutil)."""
        try:
            import psutil
            process = psutil.Process()

            # Measure initial memory
            initial_memory = process.memory_info().rss / 1024 / 1024  # MB

            # Run conversion
            result = conversion_func(*args, **kwargs)

            # Measure peak memory
            peak_memory = process.memory_info().rss / 1024 / 1024  # MB

            return result, {
                "initial_memory_mb": initial_memory,
                "peak_memory_mb": peak_memory,
                "memory_increase_mb": peak_memory - initial_memory
            }

        except ImportError:
            # psutil not available, just run the function
            result = conversion_func(*args, **kwargs)
            return result, {"error": "psutil not available for memory measurement"}


class ScenarioTestHelper:
    """Helper for testing real-world scenarios."""

    @staticmethod
    def simulate_data_engineer_workflow(temp_dir: Path) -> Tuple[bool, Dict[str, Any], List[str]]:
        """Simulate a data engineer's typical workflow."""
        results = {"steps_completed": [], "files_generated": []}
        errors = []

        try:
            # Step 1: Data engineer creates ODCS contract
            odcs_data = EndToEndTestHelper.create_production_like_odcs()
            json_path = temp_dir / "contract.json"

            with open(json_path, 'w') as f:
                json.dump(odcs_data, f, indent=2)

            results["steps_completed"].append("Created ODCS JSON contract")
            results["files_generated"].append(str(json_path))

            # Step 2: Convert to Excel for business stakeholder review
            excel_path = temp_dir / "contract_for_review.xlsx"
            success, stdout, stderr = CLITestHelper.test_odcs_to_excel_cli(json_path, excel_path)

            if not success:
                errors.append(f"CLI conversion failed: {stderr}")
                return False, results, errors

            results["steps_completed"].append("Converted to Excel via CLI")
            results["files_generated"].append(str(excel_path))

            # Step 3: Simulate business user making changes (modify Excel)
            # This would typically be done manually, but we'll simulate programmatically
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)

            # Add a comment to description
            if "Description" in wb.sheetnames:
                desc_sheet = wb["Description"]
                # Find usage row and modify it
                for row in desc_sheet.iter_rows():
                    if row[0].value == "usage":
                        row[1].value = f"{row[1].value} [REVIEWED BY BUSINESS]"
                        break

            wb.save(excel_path)
            results["steps_completed"].append("Business user reviewed and modified Excel")

            # Step 4: Convert back to ODCS JSON
            updated_json_path = temp_dir / "contract_updated.json"
            success, stdout, stderr = CLITestHelper.test_excel_to_odcs_cli(
                excel_path, updated_json_path, "json"
            )

            if not success:
                errors.append(f"CLI conversion back failed: {stderr}")
                return False, results, errors

            results["steps_completed"].append("Converted updated Excel back to JSON")
            results["files_generated"].append(str(updated_json_path))

            # Step 5: Verify the changes were preserved
            with open(updated_json_path, 'r') as f:
                updated_data = json.load(f)

            if "[REVIEWED BY BUSINESS]" in updated_data.get("description", {}).get("usage", ""):
                results["steps_completed"].append("Business changes preserved in roundtrip")
            else:
                errors.append("Business changes were lost in roundtrip conversion")

            return len(errors) == 0, results, errors

        except Exception as e:
            errors.append(f"Workflow error: {str(e)}")
            return False, results, errors

    @staticmethod
    def simulate_compliance_review_workflow(temp_dir: Path) -> Tuple[bool, Dict[str, Any], List[str]]:
        """Simulate a compliance team reviewing data contracts."""
        results = {"compliance_checks": [], "issues_found": []}
        errors = []

        try:
            # Create a contract with compliance-sensitive data
            odcs_data = EndToEndTestHelper.create_production_like_odcs()

            # Step 1: Convert to Excel for compliance review
            excel_path = temp_dir / "compliance_review.xlsx"
            converter = ODCSToExcelConverter()
            converter.generate_from_dict(odcs_data, excel_path)

            results["compliance_checks"].append("Contract exported to Excel for review")

            # Step 2: Simulate compliance checks
            parser = ExcelToODCSParser()
            parsed_data = parser.parse_from_file(excel_path)

            # Check for GDPR compliance indicators
            custom_props = parsed_data.get("customProperties", [])
            gdpr_compliant = any(
                prop.get("property") == "gdprCompliant" and prop.get("value") is True
                for prop in custom_props
            )

            if gdpr_compliant:
                results["compliance_checks"].append("GDPR compliance flag verified")
            else:
                results["issues_found"].append("Missing GDPR compliance indication")

            # Check for data classification
            data_classified = any(
                prop.get("property") == "dataClassification"
                for prop in custom_props
            )

            if data_classified:
                results["compliance_checks"].append("Data classification present")
            else:
                results["issues_found"].append("Missing data classification")

            # Check for PII handling in schema
            schema_objects = parsed_data.get("schema", [])
            pii_fields_documented = False

            for schema_obj in schema_objects:
                properties = schema_obj.get("properties", [])
                for prop in properties:
                    if "hash" in prop.get("name", "").lower():
                        pii_fields_documented = True
                        break

            if pii_fields_documented:
                results["compliance_checks"].append("PII handling documented in schema")
            else:
                results["issues_found"].append("No evidence of PII handling")

            return len(results["issues_found"]) == 0, results, errors

        except Exception as e:
            errors.append(f"Compliance workflow error: {str(e)}")
            return False, results, errors


class ErrorScenarioTestHelper:
    """Helper for testing error scenarios and edge cases."""

    @staticmethod
    def test_file_system_errors(temp_dir: Path) -> List[str]:
        """Test various file system error scenarios."""
        errors_handled = []

        # Test 1: Non-existent input file
        try:
            parser = ExcelToODCSParser()
            parser.parse_from_file(temp_dir / "nonexistent.xlsx")
        except Exception:
            errors_handled.append("Non-existent file error handled")

        # Test 2: Invalid output directory
        try:
            converter = ODCSToExcelConverter()
            invalid_path = Path("/invalid/directory/output.xlsx")
            converter.generate_from_dict({"test": "data"}, invalid_path)
        except Exception:
            errors_handled.append("Invalid output directory error handled")

        # Test 3: Permission denied (simulate by trying to write to read-only location)
        try:
            readonly_path = temp_dir / "readonly.xlsx"
            readonly_path.touch()
            readonly_path.chmod(0o444)  # Read-only

            converter = ODCSToExcelConverter()
            converter.generate_from_dict({"test": "data"}, readonly_path)
        except Exception:
            errors_handled.append("Permission denied error handled")
        finally:
            if readonly_path.exists():
                readonly_path.chmod(0o644)  # Restore permissions for cleanup

        return errors_handled

    @staticmethod
    def test_data_corruption_scenarios(temp_dir: Path) -> List[str]:
        """Test handling of corrupted or malformed data."""
        scenarios_tested = []

        # Test 1: Corrupted Excel file
        corrupted_excel = temp_dir / "corrupted.xlsx"
        with open(corrupted_excel, 'wb') as f:
            f.write(b"This is not an Excel file")

        try:
            parser = ExcelToODCSParser()
            parser.parse_from_file(corrupted_excel)
        except Exception:
            scenarios_tested.append("Corrupted Excel file handled")

        # Test 2: Malformed JSON data
        try:
            converter = ODCSToExcelConverter()
            malformed_data = {"version": None, "invalid": float('inf')}
            excel_path = temp_dir / "from_malformed.xlsx"
            converter.generate_from_dict(malformed_data, excel_path)
        except Exception:
            scenarios_tested.append("Malformed JSON data handled")

        # Test 3: Extremely large data
        try:
            huge_data = EndToEndTestHelper.create_production_like_odcs()
            # Add many properties to stress test
            for schema_obj in huge_data.get("schema", []):
                props = schema_obj.get("properties", [])
                for i in range(1000):  # Add 1000 properties
                    props.append({
                        "name": f"prop_{i}",
                        "logicalType": "string",
                        "description": "A" * 1000  # Long description
                    })

            converter = ODCSToExcelConverter()
            excel_path = temp_dir / "huge_data.xlsx"
            converter.generate_from_dict(huge_data, excel_path)
            scenarios_tested.append("Large dataset handled")

        except Exception:
            scenarios_tested.append("Large dataset error handled gracefully")

        return scenarios_tested


# Pytest fixtures specific to end-to-end tests
@pytest.fixture
def e2e_test_helper():
    """Provide E2E test helper instance."""
    return EndToEndTestHelper()


@pytest.fixture
def cli_test_helper():
    """Provide CLI test helper instance."""
    return CLITestHelper()


@pytest.fixture
def performance_test_helper():
    """Provide performance test helper instance."""
    return PerformanceTestHelper()


@pytest.fixture
def scenario_test_helper():
    """Provide scenario test helper instance."""
    return ScenarioTestHelper()


@pytest.fixture
def error_scenario_test_helper():
    """Provide error scenario test helper instance."""
    return ErrorScenarioTestHelper()


@pytest.fixture
def production_like_odcs():
    """Provide production-like ODCS data."""
    return EndToEndTestHelper.create_production_like_odcs()


@pytest.fixture
def complex_multi_domain_odcs():
    """Provide complex multi-domain ODCS data."""
    return EndToEndTestHelper.create_complex_multi_domain_odcs()


# Custom decorators for end-to-end tests
def e2e_test(func):
    """Decorator to mark a function as an end-to-end test."""
    return pytest.mark.e2e(func)


def slow_e2e_test(func):
    """Decorator to mark a function as a slow end-to-end test."""
    return pytest.mark.e2e(pytest.mark.slow(func))


def cli_test(func):
    """Decorator to mark a function as a CLI test."""
    return pytest.mark.e2e(pytest.mark.cli(func))


def performance_test(func):
    """Decorator to mark a function as a performance test."""
    return pytest.mark.e2e(pytest.mark.performance(func))


def requires_cli_setup(func):
    """Decorator for tests that require CLI setup."""
    def wrapper(*args, **kwargs):
        # Ensure CLI is available in PATH or set up environment
        return func(*args, **kwargs)
    return wrapper
