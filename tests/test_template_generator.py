"""Tests for template generator functionality."""

import pytest
from pathlib import Path
from openpyxl import load_workbook

from odcs_converter.template_generator import TemplateGenerator, TemplateType


class TestTemplateGenerator:
    """Test suite for TemplateGenerator class."""

    @pytest.fixture
    def generator(self):
        """Create a TemplateGenerator instance."""
        return TemplateGenerator()

    @pytest.fixture
    def temp_output_path(self, tmp_path):
        """Create a temporary output path."""
        return tmp_path / "test_template.xlsx"

    def test_generator_initialization(self, generator):
        """Test that generator initializes correctly."""
        assert generator is not None
        assert generator.style_config is not None
        assert "header_font" in generator.style_config
        assert "required_header_fill" in generator.style_config
        assert "optional_header_fill" in generator.style_config

    def test_minimal_template_generation(self, generator, temp_output_path):
        """Test minimal template generation."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        assert temp_output_path.exists()

        # Load and verify workbook
        wb = load_workbook(temp_output_path)

        # Check that Instructions sheet exists
        assert "📖 Instructions" in wb.sheetnames

        # Check minimal template sheets
        assert "Basic Information" in wb.sheetnames
        assert "Schema" in wb.sheetnames
        assert "Schema Properties" in wb.sheetnames

        # Verify worksheet count (3 data + 1 instructions)
        assert len(wb.sheetnames) == 4

        wb.close()

    def test_required_template_generation(self, generator, temp_output_path):
        """Test required template generation."""
        generator.generate_template(
            temp_output_path,
            template_type=TemplateType.REQUIRED,
            include_examples=True,
        )

        assert temp_output_path.exists()

        wb = load_workbook(temp_output_path)

        # Check required template sheets
        assert "📖 Instructions" in wb.sheetnames
        assert "Basic Information" in wb.sheetnames
        assert "Servers" in wb.sheetnames
        assert "Schema" in wb.sheetnames
        assert "Schema Properties" in wb.sheetnames

        # Verify worksheet count (4 data + 1 instructions)
        assert len(wb.sheetnames) == 5

        wb.close()

    def test_full_template_generation(self, generator, temp_output_path):
        """Test full template generation."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        assert temp_output_path.exists()

        wb = load_workbook(temp_output_path)

        # Check that all expected sheets exist
        expected_sheets = [
            "📖 Instructions",
            "Basic Information",
            "Tags",
            "Description",
            "Servers",
            "Schema",
            "Schema Properties",
            "Logical Type Options",
            "Quality Rules",
            "Support",
            "Pricing",
            "Team",
            "Roles",
            "SLA Properties",
            "Authoritative Definitions",
            "Custom Properties",
        ]

        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

        # Verify total worksheet count (15 data + 1 instructions)
        assert len(wb.sheetnames) == 16

        wb.close()

    def test_template_without_examples(self, generator, temp_output_path):
        """Test template generation without examples."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=False
        )

        assert temp_output_path.exists()

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # Check that only headers exist (row 1)
        # No example rows should be present (row 2 should be empty)
        assert sheet.max_row == 1  # Only header row

        wb.close()

    def test_template_with_examples(self, generator, temp_output_path):
        """Test template generation with examples."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        assert temp_output_path.exists()

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # Check that headers and examples exist
        assert sheet.max_row >= 2  # Header + at least one example row

        # Verify example data exists in row 2
        assert sheet.cell(row=2, column=1).value is not None

        wb.close()

    def test_instructions_sheet_content(self, generator, temp_output_path):
        """Test that instructions sheet has proper content."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        instructions = wb["📖 Instructions"]

        # Check for key instruction elements
        cell_values = []
        for row in instructions.iter_rows(values_only=True):
            cell_values.extend([str(cell) for cell in row if cell])

        content = " ".join(cell_values)

        # Verify key instruction content
        assert "ODCS Data Contract" in content
        assert "Red Headers" in content or "RED" in content
        assert "Blue Headers" in content or "BLUE" in content
        assert "REQUIRED" in content
        assert "OPTIONAL" in content

        wb.close()

    def test_header_styling(self, generator, temp_output_path):
        """Test that headers have correct styling."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # Check header row styling
        for col in range(1, 6):  # First 5 columns are required in minimal
            cell = sheet.cell(row=1, column=col)

            # Check font is bold
            assert cell.font.bold is True

            # Check fill color (required fields should have red fill)
            if col <= 5:  # All basic info fields in minimal are required
                assert cell.fill.start_color.index is not None

        wb.close()

    def test_required_vs_optional_headers(self, generator, temp_output_path):
        """Test that required and optional headers are styled differently."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # In full template, first 5 are required (red), rest are optional (blue)
        required_color = sheet.cell(row=1, column=1).fill.start_color.index
        optional_color = sheet.cell(row=1, column=6).fill.start_color.index

        # Colors should be different
        assert required_color != optional_color

        wb.close()

    def test_cell_comments(self, generator, temp_output_path):
        """Test that headers have help text comments."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Schema Properties"]

        # Check that at least some headers have comments
        comment_count = 0
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            if cell.comment:
                comment_count += 1

        # At least some headers should have comments
        assert comment_count > 0

        wb.close()

    def test_schema_properties_structure(self, generator, temp_output_path):
        """Test Schema Properties worksheet structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Schema Properties"]

        # Check required headers exist
        headers = [sheet.cell(row=1, column=col).value for col in range(1, 18)]

        assert "schemaName" in headers
        assert "name" in headers
        assert "logicalType" in headers

        wb.close()

    def test_servers_sheet_structure(self, generator, temp_output_path):
        """Test Servers worksheet structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Servers"]

        # Check required headers
        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "server" in headers
        assert "type" in headers

        wb.close()

    def test_quality_rules_sheet(self, generator, temp_output_path):
        """Test Quality Rules worksheet exists and has proper structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Quality Rules"]

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "schemaName" in headers
        assert "name" in headers
        assert "dimension" in headers

        wb.close()

    def test_example_values_format(self, generator, temp_output_path):
        """Test that example values have correct styling."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # Example row should have italic gray font
        example_cell = sheet.cell(row=2, column=1)

        assert example_cell.font.italic is True
        assert example_cell.font.color.index is not None  # Has color

        wb.close()

    def test_output_directory_creation(self, generator, tmp_path):
        """Test that output directory is created if it doesn't exist."""
        nested_path = tmp_path / "nested" / "directories" / "template.xlsx"

        generator.generate_template(
            nested_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        assert nested_path.exists()
        assert nested_path.parent.exists()

    def test_template_type_enum(self):
        """Test TemplateType enum values."""
        assert TemplateType.MINIMAL.value == "minimal"
        assert TemplateType.REQUIRED.value == "required"
        assert TemplateType.FULL.value == "full"

    def test_column_width_adjustment(self, generator, temp_output_path):
        """Test that column widths are adjusted appropriately."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Basic Information"]

        # Check that columns have reasonable widths
        for col_letter in ["A", "B", "C", "D", "E"]:
            width = sheet.column_dimensions[col_letter].width
            assert width >= 12  # Minimum width
            assert width <= 50  # Maximum width

        wb.close()

    def test_tags_sheet_in_full_template(self, generator, temp_output_path):
        """Test Tags sheet in full template."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Tags"]

        # Check header
        assert sheet.cell(row=1, column=1).value == "tag"

        # If examples are included, check for example tags
        if sheet.max_row > 1:
            assert sheet.cell(row=2, column=1).value is not None

        wb.close()

    def test_minimal_template_excludes_optional_sheets(
        self, generator, temp_output_path
    ):
        """Test that minimal template doesn't include optional sheets."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.MINIMAL, include_examples=True
        )

        wb = load_workbook(temp_output_path)

        # These sheets should NOT be in minimal template
        optional_sheets = [
            "Tags",
            "Description",
            "Quality Rules",
            "Support",
            "Pricing",
            "Team",
            "Roles",
        ]

        for sheet in optional_sheets:
            assert sheet not in wb.sheetnames

        wb.close()

    def test_description_sheet_in_full_template(self, generator, temp_output_path):
        """Test Description sheet structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Description"]

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "usage" in headers
        assert "purpose" in headers
        assert "limitations" in headers

        wb.close()

    def test_team_sheet_structure(self, generator, temp_output_path):
        """Test Team sheet structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Team"]

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "username" in headers
        assert "name" in headers
        assert "role" in headers

        wb.close()

    def test_roles_sheet_structure(self, generator, temp_output_path):
        """Test Roles sheet structure."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["Roles"]

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "role" in headers
        assert "access" in headers

        wb.close()

    def test_sla_properties_sheet(self, generator, temp_output_path):
        """Test SLA Properties sheet."""
        generator.generate_template(
            temp_output_path, template_type=TemplateType.FULL, include_examples=True
        )

        wb = load_workbook(temp_output_path)
        sheet = wb["SLA Properties"]

        headers = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]

        assert "property" in headers
        assert "value" in headers
        assert "unit" in headers

        wb.close()
