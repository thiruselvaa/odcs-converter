"""Test cases specifically for Excel to ODCS parsing functionality."""

import json
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import Workbook

from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.generator import ODCSToExcelConverter


class TestExcelToODCSParser:
    """Test cases for Excel to ODCS parsing."""

    @pytest.fixture
    def parser(self):
        """Create ExcelToODCSParser instance."""
        return ExcelToODCSParser()

    @pytest.fixture
    def sample_excel_file(self):
        """Create a sample Excel file for testing."""
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Basic Information sheet
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "1.0.0", "Contract Version"],
            ["kind", "DataContract", "Contract Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "test-001", "Unique Identifier"],
            ["status", "active", "Status"],
            ["name", "Test Contract", "Contract Name"]
        ]
        for row in basic_data:
            basic_sheet.append(row)

        # Tags sheet
        tags_sheet = wb.create_sheet("Tags")
        tags_sheet.append(["Tag"])
        tags_sheet.append(["test"])
        tags_sheet.append(["example"])
        tags_sheet.append(["parser"])

        # Description sheet
        desc_sheet = wb.create_sheet("Description")
        desc_data = [
            ["Field", "Value"],
            ["usage", "Test usage description"],
            ["purpose", "Testing Excel parser"],
            ["limitations", "Test environment only"]
        ]
        for row in desc_data:
            desc_sheet.append(row)

        # Servers sheet
        servers_sheet = wb.create_sheet("Servers")
        servers_data = [
            ["Server", "Type", "Description", "Environment", "Host", "Port", "Database"],
            ["test-server", "postgresql", "Test database", "test", "localhost", "5432", "testdb"],
            ["prod-server", "snowflake", "Production database", "prod", "prod.snowflake.com", "", "PROD_DB"]
        ]
        for row in servers_data:
            servers_sheet.append(row)

        # Team sheet
        team_sheet = wb.create_sheet("Team")
        team_data = [
            ["Username", "Name", "Role", "Description"],
            ["test@example.com", "Test User", "owner", "Test team member"],
            ["dev@example.com", "Developer", "contributor", "Development team"]
        ]
        for row in team_data:
            team_sheet.append(row)

        # Custom Properties sheet
        props_sheet = wb.create_sheet("Custom Properties")
        props_data = [
            ["Property", "Value"],
            ["environment", "test"],
            ["dataRetentionDays", "365"],
            ["isActive", "true"]
        ]
        for row in props_data:
            props_sheet.append(row)

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        wb.close()

        return temp_file.name

    def test_parse_from_file_success(self, parser, sample_excel_file):
        """Test successful parsing of Excel file."""
        try:
            result = parser.parse_from_file(sample_excel_file)

            # Verify basic structure
            assert isinstance(result, dict)
            assert result.get("version") == "1.0.0"
            assert result.get("kind") == "DataContract"
            assert result.get("apiVersion") == "v3.0.2"
            assert result.get("id") == "test-001"
            assert result.get("status") == "active"
            assert result.get("name") == "Test Contract"

            # Verify tags
            assert "tags" in result
            assert isinstance(result["tags"], list)
            assert "test" in result["tags"]
            assert "example" in result["tags"]
            assert "parser" in result["tags"]

            # Verify description
            assert "description" in result
            desc = result["description"]
            assert desc.get("usage") == "Test usage description"
            assert desc.get("purpose") == "Testing Excel parser"
            assert desc.get("limitations") == "Test environment only"

            # Verify servers
            assert "servers" in result
            assert isinstance(result["servers"], list)
            assert len(result["servers"]) == 2

            first_server = result["servers"][0]
            assert first_server.get("server") == "test-server"
            assert first_server.get("type") == "postgresql"
            assert first_server.get("port") == 5432  # Should be converted to int

            # Verify team
            assert "team" in result
            assert isinstance(result["team"], list)
            assert len(result["team"]) == 2

            # Verify custom properties
            assert "customProperties" in result
            assert isinstance(result["customProperties"], list)

            # Find specific custom properties
            env_prop = next((p for p in result["customProperties"] if p.get("property") == "environment"), None)
            assert env_prop is not None
            assert env_prop.get("value") == "test"

            retention_prop = next((p for p in result["customProperties"] if p.get("property") == "dataRetentionDays"), None)
            assert retention_prop is not None
            assert retention_prop.get("value") == 365  # Should be converted to int

            active_prop = next((p for p in result["customProperties"] if p.get("property") == "isActive"), None)
            assert active_prop is not None
            assert active_prop.get("value") == True  # Should be converted to boolean

        finally:
            Path(sample_excel_file).unlink(missing_ok=True)

    def test_parse_file_not_found(self, parser):
        """Test handling of non-existent Excel file."""
        with pytest.raises(FileNotFoundError):
            parser.parse_from_file("nonexistent_file.xlsx")

    def test_parse_invalid_excel_file(self, parser):
        """Test handling of invalid Excel file."""
        # Create a text file with .xlsx extension
        with tempfile.NamedTemporaryFile(mode='w', suffix='.xlsx', delete=False) as temp_file:
            temp_file.write("This is not an Excel file")
            temp_file_path = temp_file.name

        try:
            with pytest.raises(ValueError, match="Invalid Excel file format"):
                parser.parse_from_file(temp_file_path)
        finally:
            Path(temp_file_path).unlink(missing_ok=True)

    def test_convert_value_types(self, parser):
        """Test value type conversion functionality."""
        # Test boolean conversion
        assert parser._convert_value("true") == True
        assert parser._convert_value("True") == True
        assert parser._convert_value("FALSE") == False
        assert parser._convert_value("false") == False

        # Test integer conversion
        assert parser._convert_value("123") == 123
        assert parser._convert_value("0") == 0

        # Test float conversion
        assert parser._convert_value("123.45") == 123.45
        assert parser._convert_value("0.0") == 0.0

        # Test string values
        assert parser._convert_value("hello") == "hello"
        assert parser._convert_value("test string") == "test string"

        # Test None and empty values
        assert parser._convert_value(None) is None
        assert parser._convert_value("") is None
        assert parser._convert_value("   ") == ""  # Whitespace gets stripped to empty

        # Test existing types (should return as-is)
        assert parser._convert_value(True) == True
        assert parser._convert_value(42) == 42
        assert parser._convert_value(3.14) == 3.14

    def test_clean_data_removes_empty_values(self, parser):
        """Test data cleaning removes None and empty values."""
        dirty_data = {
            "valid_string": "keep this",
            "empty_string": "",
            "none_value": None,
            "valid_number": 42,
            "zero_number": 0,  # Should be kept
            "false_boolean": False,  # Should be kept
            "empty_dict": {},
            "valid_dict": {
                "nested_valid": "keep",
                "nested_empty": "",
                "nested_none": None
            },
            "empty_list": [],
            "valid_list": ["keep1", "", None, "keep2"],
            "mixed_list": [
                {"valid": "data", "empty": ""},
                {"all_empty": "", "also_empty": None},
                "string_item"
            ]
        }

        cleaned = parser._clean_data(dirty_data)

        expected = {
            "valid_string": "keep this",
            "valid_number": 42,
            "zero_number": 0,
            "false_boolean": False,
            "valid_dict": {
                "nested_valid": "keep"
            },
            "valid_list": ["keep1", "keep2"],
            "mixed_list": [
                {"valid": "data"},
                "string_item"
            ]
        }

        assert cleaned == expected

    def test_parse_empty_worksheets(self, parser):
        """Test parsing behavior with empty worksheets."""
        # Mock empty worksheets
        parser.worksheets = {
            "Basic Information": pd.DataFrame(),
            "Tags": pd.DataFrame(),
            "Empty Sheet": pd.DataFrame()
        }

        result = parser._parse_basic_information()
        assert result == {}

        result = parser._parse_tags()
        assert result == {}

    def test_parse_missing_worksheets(self, parser):
        """Test parsing behavior when expected worksheets are missing."""
        parser.worksheets = {}  # No worksheets

        # All parsing methods should return empty dicts when worksheets are missing
        assert parser._parse_basic_information() == {}
        assert parser._parse_tags() == {}
        assert parser._parse_description() == {}
        assert parser._parse_servers() == {}
        assert parser._parse_schema() == {}
        assert parser._parse_support() == {}
        assert parser._parse_pricing() == {}
        assert parser._parse_team() == {}
        assert parser._parse_roles() == {}
        assert parser._parse_sla_properties() == {}
        assert parser._parse_authoritative_definitions() == {}
        assert parser._parse_custom_properties() == {}

    @patch('pandas.read_excel')
    def test_load_worksheets_with_error(self, mock_read_excel, parser):
        """Test worksheet loading with pandas errors."""
        # Mock workbook
        mock_workbook = MagicMock()
        mock_workbook.sheetnames = ["Sheet1", "Sheet2"]
        parser.workbook = mock_workbook

        # Mock pandas to raise an error for one sheet
        def side_effect(filename, sheet_name, **kwargs):
            if sheet_name == "Sheet1":
                return pd.DataFrame({"A": [1, 2, 3]})
            else:
                raise Exception("Cannot read sheet")

        mock_read_excel.side_effect = side_effect

        # Should handle the error gracefully
        parser._load_all_worksheets()

        # Should have loaded Sheet1, Sheet2 might be empty but still loaded
        assert "Sheet1" in parser.worksheets
        # Note: Sheet2 might still be loaded as empty DataFrame even with error

    def test_validate_odcs_data_valid(self, parser):
        """Test ODCS validation with valid data."""
        valid_data = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "test-123",
            "status": "active"
        }

        assert parser.validate_odcs_data(valid_data) == True

    def test_validate_odcs_data_invalid(self, parser):
        """Test ODCS validation with invalid data."""
        invalid_data = {
            "version": "1.0.0",
            # Missing required fields
        }

        assert parser.validate_odcs_data(invalid_data) == False

    def test_parse_servers_with_port_conversion(self, parser):
        """Test server parsing with port number conversion."""
        # Mock DataFrame with server data
        servers_df = pd.DataFrame([
            {
                "Server": "db1",
                "Type": "postgresql",
                "Port": "5432",
                "Host": "localhost",
                "Database": "mydb"
            },
            {
                "Server": "db2",
                "Type": "mysql",
                "Port": "invalid_port",  # Should not crash
                "Host": "remote",
                "Database": "otherdb"
            }
        ])

        parser.worksheets = {"Servers": servers_df}

        result = parser._parse_servers()

        assert "servers" in result
        servers = result["servers"]

        # First server should have port as integer
        assert servers[0]["port"] == 5432

        # Second server should have port as string (conversion failed)
        assert servers[1]["port"] == "invalid_port"

    def test_parse_custom_properties_with_type_conversion(self, parser):
        """Test custom properties parsing with automatic type conversion."""
        # Mock DataFrame with various value types
        props_df = pd.DataFrame([
            {"Property": "stringProp", "Value": "hello"},
            {"Property": "intProp", "Value": "42"},
            {"Property": "floatProp", "Value": "3.14"},
            {"Property": "boolProp", "Value": "true"},
            {"Property": "emptyProp", "Value": ""},
            {"Property": "noneProp", "Value": None}
        ])

        parser.worksheets = {"Custom Properties": props_df}

        result = parser._parse_custom_properties()

        assert "customProperties" in result
        props = result["customProperties"]

        # Should have 5 properties (empty gets converted to None but still included)
        assert len(props) == 5

        # Find each property and check type conversion
        string_prop = next(p for p in props if p["property"] == "stringProp")
        assert string_prop["value"] == "hello"

        int_prop = next(p for p in props if p["property"] == "intProp")
        assert int_prop["value"] == 42

        float_prop = next(p for p in props if p["property"] == "floatProp")
        assert float_prop["value"] == 3.14

        bool_prop = next(p for p in props if p["property"] == "boolProp")
        assert bool_prop["value"] == True

    def test_integration_with_generated_excel(self, parser):
        """Test parsing Excel file generated by ODCSToExcelConverter."""
        # Original ODCS data
        original_data = {
            "version": "2.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "integration-test",
            "status": "draft",
            "name": "Integration Test",
            "tags": ["integration", "roundtrip"],
            "description": {
                "usage": "Integration testing",
                "purpose": "Verify roundtrip conversion"
            }
        }

        # Generate Excel file
        converter = ODCSToExcelConverter()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as excel_file:
            excel_path = excel_file.name

        try:
            converter.generate_from_dict(original_data, excel_path)

            # Parse it back
            parsed_data = parser.parse_from_file(excel_path)

            # Verify core fields are preserved
            assert parsed_data.get("version") == "2.0.0"
            assert parsed_data.get("kind") == "DataContract"
            assert parsed_data.get("apiVersion") == "v3.0.2"
            assert parsed_data.get("id") == "integration-test"
            assert parsed_data.get("status") == "draft"
            assert parsed_data.get("name") == "Integration Test"

            # Verify tags (order might differ)
            assert set(parsed_data.get("tags", [])) == {"integration", "roundtrip"}

            # Verify description
            desc = parsed_data.get("description", {})
            assert desc.get("usage") == "Integration testing"
            assert desc.get("purpose") == "Verify roundtrip conversion"

        finally:
            Path(excel_path).unlink(missing_ok=True)
