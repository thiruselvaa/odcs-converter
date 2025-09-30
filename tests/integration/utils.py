"""Utilities for integration tests."""

import json
import tempfile
from pathlib import Path
from typing import Any, Dict, Generator, List, Optional, Tuple
from unittest.mock import MagicMock, patch
import pytest
from openpyxl import Workbook, load_workbook

from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.yaml_converter import YAMLConverter


class IntegrationTestHelper:
    """Helper class for integration test operations."""

    @staticmethod
    def create_complete_odcs_dict() -> Dict[str, Any]:
        """Create complete ODCS data for integration testing."""
        return {
            "version": "2.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "integration-test-complete",
            "name": "Integration Test Contract",
            "tenant": "integration-tenant",
            "status": "active",
            "dataProduct": "Integration Test Product",
            "domain": "integration_domain",
            "contractCreatedTs": "2024-01-15T10:30:00Z",
            "tags": ["integration", "test", "complete"],
            "description": {
                "usage": "Complete integration test contract",
                "purpose": "Testing component interactions",
                "limitations": "Integration test environment only"
            },
            "servers": [
                {
                    "server": "integration-db-primary",
                    "type": "postgresql",
                    "description": "Primary integration database",
                    "environment": "integration",
                    "host": "integration-db.example.com",
                    "port": 5432,
                    "database": "integration_db",
                    "schema": "public"
                }
            ],
            "schema": [
                {
                    "name": "integration_table",
                    "logicalType": "object",
                    "physicalName": "integration_table_v1",
                    "description": "Integration test table",
                    "businessName": "Integration Test Table",
                    "dataGranularityDescription": "One record per integration entity",
                    "properties": [
                        {
                            "name": "id",
                            "logicalType": "integer",
                            "physicalType": "BIGINT",
                            "description": "Unique identifier",
                            "required": True,
                            "primaryKey": True,
                            "primaryKeyPosition": 1
                        },
                        {
                            "name": "name",
                            "logicalType": "string",
                            "physicalType": "VARCHAR(255)",
                            "description": "Entity name",
                            "required": True
                        },
                        {
                            "name": "created_at",
                            "logicalType": "timestamp",
                            "physicalType": "TIMESTAMP",
                            "description": "Creation timestamp",
                            "required": True
                        },
                        {
                            "name": "active",
                            "logicalType": "boolean",
                            "physicalType": "BOOLEAN",
                            "description": "Active status",
                            "required": False
                        }
                    ]
                }
            ],
            "support": [
                {
                    "channel": "integration-support",
                    "url": "https://support.example.com/integration",
                    "description": "Integration support channel",
                    "tool": "web",
                    "scope": "issues"
                }
            ],
            "team": [
                {
                    "username": "integration.user@example.com",
                    "name": "Integration User",
                    "role": "owner",
                    "description": "Integration test owner"
                }
            ],
            "roles": [
                {
                    "role": "integration_reader",
                    "description": "Read-only access to integration data",
                    "access": "SELECT"
                }
            ],
            "customProperties": [
                {
                    "property": "integrationEnvironment",
                    "value": "testing"
                }
            ]
        }

    @staticmethod
    def create_multi_table_odcs() -> Dict[str, Any]:
        """Create ODCS with multiple tables for complex integration testing."""
        base = IntegrationTestHelper.create_complete_odcs_dict()
        base["id"] = "integration-multi-table"
        base["schema"] = [
            {
                "name": "users",
                "logicalType": "object",
                "physicalName": "users_v1",
                "description": "Users table",
                "properties": [
                    {
                        "name": "user_id",
                        "logicalType": "integer",
                        "physicalType": "BIGINT",
                        "description": "User ID",
                        "required": True,
                        "primaryKey": True,
                        "primaryKeyPosition": 1
                    },
                    {
                        "name": "email",
                        "logicalType": "string",
                        "physicalType": "VARCHAR(255)",
                        "description": "User email",
                        "required": True
                    }
                ]
            },
            {
                "name": "orders",
                "logicalType": "object",
                "physicalName": "orders_v1",
                "description": "Orders table",
                "properties": [
                    {
                        "name": "order_id",
                        "logicalType": "integer",
                        "physicalType": "BIGINT",
                        "description": "Order ID",
                        "required": True,
                        "primaryKey": True,
                        "primaryKeyPosition": 1
                    },
                    {
                        "name": "user_id",
                        "logicalType": "integer",
                        "physicalType": "BIGINT",
                        "description": "User ID (FK)",
                        "required": True
                    },
                    {
                        "name": "amount",
                        "logicalType": "number",
                        "physicalType": "DECIMAL(10,2)",
                        "description": "Order amount",
                        "required": True
                    }
                ]
            }
        ]
        return base


class ExcelTestHelper:
    """Helper for Excel-related integration tests."""

    @staticmethod
    def create_sample_excel_workbook(odcs_data: Dict[str, Any]) -> Generator[Path, None, None]:
        """Create Excel workbook from ODCS data for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_path = Path(tmp_file.name)

        converter = ODCSToExcelConverter()
        converter.generate_from_dict(odcs_data, temp_path)

        try:
            yield temp_path
        finally:
            temp_path.unlink(missing_ok=True)

    @staticmethod
    def validate_excel_structure(excel_path: Path, expected_sheets: List[str]) -> bool:
        """Validate that Excel file has expected sheet structure."""
        try:
            workbook = load_workbook(excel_path)
            actual_sheets = workbook.sheetnames

            for sheet in expected_sheets:
                if sheet not in actual_sheets:
                    return False

            return True
        except Exception:
            return False

    @staticmethod
    def get_excel_sheet_data(excel_path: Path, sheet_name: str) -> List[List[Any]]:
        """Get data from specific Excel sheet."""
        workbook = load_workbook(excel_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data

    @staticmethod
    def create_invalid_excel_file() -> Generator[Path, None, None]:
        """Create an invalid Excel file for error testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            # Write invalid content
            tmp_file.write(b"This is not a valid Excel file")
            temp_path = Path(tmp_file.name)

        try:
            yield temp_path
        finally:
            temp_path.unlink(missing_ok=True)


class ConversionTestHelper:
    """Helper for testing bidirectional conversions."""

    @staticmethod
    def test_roundtrip_conversion(
        odcs_data: Dict[str, Any],
        temp_dir: Path
    ) -> Tuple[bool, Optional[str]]:
        """Test roundtrip conversion: ODCS -> Excel -> ODCS."""
        try:
            # Step 1: Convert ODCS to Excel
            excel_path = temp_dir / "roundtrip_test.xlsx"
            converter = ODCSToExcelConverter()
            converter.generate_from_dict(odcs_data, excel_path)

            if not excel_path.exists():
                return False, "Excel file was not created"

            # Step 2: Convert Excel back to ODCS
            parser = ExcelToODCSParser()
            converted_data = parser.parse_from_file(excel_path)

            # Step 3: Compare key fields (some fields might be normalized)
            key_fields = ["version", "kind", "apiVersion", "id", "status"]
            for field in key_fields:
                if odcs_data.get(field) != converted_data.get(field):
                    return False, f"Field '{field}' mismatch: {odcs_data.get(field)} != {converted_data.get(field)}"

            return True, None

        except Exception as e:
            return False, str(e)

    @staticmethod
    def compare_odcs_structures(
        original: Dict[str, Any],
        converted: Dict[str, Any],
        ignore_fields: Optional[List[str]] = None
    ) -> Tuple[bool, List[str]]:
        """Compare two ODCS structures and return differences."""
        ignore_fields = ignore_fields or []
        differences = []

        def _compare_recursive(obj1, obj2, path=""):
            if type(obj1) != type(obj2):
                differences.append(f"{path}: type mismatch {type(obj1)} != {type(obj2)}")
                return

            if isinstance(obj1, dict):
                all_keys = set(obj1.keys()) | set(obj2.keys())
                for key in all_keys:
                    current_path = f"{path}.{key}" if path else key
                    if key in ignore_fields:
                        continue

                    if key not in obj1:
                        differences.append(f"{current_path}: missing in original")
                    elif key not in obj2:
                        differences.append(f"{current_path}: missing in converted")
                    else:
                        _compare_recursive(obj1[key], obj2[key], current_path)

            elif isinstance(obj1, list):
                if len(obj1) != len(obj2):
                    differences.append(f"{path}: length mismatch {len(obj1)} != {len(obj2)}")
                    return

                for i, (item1, item2) in enumerate(zip(obj1, obj2)):
                    _compare_recursive(item1, item2, f"{path}[{i}]")

            else:
                if obj1 != obj2:
                    differences.append(f"{path}: value mismatch {obj1} != {obj2}")

        _compare_recursive(original, converted)
        return len(differences) == 0, differences


class ComponentTestHelper:
    """Helper for testing component interactions."""

    @staticmethod
    def test_yaml_excel_integration(
        yaml_data: Dict[str, Any],
        temp_dir: Path
    ) -> Tuple[bool, Optional[str]]:
        """Test YAML -> ODCS -> Excel integration."""
        try:
            # Step 1: Save as YAML
            yaml_path = temp_dir / "test.yaml"
            YAMLConverter.dict_to_yaml(yaml_data, yaml_path)

            # Step 2: Load YAML and convert to Excel
            loaded_data = YAMLConverter.yaml_to_dict(yaml_path)
            excel_path = temp_dir / "from_yaml.xlsx"

            converter = ODCSToExcelConverter()
            converter.generate_from_dict(loaded_data, excel_path)

            # Step 3: Verify Excel file exists and is valid
            if not excel_path.exists():
                return False, "Excel file was not created from YAML"

            # Step 4: Parse Excel back to verify structure
            parser = ExcelToODCSParser()
            parsed_data = parser.parse_from_file(excel_path)

            return True, None

        except Exception as e:
            return False, str(e)

    @staticmethod
    def test_error_handling_chain(
        invalid_data: Any,
        temp_dir: Path
    ) -> List[str]:
        """Test error handling across components."""
        errors_found = []

        # Test ODCS to Excel with invalid data
        try:
            excel_path = temp_dir / "invalid_test.xlsx"
            converter = ODCSToExcelConverter()
            converter.generate_from_dict(invalid_data, excel_path)
        except Exception as e:
            errors_found.append(f"ODCS to Excel error: {str(e)}")

        # Test YAML conversion with invalid data
        try:
            yaml_path = temp_dir / "invalid_test.yaml"
            YAMLConverter.dict_to_yaml(invalid_data, yaml_path)
        except Exception as e:
            errors_found.append(f"YAML conversion error: {str(e)}")

        return errors_found


class WorkflowTestHelper:
    """Helper for testing complete workflows."""

    @staticmethod
    def simulate_user_workflow(
        input_data: Dict[str, Any],
        workflow_type: str,
        temp_dir: Path
    ) -> Tuple[bool, Dict[str, Any], List[str]]:
        """Simulate complete user workflow."""
        results = {"files_created": [], "conversions_completed": []}
        errors = []

        try:
            if workflow_type == "odcs_to_excel":
                # Simulate: User provides ODCS JSON, wants Excel
                excel_path = temp_dir / "user_output.xlsx"
                converter = ODCSToExcelConverter()
                converter.generate_from_dict(input_data, excel_path)

                results["files_created"].append(str(excel_path))
                results["conversions_completed"].append("ODCS -> Excel")

            elif workflow_type == "excel_to_odcs":
                # Simulate: User provides Excel, wants ODCS
                # First create Excel from input data
                excel_path = temp_dir / "user_input.xlsx"
                converter = ODCSToExcelConverter()
                converter.generate_from_dict(input_data, excel_path)

                # Then parse it back
                parser = ExcelToODCSParser()
                parsed_data = parser.parse_from_file(excel_path)

                # Save as JSON
                json_path = temp_dir / "user_output.json"
                with open(json_path, 'w') as f:
                    json.dump(parsed_data, f, indent=2)

                results["files_created"].extend([str(excel_path), str(json_path)])
                results["conversions_completed"].extend(["ODCS -> Excel", "Excel -> ODCS"])

            elif workflow_type == "roundtrip":
                # Simulate: User tests roundtrip conversion
                success, error = ConversionTestHelper.test_roundtrip_conversion(input_data, temp_dir)
                if not success:
                    errors.append(f"Roundtrip failed: {error}")
                else:
                    results["conversions_completed"].append("Full Roundtrip")

            return len(errors) == 0, results, errors

        except Exception as e:
            errors.append(f"Workflow error: {str(e)}")
            return False, results, errors


# Pytest fixtures specific to integration tests
@pytest.fixture
def integration_test_helper():
    """Provide integration test helper instance."""
    return IntegrationTestHelper()


@pytest.fixture
def excel_test_helper():
    """Provide Excel test helper instance."""
    return ExcelTestHelper()


@pytest.fixture
def conversion_test_helper():
    """Provide conversion test helper instance."""
    return ConversionTestHelper()


@pytest.fixture
def component_test_helper():
    """Provide component test helper instance."""
    return ComponentTestHelper()


@pytest.fixture
def workflow_test_helper():
    """Provide workflow test helper instance."""
    return WorkflowTestHelper()


@pytest.fixture
def complete_odcs_data():
    """Provide complete ODCS data for integration testing."""
    return IntegrationTestHelper.create_complete_odcs_dict()


@pytest.fixture
def multi_table_odcs_data():
    """Provide multi-table ODCS data for complex testing."""
    return IntegrationTestHelper.create_multi_table_odcs()


# Custom decorators for integration tests
def integration_test(func):
    """Decorator to mark a function as an integration test."""
    return pytest.mark.integration(func)


def slow_integration_test(func):
    """Decorator to mark a function as a slow integration test."""
    return pytest.mark.integration(pytest.mark.slow(func))


def requires_temp_files(func):
    """Decorator for tests that require temporary file management."""
    def wrapper(*args, **kwargs):
        # This decorator can be extended to handle specific temp file requirements
        return func(*args, **kwargs)
    return wrapper
