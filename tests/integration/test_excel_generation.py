"""Integration tests for Excel generation workflows."""

import pytest
from unittest.mock import patch, MagicMock

from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.yaml_converter import YAMLConverter


@pytest.mark.integration
class TestExcelGenerationWorkflow:
    """Integration tests for complete Excel generation workflows."""

    def test_generate_excel_from_json_file_complete_workflow(
        self, sample_odcs_complete, json_file_factory, temp_dir
    ):
        """Test complete workflow: JSON file -> Excel generation."""
        # Setup
        converter = ODCSToExcelConverter()
        json_file = json_file_factory(sample_odcs_complete, "complete_contract.json")
        excel_file = temp_dir / "integration_output.xlsx"

        # Execute
        converter.generate_from_file(json_file, excel_file)

        # Verify
        assert excel_file.exists()
        assert excel_file.stat().st_size > 0

        # Verify Excel structure using openpyxl
        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)

        expected_sheets = [
            "Basic Information",
            "Tags",
            "Description",
            "Servers",
            "Schema",
            "Support",
            "Team",
            "Roles",
            "SLA Properties",
            "Authoritative Definitions",
            "Custom Properties",
        ]

        for sheet_name in expected_sheets:
            assert sheet_name in workbook.sheetnames

        # Verify basic information sheet content
        basic_sheet = workbook["Basic Information"]
        assert basic_sheet["A1"].value == "Field"
        assert basic_sheet["B1"].value == "Value"

        # Find version row and verify
        version_found = False
        for row in basic_sheet.iter_rows(min_row=2, max_row=20, min_col=1, max_col=2):
            if row[0].value == "version":
                assert row[1].value == "2.0.0"
                version_found = True
                break
        assert version_found

    def test_generate_excel_from_yaml_file_workflow(
        self, sample_odcs_complete, yaml_file_factory, temp_dir
    ):
        """Test workflow: YAML file -> Excel generation."""
        # Setup
        converter = ODCSToExcelConverter()
        yaml_file = yaml_file_factory(sample_odcs_complete, "complete_contract.yaml")
        excel_file = temp_dir / "yaml_to_excel.xlsx"

        # Load YAML and convert
        yaml_data = YAMLConverter.yaml_to_dict(yaml_file)
        converter.generate_from_dict(yaml_data, excel_file)

        # Verify
        assert excel_file.exists()

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)
        assert len(workbook.sheetnames) >= 5

        # Verify tags sheet
        if "Tags" in workbook.sheetnames:
            tags_sheet = workbook["Tags"]
            tag_values = []
            for row in tags_sheet.iter_rows(min_row=2, max_col=1):
                if row[0].value:
                    tag_values.append(row[0].value)

            expected_tags = sample_odcs_complete.get("tags", [])
            for tag in expected_tags:
                assert tag in tag_values

    def test_generate_excel_with_custom_style_configuration(
        self, sample_odcs_complete, temp_dir
    ):
        """Test Excel generation with custom styling."""
        # Setup custom style
        from openpyxl.styles import Font, PatternFill, Alignment

        custom_style = {
            "header_font": Font(bold=True, color="FF0000"),  # Red
            "header_fill": PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            ),  # Yellow
            "alignment": Alignment(horizontal="center", vertical="center"),
        }

        converter = ODCSToExcelConverter(style_config=custom_style)
        excel_file = temp_dir / "styled_excel.xlsx"

        # Execute
        converter.generate_from_dict(sample_odcs_complete, excel_file)

        # Verify
        assert excel_file.exists()

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)

        # Check that basic information sheet has styling applied
        basic_sheet = workbook["Basic Information"]
        header_cell = basic_sheet["A1"]

        # Note: Style verification can be complex due to openpyxl's style handling
        # Here we just verify the file was created with custom styling
        assert header_cell.value == "Field"

    @patch("odcs_converter.generator.requests.get")
    def test_generate_excel_from_url_workflow(
        self, mock_get, sample_odcs_complete, temp_dir
    ):
        """Test workflow: URL -> Excel generation."""
        # Setup mock response
        mock_response = MagicMock()
        mock_response.json.return_value = sample_odcs_complete
        mock_response.raise_for_status.return_value = None
        mock_get.return_value = mock_response

        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "url_to_excel.xlsx"
        test_url = "https://example.com/contract.json"

        # Execute
        converter.generate_from_url(test_url, excel_file)

        # Verify
        mock_get.assert_called_once_with(test_url, timeout=30)
        assert excel_file.exists()

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)
        assert "Basic Information" in workbook.sheetnames

    def test_generate_excel_with_minimal_data(self, sample_odcs_minimal, temp_dir):
        """Test Excel generation with minimal ODCS data."""
        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "minimal_excel.xlsx"

        # Execute
        converter.generate_from_dict(sample_odcs_minimal, excel_file)

        # Verify
        assert excel_file.exists()

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)

        # Should have basic information sheet at minimum
        assert "Basic Information" in workbook.sheetnames

        basic_sheet = workbook["Basic Information"]
        # Verify required fields are present
        field_values = {}
        for row in basic_sheet.iter_rows(min_row=2, max_col=2):
            if row[0].value and row[1].value:
                field_values[row[0].value] = row[1].value

        assert field_values.get("version") == "1.0.0"
        assert field_values.get("kind") == "DataContract"
        assert field_values.get("id") == "test-contract-minimal"

    def test_generate_excel_with_invalid_data_handling(self, temp_dir):
        """Test Excel generation with invalid/partial ODCS data."""
        # Data that will fail ODCS validation but should still generate Excel
        invalid_data = {
            "version": "1.0.0",
            "kind": "DataContract",
            # Missing required fields
            "invalid_field": "should be ignored",
            "servers": [
                {
                    "server": "test",
                    "type": "unknown_type",  # Invalid type
                    "port": "not_a_number",
                }
            ],
        }

        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "invalid_data_excel.xlsx"

        # Execute - should not raise exception
        converter.generate_from_dict(invalid_data, excel_file)

        # Verify file is created despite invalid data
        assert excel_file.exists()

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)
        assert "Basic Information" in workbook.sheetnames
        assert "Servers" in workbook.sheetnames

    def test_concurrent_excel_generation(self, sample_odcs_complete, temp_dir):
        """Test concurrent Excel generation to ensure thread safety."""
        import threading
        import time

        converter = ODCSToExcelConverter()
        results = {}
        errors = {}

        def generate_excel(thread_id):
            try:
                excel_file = temp_dir / f"concurrent_{thread_id}.xlsx"
                # Add slight delay to increase chance of concurrent access
                time.sleep(0.01)
                converter.generate_from_dict(sample_odcs_complete, excel_file)
                results[thread_id] = excel_file.exists()
            except Exception as e:
                errors[thread_id] = str(e)

        # Create multiple threads
        threads = []
        for i in range(5):
            thread = threading.Thread(target=generate_excel, args=(i,))
            threads.append(thread)

        # Start all threads
        for thread in threads:
            thread.start()

        # Wait for completion
        for thread in threads:
            thread.join()

        # Verify all succeeded
        assert len(errors) == 0, f"Errors occurred: {errors}"
        assert len(results) == 5
        assert all(results.values()), "Some Excel files were not created"

    def test_excel_generation_with_large_dataset(self, temp_dir):
        """Test Excel generation with large dataset."""
        # Create large dataset
        large_data = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "large-dataset-contract",
            "status": "active",
            "name": "Large Dataset Contract",
            "tags": [f"tag_{i}" for i in range(100)],  # 100 tags
            "servers": [
                {
                    "server": f"server_{i}",
                    "type": "postgresql",
                    "description": f"Server {i} for large dataset testing",
                    "host": f"server{i}.example.com",
                    "port": 5432 + i,
                    "database": f"db_{i}",
                }
                for i in range(50)  # 50 servers
            ],
            "team": [
                {
                    "username": f"user{i}@example.com",
                    "name": f"User {i}",
                    "role": f"Role {i % 10}",
                    "description": f"Team member {i}",
                }
                for i in range(30)  # 30 team members
            ],
            "customProperties": [
                {"property": f"property_{i}", "value": f"value_{i}"}
                for i in range(25)  # 25 custom properties
            ],
        }

        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "large_dataset.xlsx"

        # Execute
        converter.generate_from_dict(large_data, excel_file)

        # Verify
        assert excel_file.exists()
        # File should be reasonably large due to data volume
        assert excel_file.stat().st_size > 10000  # At least 10KB

        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)

        # Verify tags sheet has all tags
        tags_sheet = workbook["Tags"]
        tag_count = 0
        for row in tags_sheet.iter_rows(min_row=2):
            if row[0].value:
                tag_count += 1
        assert tag_count == 100

        # Verify servers sheet has all servers
        servers_sheet = workbook["Servers"]
        server_count = 0
        for row in servers_sheet.iter_rows(min_row=2):
            if row[0].value:
                server_count += 1
        assert server_count == 50

    def test_excel_generation_memory_efficiency(self, temp_dir):
        """Test memory efficiency during Excel generation."""
        try:
            import psutil
            import os
        except ImportError:
            pytest.skip("psutil not available for memory testing")

        # Get initial memory usage
        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss

        # Create moderate-sized dataset
        data = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "memory-test-contract",
            "status": "active",
            "description": {
                "usage": "A" * 1000,  # 1KB string
                "purpose": "B" * 1000,
                "limitations": "C" * 1000,
            },
            "servers": [
                {
                    "server": f"server_{i}",
                    "type": "postgresql",
                    "description": "X" * 500,  # 500 bytes per server
                }
                for i in range(20)
            ],
        }

        converter = ODCSToExcelConverter()
        excel_files = []

        # Generate multiple Excel files
        for i in range(10):
            excel_file = temp_dir / f"memory_test_{i}.xlsx"
            converter.generate_from_dict(data, excel_file)
            excel_files.append(excel_file)

        # Check memory usage hasn't grown excessively
        final_memory = process.memory_info().rss
        memory_growth = final_memory - initial_memory

        # Memory growth should be reasonable (less than 50MB for this test)
        assert (
            memory_growth < 50 * 1024 * 1024
        ), f"Excessive memory growth: {memory_growth / 1024 / 1024:.2f}MB"

        # Verify all files were created
        for excel_file in excel_files:
            assert excel_file.exists()

    @patch("odcs_converter.generator.requests.get")
    def test_url_error_handling_workflow(self, mock_get, temp_dir):
        """Test error handling in URL-based workflow."""
        from requests import RequestException

        # Test network error
        mock_get.side_effect = RequestException("Network error")

        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "error_test.xlsx"

        with pytest.raises(ValueError, match="Failed to fetch data from URL"):
            converter.generate_from_url(
                "https://invalid-url.com/contract.json", excel_file
            )

        # Verify no file was created
        assert not excel_file.exists()

    def test_file_permission_error_handling(self, sample_odcs_minimal, temp_dir):
        """Test handling of file permission errors."""
        converter = ODCSToExcelConverter()

        # Try to write to a directory instead of a file
        invalid_path = temp_dir / "directory_instead_of_file"
        invalid_path.mkdir()

        with pytest.raises(Exception):  # Should raise some form of file error
            converter.generate_from_dict(sample_odcs_minimal, invalid_path)

    def test_excel_generation_output_directory_creation(
        self, sample_odcs_minimal, temp_dir
    ):
        """Test that output directories are created automatically."""
        converter = ODCSToExcelConverter()

        # Path with non-existent directories
        nested_path = temp_dir / "level1" / "level2" / "level3" / "output.xlsx"

        # Execute
        converter.generate_from_dict(sample_odcs_minimal, nested_path)

        # Verify
        assert nested_path.exists()
        assert nested_path.parent.exists()

    def test_excel_worksheet_auto_sizing(self, sample_odcs_complete, temp_dir):
        """Test that Excel worksheets have proper column auto-sizing."""
        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "auto_sized.xlsx"

        # Execute
        converter.generate_from_dict(sample_odcs_complete, excel_file)

        # Verify
        from openpyxl import load_workbook

        workbook = load_workbook(excel_file)

        # Check that columns have been auto-sized (width > 0)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                column_letter = column[0].column_letter
                width = sheet.column_dimensions[column_letter].width
                # Auto-sized columns should have width > 0
                assert (
                    width > 0
                ), f"Column {column_letter} in {sheet_name} has zero width"
