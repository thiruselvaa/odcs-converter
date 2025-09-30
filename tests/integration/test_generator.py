"""Tests for ODCS Converter - bidirectional conversion between ODCS and Excel."""

import json
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
from openpyxl import load_workbook

from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.yaml_converter import YAMLConverter
from odcs_converter.models import ODCSDataContract


class TestODCSToExcelConverter:
    """Test cases for ODCS to Excel conversion."""

    @pytest.fixture
    def sample_odcs_data(self):
        """Sample ODCS data for testing."""
        return {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "test-contract-001",
            "name": "Test Contract",
            "status": "active",
            "tags": ["test", "example"],
            "description": {
                "usage": "Test usage",
                "purpose": "Testing purposes"
            },
            "servers": [
                {
                    "server": "test-server",
                    "type": "snowflake",
                    "description": "Test server"
                }
            ],
            "team": [
                {
                    "username": "test@example.com",
                    "name": "Test User",
                    "role": "owner"
                }
            ]
        }

    @pytest.fixture
    def converter(self):
        """Create ODCSToExcelConverter instance."""
        return ODCSToExcelConverter()

    def test_generate_from_dict(self, converter, sample_odcs_data):
        """Test generating Excel from ODCS dictionary."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            output_path = tmp_file.name

        try:
            converter.generate_from_dict(sample_odcs_data, output_path)

            # Verify file was created
            assert Path(output_path).exists()

            # Verify Excel structure
            workbook = load_workbook(output_path)
            expected_sheets = [
                "Basic Information", "Tags", "Description",
                "Servers", "Team"
            ]

            for sheet_name in expected_sheets:
                assert sheet_name in workbook.sheetnames

        finally:
            Path(output_path).unlink(missing_ok=True)

    def test_generate_from_file(self, converter, sample_odcs_data):
        """Test generating Excel from JSON file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as json_file:
            json.dump(sample_odcs_data, json_file)
            json_file_path = json_file.name

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as excel_file:
            excel_file_path = excel_file.name

        try:
            converter.generate_from_file(json_file_path, excel_file_path)
            assert Path(excel_file_path).exists()

        finally:
            Path(json_file_path).unlink(missing_ok=True)
            Path(excel_file_path).unlink(missing_ok=True)

    @patch('requests.get')
    def test_generate_from_url(self, mock_get, converter, sample_odcs_data):
        """Test generating Excel from URL."""
        # Mock HTTP response
        mock_response = MagicMock()
        mock_response.json.return_value = sample_odcs_data
        mock_response.raise_for_status.return_value = None
        mock_get.return_value = mock_response

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            output_path = tmp_file.name

        try:
            converter.generate_from_url("https://example.com/contract.json", output_path)
            assert Path(output_path).exists()

        finally:
            Path(output_path).unlink(missing_ok=True)

    def test_file_not_found_error(self, converter):
        """Test handling of missing input file."""
        with pytest.raises(FileNotFoundError):
            converter.generate_from_file("nonexistent.json", "output.xlsx")

    def test_invalid_url_error(self, converter):
        """Test handling of invalid URL."""
        with patch('odcs_converter.generator.requests.get') as mock_get:
            from requests import RequestException
            mock_get.side_effect = RequestException("Network error")

            with pytest.raises(ValueError, match="Failed to fetch data from URL"):
                converter.generate_from_url("invalid-url", "output.xlsx")


class TestExcelToODCSParser:
    """Test cases for Excel to ODCS conversion."""

    @pytest.fixture
    def parser(self):
        """Create ExcelToODCSParser instance."""
        return ExcelToODCSParser()

    @pytest.fixture
    def sample_excel_data(self):
        """Sample Excel data structure."""
        return {
            "Basic Information": {
                "Field": ["version", "kind", "apiVersion", "id", "status"],
                "Value": ["1.0.0", "DataContract", "v3.0.2", "test-001", "active"],
                "Description": ["Version", "Kind", "API Version", "ID", "Status"]
            },
            "Tags": {
                "Tag": ["test", "example", "sample"]
            }
        }

    def test_convert_value(self, parser):
        """Test value conversion utilities."""
        assert parser._convert_value("true") == True
        assert parser._convert_value("false") == False
        assert parser._convert_value("123") == 123
        assert parser._convert_value("123.45") == 123.45
        assert parser._convert_value("text") == "text"
        assert parser._convert_value("") is None
        assert parser._convert_value(None) is None

    def test_clean_data(self, parser):
        """Test data cleaning functionality."""
        dirty_data = {
            "field1": "value1",
            "field2": None,
            "field3": "",
            "field4": {
                "nested1": "value",
                "nested2": None,
                "nested3": ""
            },
            "field5": ["item1", "", None, "item2"],
            "field6": []
        }

        cleaned = parser._clean_data(dirty_data)

        expected = {
            "field1": "value1",
            "field4": {"nested1": "value"},
            "field5": ["item1", "item2"]
        }

        assert cleaned == expected

    def test_parse_basic_information(self, parser):
        """Test parsing basic information."""
        # This would require setting up mock DataFrame data
        # For now, test the method exists and handles empty data
        parser.worksheets = {}
        result = parser._parse_basic_information()
        assert result == {}

    def test_validate_odcs_data_valid(self, parser):
        """Test ODCS data validation with valid data."""
        valid_data = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "test-001",
            "status": "active"
        }

        assert parser.validate_odcs_data(valid_data) == True

    def test_validate_odcs_data_invalid(self, parser):
        """Test ODCS data validation with invalid data."""
        invalid_data = {
            "version": "1.0.0",
            # Missing required fields
        }

        assert parser.validate_odcs_data(invalid_data) == False


class TestYAMLConverter:
    """Test cases for YAML conversion utilities."""

    @pytest.fixture
    def sample_data(self):
        """Sample data for YAML conversion."""
        return {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "description": {
                "usage": "Test usage",
                "purpose": "Testing"
            },
            "tags": ["test", "yaml"]
        }

    def test_dict_to_yaml_string(self, sample_data):
        """Test converting dictionary to YAML string."""
        yaml_string = YAMLConverter.dict_to_yaml_string(sample_data)

        assert "version: 1.0.0" in yaml_string
        assert "kind: DataContract" in yaml_string
        assert "- test" in yaml_string
        assert "- yaml" in yaml_string

    def test_yaml_string_to_dict(self, sample_data):
        """Test converting YAML string to dictionary."""
        yaml_string = YAMLConverter.dict_to_yaml_string(sample_data)
        converted_data = YAMLConverter.yaml_string_to_dict(yaml_string)

        assert converted_data == sample_data

    def test_dict_to_yaml_file(self, sample_data):
        """Test converting dictionary to YAML file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.yaml', delete=False) as tmp_file:
            yaml_file_path = tmp_file.name

        try:
            YAMLConverter.dict_to_yaml(sample_data, yaml_file_path)

            assert Path(yaml_file_path).exists()

            # Read back and verify
            loaded_data = YAMLConverter.yaml_to_dict(yaml_file_path)
            assert loaded_data == sample_data

        finally:
            Path(yaml_file_path).unlink(missing_ok=True)

    def test_yaml_to_dict_file_not_found(self):
        """Test handling of missing YAML file."""
        with pytest.raises(FileNotFoundError):
            YAMLConverter.yaml_to_dict("nonexistent.yaml")

    def test_is_yaml_file(self):
        """Test YAML file detection."""
        assert YAMLConverter.is_yaml_file("test.yaml") == True
        assert YAMLConverter.is_yaml_file("test.yml") == True
        assert YAMLConverter.is_yaml_file("test.json") == False
        assert YAMLConverter.is_yaml_file("test.txt") == False

    def test_normalize_yaml_extension(self):
        """Test YAML extension normalization."""
        result1 = YAMLConverter.normalize_yaml_extension("test.txt", prefer_yaml=True)
        assert result1 == Path("test.yaml")

        result2 = YAMLConverter.normalize_yaml_extension("test.txt", prefer_yaml=False)
        assert result2 == Path("test.yml")

        result3 = YAMLConverter.normalize_yaml_extension("test.yaml")
        assert result3 == Path("test.yaml")

    def test_invalid_yaml_string(self):
        """Test handling of invalid YAML string."""
        invalid_yaml = "invalid: yaml: content: [unclosed"

        with pytest.raises(ValueError, match="Invalid YAML format"):
            YAMLConverter.yaml_string_to_dict(invalid_yaml)


class TestIntegrationBidirectional:
    """Integration tests for bidirectional conversion."""

    @pytest.fixture
    def sample_odcs_data(self):
        """Complete ODCS sample data."""
        return {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "integration-test-001",
            "name": "Integration Test Contract",
            "status": "active",
            "tags": ["integration", "test", "bidirectional"],
            "description": {
                "usage": "Integration testing",
                "purpose": "Test bidirectional conversion",
                "limitations": "Test environment only"
            },
            "servers": [
                {
                    "server": "test-server",
                    "type": "snowflake",
                    "description": "Integration test server",
                    "database": "TEST_DB",
                    "schema": "PUBLIC"
                }
            ],
            "team": [
                {
                    "username": "tester@example.com",
                    "name": "Integration Tester",
                    "role": "owner"
                }
            ],
            "customProperties": [
                {
                    "property": "testProperty",
                    "value": "testValue"
                }
            ]
        }

    def test_odcs_to_excel_to_odcs_roundtrip_json(self, sample_odcs_data):
        """Test complete roundtrip: ODCS JSON → Excel → ODCS JSON."""
        # Create temporary files
        json_input = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)
        excel_temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        json_output = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False)

        try:
            # Step 1: Write original ODCS data
            json.dump(sample_odcs_data, json_input)
            json_input.close()

            # Step 2: Convert ODCS → Excel
            converter = ODCSToExcelConverter()
            converter.generate_from_file(json_input.name, excel_temp.name)

            # Verify Excel file exists and has expected structure
            assert Path(excel_temp.name).exists()
            workbook = load_workbook(excel_temp.name)
            assert "Basic Information" in workbook.sheetnames
            assert "Tags" in workbook.sheetnames
            assert "Team" in workbook.sheetnames

            # Step 3: Convert Excel → ODCS
            parser = ExcelToODCSParser()
            parsed_data = parser.parse_from_file(excel_temp.name)

            # Step 4: Write parsed data
            json_output.close()
            with open(json_output.name, 'w') as f:
                json.dump(parsed_data, f, indent=2)

            # Step 5: Verify essential data integrity
            assert parsed_data.get("version") == sample_odcs_data["version"]
            assert parsed_data.get("kind") == sample_odcs_data["kind"]
            assert parsed_data.get("id") == sample_odcs_data["id"]
            assert parsed_data.get("status") == sample_odcs_data["status"]

            # Verify tags are preserved (order might differ)
            if "tags" in parsed_data:
                assert set(parsed_data["tags"]) == set(sample_odcs_data["tags"])

        finally:
            # Cleanup
            Path(json_input.name).unlink(missing_ok=True)
            Path(excel_temp.name).unlink(missing_ok=True)
            Path(json_output.name).unlink(missing_ok=True)

    def test_odcs_to_excel_to_odcs_roundtrip_yaml(self, sample_odcs_data):
        """Test complete roundtrip: ODCS YAML → Excel → ODCS YAML."""
        # Create temporary files
        yaml_input = tempfile.NamedTemporaryFile(mode='w', suffix='.yaml', delete=False)
        excel_temp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        yaml_output = tempfile.NamedTemporaryFile(mode='w', suffix='.yaml', delete=False)

        try:
            # Step 1: Write original ODCS data as YAML
            yaml_input.close()
            YAMLConverter.dict_to_yaml(sample_odcs_data, yaml_input.name)

            # Step 2: Convert ODCS YAML → Excel
            converter = ODCSToExcelConverter()
            yaml_data = YAMLConverter.yaml_to_dict(yaml_input.name)
            converter.generate_from_dict(yaml_data, excel_temp.name)

            # Step 3: Convert Excel → ODCS YAML
            parser = ExcelToODCSParser()
            parsed_data = parser.parse_from_file(excel_temp.name)

            yaml_output.close()
            YAMLConverter.dict_to_yaml(parsed_data, yaml_output.name)

            # Step 4: Verify data integrity
            final_data = YAMLConverter.yaml_to_dict(yaml_output.name)

            assert final_data.get("version") == sample_odcs_data["version"]
            assert final_data.get("kind") == sample_odcs_data["kind"]
            assert final_data.get("id") == sample_odcs_data["id"]

        finally:
            # Cleanup
            Path(yaml_input.name).unlink(missing_ok=True)
            Path(excel_temp.name).unlink(missing_ok=True)
            Path(yaml_output.name).unlink(missing_ok=True)
