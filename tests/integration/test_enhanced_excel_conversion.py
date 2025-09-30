"""Integration tests for enhanced Excel conversion with all new ODCS v3.0.2 fields."""

from typing import Dict, Any

import pytest

from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.models import ODCSDataContract


class TestEnhancedExcelConversion:
    """Test enhanced Excel conversion with all new fields."""

    @pytest.fixture
    def enhanced_odcs_contract(self) -> Dict[str, Any]:
        """Complete ODCS contract with all enhanced features."""
        return {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "enhanced-test-contract",
            "name": "Enhanced Test Contract",
            "status": "active",
            "domain": "test_domain",
            "dataProduct": "Enhanced Test Product",
            "tenant": "test-tenant",
            "tags": ["enhanced", "test", "complete"],
            "description": {
                "purpose": "Test contract with all enhanced features",
                "usage": "Testing enhanced ODCS v3.0.2 capabilities",
                "limitations": "Test environment only",
            },
            "schema": [
                {
                    "name": "users",
                    "logicalType": "object",
                    "physicalType": "table",
                    "physicalName": "users_v1",
                    "description": "User information table",
                    "businessName": "User Data",
                    "dataGranularityDescription": "One record per user",
                    "tags": ["users", "core"],
                    "properties": [
                        {
                            "name": "user_id",
                            "logicalType": "integer",
                            "logicalTypeOptions": {
                                "minimum": 1,
                                "maximum": 999999999,
                                "exclusiveMinimum": False,
                                "exclusiveMaximum": False,
                            },
                            "physicalType": "BIGINT",
                            "description": "Unique user identifier",
                            "businessName": "User ID",
                            "required": True,
                            "unique": True,
                            "primaryKey": True,
                            "primaryKeyPosition": 1,
                            "classification": "public",
                            "criticalDataElement": True,
                            "examples": [1, 12345, 999999],
                            "tags": ["primary_key", "identifier"],
                            "quality": [
                                {
                                    "name": "Unique Check",
                                    "description": "Ensure user IDs are unique",
                                    "type": "library",
                                    "rule": "uniqueness",
                                    "dimension": "uniqueness",
                                    "mustBe": 100,
                                    "unit": "percent",
                                    "severity": "critical",
                                    "businessImpact": "Duplicate users in system",
                                    "tags": ["uniqueness", "critical"],
                                },
                                {
                                    "name": "Not Null Check",
                                    "type": "sql",
                                    "query": "SELECT COUNT(*) FROM ${object} WHERE ${property} IS NULL",
                                    "mustBe": 0,
                                    "dimension": "completeness",
                                    "scheduler": "cron",
                                    "schedule": "0 9 * * *",
                                },
                            ],
                        },
                        {
                            "name": "email",
                            "logicalType": "string",
                            "logicalTypeOptions": {
                                "format": "email",
                                "minLength": 5,
                                "maxLength": 255,
                                "pattern": r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$",
                            },
                            "physicalType": "VARCHAR(255)",
                            "description": "User email address",
                            "businessName": "Email Address",
                            "required": True,
                            "unique": True,
                            "classification": "restricted",
                            "encryptedName": "email_encrypted",
                            "criticalDataElement": True,
                            "transformSourceObjects": ["user_profiles", "contacts"],
                            "transformLogic": "COALESCE(up.primary_email, c.email)",
                            "transformDescription": "Use primary email from profiles, fallback to contacts",
                            "examples": ["john@example.com", "jane@company.org"],
                            "tags": ["pii", "contact", "required"],
                            "quality": [
                                {
                                    "name": "Email Format Check",
                                    "rule": "validEmail",
                                    "validValues": [
                                        "@gmail.com",
                                        "@company.com",
                                        "@example.org",
                                    ],
                                    "mustBe": 100,
                                    "unit": "percent",
                                    "dimension": "conformity",
                                },
                                {
                                    "name": "Custom Validation",
                                    "type": "custom",
                                    "engine": "soda",
                                    "implementation": """
                                    type: email_format
                                    column: email
                                    must_be_valid: true
                                    """,
                                    "dimension": "conformity",
                                },
                            ],
                        },
                        {
                            "name": "age",
                            "logicalType": "integer",
                            "logicalTypeOptions": {
                                "minimum": 0,
                                "maximum": 150,
                                "exclusiveMinimum": True,
                                "exclusiveMaximum": False,
                            },
                            "physicalType": "INTEGER",
                            "description": "User age in years",
                            "required": False,
                            "classification": "public",
                            "examples": [25, 30, 45],
                            "quality": [
                                {
                                    "name": "Age Range Check",
                                    "mustBeGreaterThan": 0,
                                    "mustBeLessThan": 150,
                                    "dimension": "conformity",
                                }
                            ],
                        },
                        {
                            "name": "preferences",
                            "logicalType": "array",
                            "logicalTypeOptions": {
                                "minItems": 0,
                                "maxItems": 20,
                                "uniqueItems": True,
                            },
                            "physicalType": "JSON",
                            "description": "User preferences array",
                            "required": False,
                            "items": {
                                "name": "preference",
                                "logicalType": "string",
                                "logicalTypeOptions": {
                                    "maxLength": 50,
                                },
                            },
                            "examples": [
                                ["dark_mode", "notifications"],
                                ["privacy", "emails"],
                            ],
                        },
                    ],
                    "quality": [
                        {
                            "name": "Table Row Count",
                            "rule": "rowCount",
                            "mustBeBetween": [1000, 100000],
                            "unit": "rows",
                            "dimension": "completeness",
                            "severity": "high",
                            "businessImpact": "Insufficient user data for analytics",
                        }
                    ],
                    "authoritativeDefinitions": [
                        {
                            "url": "https://wiki.company.com/users-table",
                            "type": "businessDefinition",
                        }
                    ],
                }
            ],
            "support": [
                {
                    "channel": "data-team-slack",
                    "url": "https://company.slack.com/channels/data-team",
                    "description": "Data team support channel",
                    "tool": "slack",
                    "scope": "interactive",
                }
            ],
            "team": [
                {
                    "username": "data.engineer@company.com",
                    "name": "Data Engineer",
                    "role": "owner",
                    "description": "Contract owner and maintainer",
                    "dateIn": "2024-01-01",
                }
            ],
            "roles": [
                {
                    "role": "users_read",
                    "access": "read",
                    "description": "Read-only access to users table",
                    "firstLevelApprovers": "Data Team Lead",
                    "secondLevelApprovers": "Data Engineering Manager",
                }
            ],
            "slaProperties": [
                {
                    "property": "latency",
                    "value": 4,
                    "valueExt": 6,
                    "unit": "h",
                    "element": "users.updated_at",
                    "driver": "operational",
                },
                {
                    "property": "frequency",
                    "value": 1,
                    "unit": "d",
                    "element": "users",
                    "driver": "analytics",
                },
            ],
            "authoritativeDefinitions": [
                {
                    "url": "https://docs.company.com/users-contract",
                    "type": "businessDefinition",
                }
            ],
            "customProperties": [
                {
                    "property": "dataRetention",
                    "value": "7 years",
                },
                {
                    "property": "encryptionRequired",
                    "value": True,
                },
            ],
        }

    def test_enhanced_excel_generation(self, enhanced_odcs_contract, temp_dir):
        """Test Excel generation with all enhanced features."""
        converter = ODCSToExcelConverter()
        excel_path = temp_dir / "enhanced_test.xlsx"

        # Generate Excel file
        converter.generate_from_dict(enhanced_odcs_contract, excel_path)

        # Verify file was created
        assert excel_path.exists()
        assert excel_path.stat().st_size > 0

        # Load workbook to verify worksheet creation
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)

        # Verify all expected worksheets exist
        expected_sheets = [
            "Basic Information",
            "Tags",
            "Description",
            "Servers",
            "Schema",
            "Schema Properties",
            "Logical Type Options",
            "Quality Rules",
            "Support",
            "Pricing",
            "Team",
            "Roles",
            "SLA Properties",
            "Authoritative Definitions",
            "Custom Properties",
        ]

        actual_sheets = wb.sheetnames
        for sheet in expected_sheets:
            if sheet not in ["Servers", "Pricing"]:  # These might be empty in test data
                assert sheet in actual_sheets, f"Missing worksheet: {sheet}"

        # Verify Schema Properties sheet has data
        if "Schema Properties" in actual_sheets:
            props_sheet = wb["Schema Properties"]
            assert props_sheet.max_row > 1  # Should have header + data rows

        # Verify Logical Type Options sheet has data
        if "Logical Type Options" in actual_sheets:
            options_sheet = wb["Logical Type Options"]
            assert options_sheet.max_row > 1  # Should have header + data rows

        # Verify Quality Rules sheet has data
        if "Quality Rules" in actual_sheets:
            quality_sheet = wb["Quality Rules"]
            assert quality_sheet.max_row > 1  # Should have header + data rows

    def test_enhanced_excel_parsing(self, enhanced_odcs_contract, temp_dir):
        """Test Excel parsing with all enhanced features."""
        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()
        excel_path = temp_dir / "enhanced_test.xlsx"

        # Generate Excel file
        converter.generate_from_dict(enhanced_odcs_contract, excel_path)

        # Parse Excel file back to ODCS
        parsed_data = parser.parse_from_file(excel_path)

        # Verify basic structure
        assert parsed_data["version"] == enhanced_odcs_contract["version"]
        assert parsed_data["id"] == enhanced_odcs_contract["id"]
        assert parsed_data["apiVersion"] == enhanced_odcs_contract["apiVersion"]

        # Verify schema structure
        assert "schema" in parsed_data
        assert len(parsed_data["schema"]) == len(enhanced_odcs_contract["schema"])

        # Verify first schema object
        parsed_obj = parsed_data["schema"][0]
        original_obj = enhanced_odcs_contract["schema"][0]

        assert parsed_obj["name"] == original_obj["name"]
        assert parsed_obj["description"] == original_obj["description"]
        assert parsed_obj["businessName"] == original_obj["businessName"]

        # Verify properties exist
        assert "properties" in parsed_obj
        assert len(parsed_obj["properties"]) > 0

        # Verify enhanced property fields are preserved
        user_id_prop = next(
            (p for p in parsed_obj["properties"] if p["name"] == "user_id"), None
        )
        assert user_id_prop is not None
        assert user_id_prop["primaryKey"] is True
        assert user_id_prop["primaryKeyPosition"] == 1
        assert user_id_prop["criticalDataElement"] is True

        email_prop = next(
            (p for p in parsed_obj["properties"] if p["name"] == "email"), None
        )
        assert email_prop is not None
        assert email_prop["encryptedName"] == "email_encrypted"
        assert "transformLogic" in email_prop
        assert len(email_prop.get("transformSourceObjects", [])) > 0

        # Verify SLA properties with valueExt
        if "slaProperties" in parsed_data:
            latency_sla = next(
                (s for s in parsed_data["slaProperties"] if s["property"] == "latency"),
                None,
            )
            if latency_sla:
                assert latency_sla.get("valueExt") == 6

    def test_round_trip_conversion_with_validation(
        self, enhanced_odcs_contract, temp_dir
    ):
        """Test complete round-trip conversion with Pydantic validation."""
        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()
        excel_path = temp_dir / "roundtrip_test.xlsx"

        # Generate Excel from original data
        converter.generate_from_dict(enhanced_odcs_contract, excel_path)

        # Parse back to dictionary
        parsed_data = parser.parse_from_file(excel_path)

        # Validate with Pydantic model
        contract = ODCSDataContract(**parsed_data)

        # Verify validation passes
        assert contract.version == enhanced_odcs_contract["version"]
        assert contract.id == enhanced_odcs_contract["id"]
        assert contract.name == enhanced_odcs_contract["name"]

        # Verify schema validation
        assert len(contract.schema_) == 1
        schema_obj = contract.schema_[0]
        assert schema_obj.name == "users"
        assert len(schema_obj.properties) > 0

        # Verify primary key validation still works
        user_id_prop = next(
            (p for p in schema_obj.properties if p.name == "user_id"), None
        )
        assert user_id_prop is not None
        assert user_id_prop.primaryKey is True
        assert user_id_prop.primaryKeyPosition == 1

    def test_logical_type_options_round_trip(self, temp_dir):
        """Test logical type options round-trip conversion."""
        contract_with_options = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "logical-type-test",
            "status": "active",
            "schema": [
                {
                    "name": "test_table",
                    "properties": [
                        {
                            "name": "email_field",
                            "logicalType": "string",
                            "logicalTypeOptions": {
                                "format": "email",
                                "minLength": 5,
                                "maxLength": 320,
                                "pattern": r"^[^@]+@[^@]+\.[^@]+$",
                            },
                        },
                        {
                            "name": "number_field",
                            "logicalType": "number",
                            "logicalTypeOptions": {
                                "minimum": 0,
                                "maximum": 1000,
                                "exclusiveMinimum": True,
                                "exclusiveMaximum": False,
                                "multipleOf": 0.01,
                            },
                        },
                        {
                            "name": "array_field",
                            "logicalType": "array",
                            "logicalTypeOptions": {
                                "minItems": 1,
                                "maxItems": 10,
                                "uniqueItems": True,
                            },
                        },
                    ],
                }
            ],
        }

        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()
        excel_path = temp_dir / "logical_type_options_test.xlsx"

        # Round-trip conversion
        converter.generate_from_dict(contract_with_options, excel_path)
        parsed_data = parser.parse_from_file(excel_path)

        # Validate with Pydantic
        contract = ODCSDataContract(**parsed_data)

        # Verify logical type options are preserved
        test_table = contract.schema_[0]

        email_prop = next(p for p in test_table.properties if p.name == "email_field")
        assert email_prop.logicalTypeOptions is not None
        assert email_prop.logicalTypeOptions.format == "email"
        assert email_prop.logicalTypeOptions.minLength == 5
        assert email_prop.logicalTypeOptions.maxLength == 320

        number_prop = next(p for p in test_table.properties if p.name == "number_field")
        assert number_prop.logicalTypeOptions is not None
        assert number_prop.logicalTypeOptions.minimum == 0
        assert number_prop.logicalTypeOptions.maximum == 1000
        assert number_prop.logicalTypeOptions.exclusiveMinimum is True
        assert number_prop.logicalTypeOptions.multipleOf == 0.01

        array_prop = next(p for p in test_table.properties if p.name == "array_field")
        assert array_prop.logicalTypeOptions is not None
        assert array_prop.logicalTypeOptions.minItems == 1
        assert array_prop.logicalTypeOptions.maxItems == 10
        assert array_prop.logicalTypeOptions.uniqueItems is True

    def test_quality_rules_round_trip(self, temp_dir):
        """Test quality rules round-trip conversion."""
        contract_with_quality = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "quality-rules-test",
            "status": "active",
            "schema": [
                {
                    "name": "test_table",
                    "quality": [
                        {
                            "name": "Row Count Check",
                            "rule": "rowCount",
                            "mustBeBetween": [1000, 10000],
                            "unit": "rows",
                            "dimension": "completeness",
                        }
                    ],
                    "properties": [
                        {
                            "name": "test_field",
                            "quality": [
                                {
                                    "name": "Not Null",
                                    "type": "sql",
                                    "query": "SELECT COUNT(*) FROM ${object} WHERE ${property} IS NULL",
                                    "mustBe": 0,
                                    "dimension": "completeness",
                                    "severity": "critical",
                                },
                                {
                                    "name": "Valid Values",
                                    "rule": "validValues",
                                    "validValues": ["A", "B", "C"],
                                    "mustBe": 100,
                                    "unit": "percent",
                                },
                                {
                                    "name": "Custom Rule",
                                    "type": "custom",
                                    "engine": "greatExpectations",
                                    "implementation": "expect_column_values_to_be_in_set",
                                },
                            ],
                        }
                    ],
                }
            ],
        }

        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()
        excel_path = temp_dir / "quality_rules_test.xlsx"

        # Round-trip conversion
        converter.generate_from_dict(contract_with_quality, excel_path)
        parsed_data = parser.parse_from_file(excel_path)

        # Validate with Pydantic
        contract = ODCSDataContract(**parsed_data)

        # Verify quality rules are preserved
        test_table = contract.schema_[0]

        # Object-level quality rules
        assert len(test_table.quality) == 1
        row_count_rule = test_table.quality[0]
        assert row_count_rule.name == "Row Count Check"
        assert row_count_rule.rule == "rowCount"
        assert row_count_rule.mustBeBetween == [1000, 10000]

        # Property-level quality rules
        test_field = test_table.properties[0]
        assert len(test_field.quality) == 3

        not_null_rule = next(r for r in test_field.quality if r.name == "Not Null")
        assert not_null_rule.type == "sql"
        assert not_null_rule.mustBe == 0

        valid_values_rule = next(
            r for r in test_field.quality if r.name == "Valid Values"
        )
        assert valid_values_rule.validValues == ["A", "B", "C"]

        custom_rule = next(r for r in test_field.quality if r.name == "Custom Rule")
        assert custom_rule.engine == "greatExpectations"

    def test_error_handling_with_enhanced_features(self, temp_dir):
        """Test error handling with enhanced features."""
        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()

        # Test with invalid logical type options
        invalid_contract = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "invalid-test",
            "status": "active",
            "schema": [
                {
                    "name": "test_table",
                    "properties": [
                        {
                            "name": "invalid_primary_key",
                            "primaryKey": True,
                            "primaryKeyPosition": -1,  # Invalid position
                        }
                    ],
                }
            ],
        }

        excel_path = temp_dir / "invalid_test.xlsx"

        # Should still generate Excel (validation happens during model creation)
        converter.generate_from_dict(invalid_contract, excel_path)
        assert excel_path.exists()

        # Should parse back successfully (raw data)
        parsed_data = parser.parse_from_file(excel_path)
        assert "schema" in parsed_data

        # But validation should fail
        with pytest.raises(Exception):  # ValidationError
            ODCSDataContract(**parsed_data)

    def test_performance_with_large_enhanced_contract(self, temp_dir):
        """Test performance with large contract containing enhanced features."""
        import time

        # Create a large contract with many properties and quality rules
        large_contract = {
            "version": "1.0.0",
            "kind": "DataContract",
            "apiVersion": "v3.0.2",
            "id": "large-enhanced-test",
            "status": "active",
            "schema": [],
        }

        # Add 10 schema objects with 50 properties each
        for obj_idx in range(10):
            obj = {
                "name": f"table_{obj_idx}",
                "properties": [],
                "quality": [
                    {
                        "name": f"Table {obj_idx} Row Count",
                        "rule": "rowCount",
                        "mustBeBetween": [1000, 100000],
                    }
                ],
            }

            # Add 50 properties per object
            for prop_idx in range(50):
                prop = {
                    "name": f"field_{prop_idx}",
                    "logicalType": "string",
                    "logicalTypeOptions": {"maxLength": 255},
                    "quality": [
                        {
                            "name": f"Field {prop_idx} Not Null",
                            "mustBe": 100,
                            "unit": "percent",
                        }
                    ],
                }
                obj["properties"].append(prop)

            large_contract["schema"].append(obj)

        converter = ODCSToExcelConverter()
        parser = ExcelToODCSParser()
        excel_path = temp_dir / "large_enhanced_test.xlsx"

        # Time the generation
        start_time = time.time()
        converter.generate_from_dict(large_contract, excel_path)
        generation_time = time.time() - start_time

        # Time the parsing
        start_time = time.time()
        parsed_data = parser.parse_from_file(excel_path)
        parsing_time = time.time() - start_time

        # Verify results
        assert excel_path.exists()
        assert len(parsed_data["schema"]) == 10
        assert len(parsed_data["schema"][0]["properties"]) == 50

        # Performance assertions (should be reasonable for large contracts)
        assert generation_time < 30  # Should generate in under 30 seconds
        assert parsing_time < 30  # Should parse in under 30 seconds

        print(f"Large contract generation time: {generation_time:.2f}s")
        print(f"Large contract parsing time: {parsing_time:.2f}s")
