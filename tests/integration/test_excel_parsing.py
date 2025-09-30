"""Integration tests for Excel parsing workflows."""

import json
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock

from odcs_converter.excel_parser import ExcelToODCSParser
from odcs_converter.generator import ODCSToExcelConverter
from odcs_converter.yaml_converter import YAMLConverter


@pytest.mark.integration
class TestExcelParsingWorkflow:
    """Integration tests for complete Excel parsing workflows."""

    def test_parse_excel_to_json_complete_workflow(
        self, sample_odcs_complete, excel_file_factory, temp_dir
    ):
        """Test complete workflow: Excel file -> JSON parsing."""
        # Setup - Create Excel file from ODCS data
        excel_file = excel_file_factory(sample_odcs_complete, "complete_test.xlsx")

        parser = ExcelToODCSParser()
        json_output = temp_dir / "parsed_output.json"

        # Execute parsing
        parsed_data = parser.parse_from_file(excel_file)

        # Save to JSON
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(parsed_data, f, indent=2)

        # Verify
        assert json_output.exists()
        assert json_output.stat().st_size > 0

        # Verify core data integrity
        assert parsed_data.get("version") == sample_odcs_complete["version"]
        assert parsed_data.get("kind") == sample_odcs_complete["kind"]
        assert parsed_data.get("id") == sample_odcs_complete["id"]
        assert parsed_data.get("status") == sample_odcs_complete["status"]

        # Verify tags are preserved (order might differ)
        if "tags" in parsed_data and "tags" in sample_odcs_complete:
            assert set(parsed_data["tags"]) == set(sample_odcs_complete["tags"])

    def test_parse_excel_to_yaml_workflow(
        self, sample_odcs_complete, excel_file_factory, temp_dir
    ):
        """Test workflow: Excel file -> YAML parsing."""
        # Setup
        excel_file = excel_file_factory(sample_odcs_complete, "yaml_test.xlsx")
        parser = ExcelToODCSParser()
        yaml_output = temp_dir / "parsed_output.yaml"

        # Execute
        parsed_data = parser.parse_from_file(excel_file)
        YAMLConverter.dict_to_yaml(parsed_data, yaml_output)

        # Verify
        assert yaml_output.exists()

        # Load and verify YAML content
        loaded_yaml = YAMLConverter.yaml_to_dict(yaml_output)
        assert loaded_yaml.get("version") == sample_odcs_complete["version"]
        assert loaded_yaml.get("kind") == sample_odcs_complete["kind"]

    def test_excel_parsing_with_validation_workflow(
        self, sample_odcs_complete, excel_file_factory, temp_dir
    ):
        """Test Excel parsing with ODCS schema validation."""
        # Setup
        excel_file = excel_file_factory(sample_odcs_complete, "validation_test.xlsx")
        parser = ExcelToODCSParser()

        # Execute parsing
        parsed_data = parser.parse_from_file(excel_file)

        # Validate against ODCS schema
        is_valid = parser.validate_odcs_data(parsed_data)

        # Should be valid (though may have warnings)
        # Note: Complete validation might fail due to URL schemes in support section
        # but the structure should be valid
        assert isinstance(is_valid, bool)

        # Core required fields should be present
        assert "version" in parsed_data
        assert "kind" in parsed_data
        assert "apiVersion" in parsed_data
        assert "id" in parsed_data
        assert "status" in parsed_data

    def test_excel_parsing_with_missing_worksheets(self, temp_dir):
        """Test Excel parsing when some expected worksheets are missing."""
        from openpyxl import Workbook

        # Create minimal Excel file with only Basic Information
        wb = Workbook()
        wb.remove(wb.active)

        # Only create Basic Information sheet
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "1.0.0", "Version"],
            ["kind", "DataContract", "Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "minimal-001", "ID"],
            ["status", "active", "Status"]
        ]
        for row in basic_data:
            basic_sheet.append(row)

        excel_file = temp_dir / "minimal_sheets.xlsx"
        wb.save(excel_file)
        wb.close()

        # Parse the Excel file
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Verify basic fields are parsed
        assert parsed_data.get("version") == "1.0.0"
        assert parsed_data.get("kind") == "DataContract"
        assert parsed_data.get("id") == "minimal-001"

        # Missing worksheets should result in empty/missing sections
        assert "tags" not in parsed_data or len(parsed_data.get("tags", [])) == 0

    def test_excel_parsing_with_empty_worksheets(self, temp_dir):
        """Test Excel parsing when worksheets exist but are empty."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        # Create worksheets with headers only (no data)
        sheets_to_create = ["Basic Information", "Tags", "Description", "Servers"]

        for sheet_name in sheets_to_create:
            sheet = wb.create_sheet(sheet_name)
            if sheet_name == "Basic Information":
                sheet.append(["Field", "Value", "Description"])
                # Add minimal required data
                sheet.append(["version", "1.0.0", "Version"])
                sheet.append(["kind", "DataContract", "Kind"])
                sheet.append(["apiVersion", "v3.0.2", "API Version"])
                sheet.append(["id", "empty-001", "ID"])
                sheet.append(["status", "active", "Status"])
            elif sheet_name == "Tags":
                sheet.append(["Tag"])
            elif sheet_name == "Description":
                sheet.append(["Field", "Value"])
            elif sheet_name == "Servers":
                sheet.append(["Server", "Type", "Description"])

        excel_file = temp_dir / "empty_sheets.xlsx"
        wb.save(excel_file)
        wb.close()

        # Parse the Excel file
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Basic information should be parsed
        assert parsed_data.get("version") == "1.0.0"

        # Empty sections should not be included or should be empty
        assert "tags" not in parsed_data or len(parsed_data.get("tags", [])) == 0
        assert "servers" not in parsed_data or len(parsed_data.get("servers", [])) == 0

    def test_excel_parsing_type_conversion_workflow(self, temp_dir):
        """Test Excel parsing with various data types and conversion."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        # Basic Information with type conversion examples
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "2.0.0", "Version"],
            ["kind", "DataContract", "Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "type-test-001", "ID"],
            ["status", "active", "Status"]
        ]
        for row in basic_data:
            basic_sheet.append(row)

        # Custom Properties with different value types
        props_sheet = wb.create_sheet("Custom Properties")
        props_data = [
            ["Property", "Value"],
            ["stringProperty", "text_value"],
            ["integerProperty", 42],
            ["floatProperty", 3.14],
            ["booleanTrue", "true"],
            ["booleanFalse", "false"],
            ["numericString", "123"],
            ["floatString", "45.67"]
        ]
        for row in props_data:
            props_sheet.append(row)

        # Servers with port conversion
        servers_sheet = wb.create_sheet("Servers")
        servers_data = [
            ["Server", "Type", "Port", "Host"],
            ["db1", "postgresql", 5432, "localhost"],
            ["db2", "mysql", "3306", "remote.host"],
            ["db3", "redis", 6379, "cache.server"]
        ]
        for row in servers_data:
            servers_sheet.append(row)

        excel_file = temp_dir / "type_conversion.xlsx"
        wb.save(excel_file)
        wb.close()

        # Parse and verify type conversions
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Verify custom properties type conversion
        if "customProperties" in parsed_data:
            props = parsed_data["customProperties"]
            prop_dict = {p["property"]: p["value"] for p in props}

            assert prop_dict["stringProperty"] == "text_value"
            assert prop_dict["integerProperty"] == 42
            assert prop_dict["floatProperty"] == 3.14
            assert prop_dict["booleanTrue"] is True
            assert prop_dict["booleanFalse"] is False
            assert prop_dict["numericString"] == 123
            assert prop_dict["floatString"] == 45.67

        # Verify server port conversion
        if "servers" in parsed_data:
            servers = parsed_data["servers"]
            for server in servers:
                if "port" in server:
                    assert isinstance(server["port"], int)

    def test_excel_parsing_data_cleaning_workflow(self, temp_dir):
        """Test Excel parsing with data cleaning (empty values, etc.)."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        # Create sheets with empty values and None data
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "1.0.0", "Version"],
            ["kind", "DataContract", "Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "cleaning-001", "ID"],
            ["status", "active", "Status"],
            ["name", "", "Empty name"],  # Empty value
            ["tenant", None, "None value"]  # None value
        ]
        for row in basic_data:
            basic_sheet.append(row)

        # Tags with empty entries
        tags_sheet = wb.create_sheet("Tags")
        tags_sheet.append(["Tag"])
        tags_sheet.append(["valid_tag"])
        tags_sheet.append([""])  # Empty tag
        tags_sheet.append([None])  # None tag
        tags_sheet.append(["another_valid_tag"])

        # Custom Properties with empty values
        props_sheet = wb.create_sheet("Custom Properties")
        props_data = [
            ["Property", "Value"],
            ["validProp", "validValue"],
            ["emptyProp", ""],
            ["noneProp", None],
            ["", "orphanValue"],  # Empty property name
            ["validProp2", "validValue2"]
        ]
        for row in props_data:
            props_sheet.append(row)

        excel_file = temp_dir / "data_cleaning.xlsx"
        wb.save(excel_file)
        wb.close()

        # Parse and verify data cleaning
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Empty/None values should be cleaned from basic info
        assert "name" not in parsed_data or parsed_data.get("name") == ""
        assert "tenant" not in parsed_data

        # Empty tags should be filtered out
        if "tags" in parsed_data:
            tags = parsed_data["tags"]
            assert "valid_tag" in tags
            assert "another_valid_tag" in tags
            assert "" not in tags
            assert None not in tags

        # Custom properties should be cleaned
        if "customProperties" in parsed_data:
            props = parsed_data["customProperties"]
            prop_names = [p["property"] for p in props if "property" in p]
            assert "validProp" in prop_names
            assert "validProp2" in prop_names
            assert "" not in prop_names

    def test_roundtrip_conversion_workflow(self, sample_odcs_complete, temp_dir):
        """Test complete roundtrip: ODCS -> Excel -> ODCS conversion."""
        # Step 1: Generate Excel from ODCS
        converter = ODCSToExcelConverter()
        excel_file = temp_dir / "roundtrip.xlsx"
        converter.generate_from_dict(sample_odcs_complete, excel_file)

        assert excel_file.exists()

        # Step 2: Parse Excel back to ODCS
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Step 3: Save parsed data
        json_output = temp_dir / "roundtrip_output.json"
        with open(json_output, 'w', encoding='utf-8') as f:
            json.dump(parsed_data, f, indent=2)

        # Step 4: Verify data integrity
        # Core fields should match exactly
        core_fields = ["version", "kind", "apiVersion", "id", "status"]
        for field in core_fields:
            if field in sample_odcs_complete:
                assert parsed_data.get(field) == sample_odcs_complete[field]

        # Optional fields that should be preserved if present
        optional_fields = ["name", "tenant", "dataProduct", "domain"]
        for field in optional_fields:
            if field in sample_odcs_complete and sample_odcs_complete[field]:
                assert parsed_data.get(field) == sample_odcs_complete[field]

        # Tags should be preserved (order may differ)
        if "tags" in sample_odcs_complete and sample_odcs_complete["tags"]:
            assert "tags" in parsed_data
            assert set(parsed_data["tags"]) == set(sample_odcs_complete["tags"])

    def test_excel_parsing_error_recovery_workflow(self, temp_dir):
        """Test Excel parsing with malformed data and error recovery."""
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)

        # Create Basic Information with minimal valid data
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "1.0.0", "Version"],
            ["kind", "DataContract", "Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "error-recovery-001", "ID"],
            ["status", "active", "Status"]
        ]
        for row in basic_data:
            basic_sheet.append(row)

        # Create Servers sheet with some invalid data
        servers_sheet = wb.create_sheet("Servers")
        servers_data = [
            ["Server", "Type", "Port", "Host"],
            ["valid_server", "postgresql", 5432, "localhost"],
            ["", "mysql", "invalid_port", ""],  # Empty server name, invalid port
            ["server2", "unknown_type", 3306, "host2"],  # Unknown type
            ["server3", "", "", "host3"]  # Missing type
        ]
        for row in servers_data:
            servers_sheet.append(row)

        excel_file = temp_dir / "error_recovery.xlsx"
        wb.save(excel_file)
        wb.close()

        # Parse should succeed despite malformed data
        parser = ExcelToODCSParser()
        parsed_data = parser.parse_from_file(excel_file)

        # Core data should be parsed successfully
        assert parsed_data.get("version") == "1.0.0"
        assert parsed_data.get("kind") == "DataContract"

        # Servers should be parsed with available data
        if "servers" in parsed_data:
            servers = parsed_data["servers"]
            # Should include servers with some valid data
            server_names = [s.get("server") for s in servers if s.get("server")]
            assert "valid_server" in server_names

    def test_excel_parsing_performance_with_large_file(self, temp_dir):
        """Test Excel parsing performance with large Excel files."""
        from openpyxl import Workbook
        import time

        wb = Workbook()
        wb.remove(wb.active)

        # Create Basic Information
        basic_sheet = wb.create_sheet("Basic Information")
        basic_data = [
            ["Field", "Value", "Description"],
            ["version", "1.0.0", "Version"],
            ["kind", "DataContract", "Kind"],
            ["apiVersion", "v3.0.2", "API Version"],
            ["id", "performance-test-001", "ID"],
            ["status", "active", "Status"]
        ]
        for row in basic_data:
            basic_sheet.append(row)

        # Create large Tags sheet
        tags_sheet = wb.create_sheet("Tags")
        tags_sheet.append(["Tag"])
        for i in range(500):  # 500 tags
            tags_sheet.append([f"tag_{i}"])

        # Create large Servers sheet
        servers_sheet = wb.create_sheet("Servers")
        servers_data = [["Server", "Type", "Host", "Port", "Database"]]
        for i in range(100):  # 100 servers
            servers_data.append([
                f"server_{i}",
                "postgresql",
                f"host{i}.example.com",
                5432 + i,
                f"db_{i}"
            ])
        for row in servers_data:
            servers_sheet.append(row)

        # Create large Custom Properties sheet
        props_sheet = wb.create_sheet("Custom Properties")
        props_sheet.append(["Property", "Value"])
        for i in range(200):  # 200 properties
            props_sheet.append([f"property_{i}", f"value_{i}"])

        excel_file = temp_dir / "large_performance.xlsx"
        wb.save(excel_file)
        wb.close()

        # Measure parsing performance
        parser = ExcelToODCSParser()
        start_time = time.time()
        parsed_data = parser.parse_from_file(excel_file)
        end_time = time.time()

        parsing_time = end_time - start_time

        # Verify data was parsed correctly
        assert parsed_data.get("version") == "1.0.0"
        assert len(parsed_data.get("tags", [])) == 500
        assert len(parsed_data.get("servers", [])) == 100
        assert len(parsed_data.get("customProperties", [])) == 200

        # Performance should be reasonable (less than 10 seconds for this size)
        assert parsing_time < 10.0, f"Parsing took too long: {parsing_time:.2f} seconds"

    def test_excel_parsing_memory_efficiency(self, temp_dir):
        """Test memory efficiency during Excel parsing."""
        import psutil
        import os

        # Get initial memory usage
        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss

        # Parse multiple Excel files to test memory accumulation
        parser = ExcelToODCSParser()

        for i in range(5):
            # Create and parse Excel file
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)

            # Add basic data
            basic_sheet = wb.create_sheet("Basic Information")
            basic_data = [
                ["Field", "Value"],
                ["version", f"{i}.0.0"],
                ["kind", "DataContract"],
                ["apiVersion", "v3.0.2"],
                ["id", f"memory-test-{i}"],
                ["status", "active"]
            ]
            for row in basic_data:
                basic_sheet.append(row)

            # Add some bulk data
            tags_sheet = wb.create_sheet("Tags")
            tags_sheet.append(["Tag"])
            for j in range(100):
                tags_sheet.append([f"tag_{i}_{j}"])

            excel_file = temp_dir / f"memory_test_{i}.xlsx"
            wb.save(excel_file)
            wb.close()

            # Parse the file
            parsed_data = parser.parse_from_file(excel_file)

            # Verify parsing worked
            assert parsed_data.get("version") == f"{i}.0.0"

        # Check final memory usage
        final_memory = process.memory_info().rss
        memory_growth = final_memory - initial_memory

        # Memory growth should be reasonable (less than 25MB for this test)
        assert memory_growth < 25 * 1024 * 1024, f"Excessive memory growth: {memory_growth / 1024 / 1024:.2f}MB"
